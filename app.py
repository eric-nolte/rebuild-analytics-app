
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from pathlib import Path
import base64
import json
import re
import urllib.request
import zipfile

import altair as alt
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_VERSION = "V18.1"
APP_LAST_UPDATED = "2026-07-08"
METHODOLOGY_VERSION = "2026.07-PowerBI-CrossType-Outlier-v18"
OUTLIER_RULE_VERSION = "Cost Log-IQR + CPT+H Cross-Type Outlier by Machine"
DEALER_RATE_VERSION = "Built-in Expanded Dealer Rates 2016-2026"
SECURITY_CONTROL_VERSION = "Phase 1 Security Controls"
EXPORT_FORMAT_VERSION = "V18.1 Power BI Model Handoff Upgrade"
CONFIDENTIALITY_LABEL = ""
MAX_UPLOAD_MB = 50
DEFAULT_MAX_ROWS_WARNING = 25000
POWERBI_FULL_EXPORT_TABLES = [
    "Fact_Rebuild_Rows", "Fact_Valid_Rebuild_Rows",
    "Fact_Global_RebuildType_AvgCost", "Fact_Region_RebuildType_AvgCost",
    "Fact_Machine_RebuildType_AvgCost", "Fact_MachineGroup_RebuildType_AvgCost",
    "Fact_MachineRegion_RebuildType_AvgCost", "Fact_Machine_Insights",
    "Fact_Machine_Ranking", "Fact_Dealer_Performance", "Fact_Region_Performance",
    "Fact_Exception_Rows", "Fact_Outlier_Rows", "Fact_CrossType_Outliers",
    "Fact_Rate_Coverage", "Fact_Data_Quality", "DataQuality_Summary",
    "Dim_Machine", "Dim_Machine_Group", "Dim_Dealer", "Dim_Rebuild_Type", "Dim_Region", "Dim_Service_Year",
    "Machine_Grouping", "Run_Metadata", "Filter_Summary", "Relationship_Guide", "PowerBI_Relationship_Checks",
    "DAX_Starter", "PowerBI_Instructions", "PowerBI_Report_Layout", "PowerBI_Table_Dictionary",
    "PowerBI_Sheet_Name_Map", "PowerBI_Pipeline_Guide", "PowerBI_Build_Checklist",
    "Export_Mode_Dictionary", "Required_Files", "Testing_Checklist", "Update_Process",
    "Known_Limitations", "Data_Dictionary", "Parameters", "PowerBI_Readiness"
]
DEFAULT_POWERBI_TABLES_STANDARD = POWERBI_FULL_EXPORT_TABLES
DEFAULT_POWERBI_TABLES_DETAILED = POWERBI_FULL_EXPORT_TABLES
DEFAULT_POWERBI_TABLES_FULL_AUDIT = POWERBI_FULL_EXPORT_TABLES
POWERBI_TABLE_OPTIONS = POWERBI_FULL_EXPORT_TABLES
POWERBI_PRESET_MAP = {"Full Detailed": POWERBI_FULL_EXPORT_TABLES}
POWERBI_SHEET_NAME_MAP = {
    "Fact_Machine_RebuildType_AvgCost": "Fact_Machine_RT_AvgCost",
    "Fact_MachineGroup_RebuildType_AvgCost": "Fact_MachGroup_RT_AvgCost",
    "Fact_MachineRegion_RebuildType_AvgCost": "Fact_MachRegion_RT_AvgCost",
    "Fact_Global_RebuildType_AvgCost": "Fact_Global_RT_AvgCost",
    "Fact_Region_RebuildType_AvgCost": "Fact_Region_RT_AvgCost",
}

st.set_page_config(layout="wide", page_title="Rebuild Analytics Platform")

# =====================================================
# CAT-INSPIRED BRANDING / THEME
# =====================================================
def find_logo_file():
    for name in ["cat_logo.png", "caterpillar_logo.png", "CAT_logo.png", "logo.png"]:
        path = Path(name)
        if path.exists():
            return path
    return None


def img_to_base64(path):
    try:
        return base64.b64encode(path.read_bytes()).decode("utf-8")
    except Exception:
        return None


def inject_cat_theme():
    st.markdown(
        """
        <style>
        :root { --cat-yellow:#FFC500; --cat-black:#000000; --cat-charcoal:#1F1F1F; --cat-gray:#7A7A7A; --cat-light:#F4F4F4; --cat-border:#D9D9D9; }
        .stApp { background: var(--cat-light); color: var(--cat-charcoal); }
        .block-container { padding-top: 1rem; padding-left: 2rem; padding-right: 2rem; max-width: 1550px; }
        .cat-header { background: linear-gradient(90deg,#000000 0%,#1F1F1F 72%,#FFC500 72%,#FFC500 100%); padding: 22px 28px; border-radius: 4px; margin-bottom: 20px; box-shadow: 0 4px 14px rgba(0,0,0,.22); border-left: 8px solid #FFC500; display: flex; justify-content: space-between; align-items: center; gap: 24px; }
        .cat-header-title { color: #FFFFFF; font-size: 34px; font-weight: 900; letter-spacing: .6px; line-height: 1.05; margin: 0; }
        .cat-header-subtitle { color: #D9D9D9; font-size: 14px; margin-top: 9px; font-weight: 600; }
        .cat-version-pill { display: inline-block; background: #FFC500; color: #000000; font-weight: 900; padding: 5px 12px; border-radius: 2px; margin-top: 12px; font-size: 12px; }
        .cat-logo-box { background:#FFC500; color:#000000; font-weight: 1000; font-size: 34px; letter-spacing: -1px; padding: 12px 18px; border: 3px solid #000000; min-width: 95px; text-align:center; }
        .cat-logo-img { max-height: 72px; max-width: 210px; background:#FFFFFF; padding:6px; border-radius:3px; border:2px solid #000000; }
        .method-card { background:#FFFFFF; border-left:8px solid #FFC500; padding:15px 18px; margin:12px 0 18px 0; box-shadow:0 2px 8px rgba(0,0,0,.08); border-radius:4px; }
        .method-title { font-size:18px; font-weight:900; color:#000000; margin-bottom:5px; }
        .method-body { font-size:14px; color:#333333; line-height:1.45; }
        h1,h2,h3 { color:#1F1F1F; font-weight:850; }
        h2 { border-left:6px solid #FFC500; padding-left:10px; }
        [data-testid="stMetric"] { background:#FFFFFF; border:1px solid #D9D9D9; border-top:5px solid #FFC500; border-radius:4px; padding:15px 17px; box-shadow:0 2px 8px rgba(0,0,0,.08); }
        [data-testid="stMetricLabel"] { color:#4A4A4A; font-weight:800; }
        [data-testid="stMetricValue"] { color:#000000; font-weight:950; }
        button[data-baseweb="tab"] { background:#FFFFFF; color:#1F1F1F; border-radius:2px 2px 0 0; font-weight:800; border:1px solid #D9D9D9; margin-right:4px; padding:10px 14px; }
        button[data-baseweb="tab"][aria-selected="true"] { background:#000000; color:#FFC500; border-bottom:4px solid #FFC500; }
        .stButton > button { background:#FFC500; color:#000000; border:2px solid #000000; border-radius:3px; font-weight:900; }
        .stButton > button:hover { background:#000000; color:#FFC500; border:2px solid #FFC500; }
        .stDownloadButton > button { background:#000000; color:#FFC500; border:2px solid #FFC500; border-radius:3px; font-weight:900; }
        .stDownloadButton > button:hover { background:#FFC500; color:#000000; border:2px solid #000000; }
        label { font-weight:800 !important; color:#1F1F1F !important; }
        div[data-testid="stAlert"] { border-radius:4px; border-left:6px solid #FFC500; }
        [data-testid="stDataFrame"] { background:#FFFFFF; border:1px solid #D9D9D9; border-radius:4px; box-shadow:0 2px 8px rgba(0,0,0,.06); }
        details { background:#FFFFFF !important; border:1px solid #D9D9D9 !important; border-radius:4px !important; }
        summary { font-weight:900 !important; color:#111111 !important; }
        [data-testid="stVegaLiteChart"] { background:#FFFFFF; border:1px solid #D9D9D9; border-radius:4px; padding:10px; box-shadow:0 2px 8px rgba(0,0,0,.06); }
        div[data-baseweb="select"] > div { border:2px solid #1F1F1F !important; border-radius:3px !important; background:#FFFFFF !important; }
        div[data-baseweb="select"]:focus-within > div { border-color:#FFC500 !important; box-shadow:0 0 0 2px rgba(255,197,0,.35) !important; }
        span[data-baseweb="tag"], div[data-baseweb="tag"] { background:#FFC500 !important; color:#000000 !important; border:1px solid #000000 !important; font-weight:800 !important; }
        span[data-baseweb="tag"] span, div[data-baseweb="tag"] span { color:#000000 !important; font-weight:800 !important; }
        div[role="option"][aria-selected="true"], ul[role="listbox"] li[aria-selected="true"] { background:#FFC500 !important; color:#000000 !important; font-weight:900 !important; }
        div[role="option"]:hover, ul[role="listbox"] li:hover { background:#1F1F1F !important; color:#FFC500 !important; }
        div[data-testid="stCheckbox"] [data-baseweb="checkbox"] > div:first-child { border-color:#000000 !important; }
        div[data-testid="stCheckbox"] input:checked + div { background-color:#FFC500 !important; border-color:#000000 !important; color:#000000 !important; }
        .cat-footer { margin-top:35px; padding:15px; background:#1F1F1F; color:#D9D9D9; font-size:12px; border-top:5px solid #FFC500; border-radius:3px; }
        .cat-footer strong { color:#FFC500; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_header():
    logo = find_logo_file()
    if logo:
        b64 = img_to_base64(logo)
        logo_html = f'<img class="cat-logo-img" src="data:image/png;base64,{b64}" />' if b64 else '<div class="cat-logo-box">CAT</div>'
    else:
        logo_html = '<div class="cat-logo-box">CAT</div>'
    st.markdown(
        f"""
        <div class="cat-header">
          <div>
            <div class="cat-header-title">REBUILD ANALYTICS PLATFORM</div>
            <div class="cat-header-subtitle">USD-normalized rebuild analytics | Dealer-year labor rates | BLS CPI-U inflation | Outlier and exception intelligence</div>
            <div class="cat-version-pill">{APP_VERSION} &nbsp; | &nbsp; Updated {APP_LAST_UPDATED}</div>
          </div>
          <div>{logo_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

inject_cat_theme()
render_header()

# =====================================================
# CONFIG
# =====================================================
CERTIFIED_REBUILD_TYPES = {
    "CER": "COMMERCIAL ENGINE REBUILD",
    "CGR-E": "CERTIFIED GENSET CAPTIVE ENGINE",
    "CHR": "CERTIFIED HYDRAULICS REBUILD",
    "CMCR A-C": "CERTIFIED MACHINE COMPONENT REBUILD AXLE (CENTER)",
    "CMCR A-F": "CERTIFIED MACHINE COMPONENT REBUILD AXLE (FRONT)",
    "CMCR A-R": "CERTIFIED MACHINE COMPONENT REBUILD AXLE (REAR)",
    "CMCR D-C": "CERTIFIED MACHINE COMPONENT REBUILD DIFFERENTIAL (CENTER)",
    "CMCR D-F": "CERTIFIED MACHINE COMPONENT REBUILD DIFFERENTIAL (FRONT)",
    "CMCR D-R": "CERTIFIED MACHINE COMPONENT REBUILD DIFFERENTIAL (REAR)",
    "CMCR-E": "CERTIFIED MACHINE COMPONENT REBUILD - ENGINE",
    "CMCR-F": "CERTIFIED MACHINE COMPONENT REBUILD FINAL DRIVE",
    "CMCR-HFP": "CERTIFIED MACHINE COMPONENT REBUILD HYDRAULIC FAN PUMP AND MOTOR",
    "CMCR-HIP": "CERTIFIED MACHINE COMPONENT REBUILD HYDRAULIC IMPLEMENT PUMP",
    "CMCR-HP": "CERTIFIED MACHINE COMPONENT REBUILD - HYDRAULIC PUMP",
    "CMCR-HSP": "CERTIFIED MACHINE COMPONENT REBUILD HYDRAULIC STEERING PUMP",
    "CMCR-T": "CERTIFIED MACHINE COMPONENT REBUILD - TRANSMISSION AND TORQUE CONVERTOR",
    "CMR": "CERTIFIED MACHINE REBUILD",
    "CMR-U": "CERTIFIED MACHINE REBUILD UPGRADE",
    "CPT+H": "CERTIFIED POWERTRAIN + HYDRAULICS",
    "CPT-O": "CERTIFIED POWERTRAIN",
}
VALID_REGIONS = ["AFRICA", "ANZP/INDONESIA", "CHINA", "EASTERN US", "EUROPE", "INDIA", "JAPAN/ASIA", "LATIN AMERICA", "MIDDLE EAST & EURASIA", "WESTERN US & CANADA"]
REBUILD_TYPE_ALIASES = {"CERTIFIED MACHINE REBUILD": "CMR", "CMR - CERTIFIED MACHINE REBUILD": "CMR", "CERTIFIED MACHINE REBUILD UPGRADE": "CMR-U", "CERTIFIED POWERTRAIN + HYDRAULICS": "CPT+H", "CPT PLUS H": "CPT+H", "CERTIFIED POWERTRAIN": "CPT-O", "CPT 777": "CPT-O"}
METHOD_LOCK_TEXT = """
**Active Methodology Lock**

- **Cost basis:** Adjusted Total Cost USD = PARTS DN USD + Labor Cost USD. PLUS PARTS DN is ignored entirely.
- **Labor:** Labor Cost USD = REBUILD WORK HRS × dealer service-year labor rate converted to USD when applicable.
- **Currency:** Source currency is converted to USD before inflation and analysis.
- **Inflation:** BLS CPI-U All Items, U.S. city average, not seasonally adjusted. Inflation is applied by component.
- **Outliers:** Cost-only log(cost) + IQR by Machine + CCR TYPE.
- **Sample-size rules:** <5 = no removal; 5–7 = 2.0×IQR; 8+ = 1.5×IQR.
- **SMU:** SMU 0 or 1 is a data-quality flag only; SMU is not used as a statistical outlier basis.
- **Cross-type outliers:** CPT+H rows above machine-level valid CMR median × 1.10 are treated as outliers/excluded rows, only when at least 3 valid CMR rows exist.
- **CPT+H reporting:** There is one reported CPT+H value. Statistical cost outliers and cross-type CPT+H outliers are excluded from core averages and exported for audit.
- **V16.1 controls:** Strict Mode and configurable thresholds are optional governance controls; defaults preserve the locked methodology unless manually changed.
"""

if "run_clicked" not in st.session_state:
    st.session_state.run_clicked = False
if "analysis" not in st.session_state:
    st.session_state.analysis = None

# =====================================================
# CHART HELPERS
# =====================================================
CAT_DOMAIN = list(CERTIFIED_REBUILD_TYPES.keys()) + ["CPT+H Adjusted"]
CAT_RANGE = ["#000000", "#FFC500", "#7A7A7A", "#4D4D4D", "#FFCD00", "#2B2B2B", "#A6A6A6", "#6B5B00", "#C49700", "#595959", "#B38F00", "#333333", "#D9A300", "#808080", "#F2B705", "#1F1F1F", "#C0C0C0", "#8A6F00", "#E0B000", "#666666", "#FFC500"]
CAT_COLOR_SCALE = alt.Scale(domain=CAT_DOMAIN, range=CAT_RANGE)

def cat_scatter_chart(data, x, y, color="CCR TYPE", tooltip=None, height=420):
    if data.empty:
        st.info("No chart data available.")
        return
    tooltip = tooltip or [x, y, color]
    chart = (alt.Chart(data).mark_circle(size=78, opacity=0.82, stroke="#000000", strokeWidth=0.35)
        .encode(
            x=alt.X(f"{x}:Q", title=x, axis=alt.Axis(labelColor="#1F1F1F", titleColor="#1F1F1F", gridColor="#E6E6E6")),
            y=alt.Y(f"{y}:Q", title=y, axis=alt.Axis(format="$,.0f", labelColor="#1F1F1F", titleColor="#1F1F1F", gridColor="#E6E6E6")),
            color=alt.Color(f"{color}:N", scale=CAT_COLOR_SCALE, legend=alt.Legend(title=color, orient="right")),
            tooltip=tooltip)
        .properties(height=height).configure_view(stroke="#D9D9D9").configure_axis(labelFontSize=12, titleFontSize=13).configure_legend(labelFontSize=12, titleFontSize=13))
    st.altair_chart(chart, use_container_width=True)

def cat_bar_chart(data, x, y, color=None, tooltip=None, height=360):
    if data.empty:
        st.info("No chart data available.")
        return
    tooltip = tooltip or [x, y]
    color_encoding = alt.value("#FFC500") if color is None else alt.Color(f"{color}:N", scale=CAT_COLOR_SCALE, legend=alt.Legend(title=color))
    chart = (alt.Chart(data).mark_bar(cornerRadiusTopLeft=2, cornerRadiusTopRight=2, stroke="#000000", strokeWidth=0.45)
        .encode(
            x=alt.X(f"{x}:N", title=x, sort="-y", axis=alt.Axis(labelAngle=-35, labelColor="#1F1F1F", titleColor="#1F1F1F")),
            y=alt.Y(f"{y}:Q", title=y, axis=alt.Axis(format="$,.0f", labelColor="#1F1F1F", titleColor="#1F1F1F", gridColor="#E6E6E6")),
            color=color_encoding,
            tooltip=tooltip)
        .properties(height=height).configure_view(stroke="#D9D9D9").configure_axis(labelFontSize=12, titleFontSize=13).configure_legend(labelFontSize=12, titleFontSize=13))
    st.altair_chart(chart, use_container_width=True)


# =====================================================
# CCR DISPLAY / REBUILD-TYPE CHART HELPERS
# =====================================================
CCR_DISPLAY_BASE_ORDER = {
    "CMR": 1,
    "CPT+H": 2,
    "CPT-O": 3,
}
COST_VIEW_SCALE = alt.Scale(domain=["Reported"], range=["#FFC500"])


def ccr_display_label(ccr_type, cost_view=None):
    ccr = str(ccr_type).strip()
    return ccr


def ccr_display_order(label):
    label = str(label).strip()
    if label in CCR_DISPLAY_BASE_ORDER:
        return CCR_DISPLAY_BASE_ORDER[label]
    return 100 + abs(hash(label)) % 1000


def add_ccr_display_columns(df):
    out = df.copy()
    if out.empty:
        return out
    if "Cost View" in out.columns:
        out = out.drop(columns=["Cost View"])
    out["CCR Display"] = out.apply(lambda r: ccr_display_label(r.get("CCR TYPE", "")), axis=1)
    out["CCR Display Order"] = out["CCR Display"].apply(ccr_display_order)
    return out.sort_values([c for c in ["CCR Display Order", "CCR Display"] if c in out.columns])


def cat_rebuild_type_bar_chart(data, value_col="Avg_Cost", height=360):
    """Bar chart that treats CMR, CPT+H Standard, CPT+H Adjusted, and CPT-O as clean sections."""
    if data is None or data.empty:
        st.info("No rebuild-type chart data available.")
        return
    chart_df = add_ccr_display_columns(data).copy()
    chart_df = chart_df.rename(columns={value_col: "Average Cost"})
    sort_values = chart_df.sort_values("CCR Display Order")["CCR Display"].drop_duplicates().tolist()
    chart = (
        alt.Chart(chart_df)
        .mark_bar(cornerRadiusTopLeft=2, cornerRadiusTopRight=2, stroke="#000000", strokeWidth=0.45)
        .encode(
            x=alt.X("CCR Display:N", title="Rebuild Type", sort=sort_values,
                    axis=alt.Axis(labelAngle=-25, labelColor="#1F1F1F", titleColor="#1F1F1F")),
            y=alt.Y("Average Cost:Q", title="Average Cost", axis=alt.Axis(format="$,.0f", labelColor="#1F1F1F", titleColor="#1F1F1F", gridColor="#E6E6E6")),
            color=alt.Color("CCR Display:N", scale=CAT_COLOR_SCALE, legend=None),
            tooltip=["CCR Display", "CCR TYPE", alt.Tooltip("Average Cost:Q", format="$,.0f"), "Count"],
        )
        .properties(height=height)
        .configure_view(stroke="#D9D9D9")
        .configure_axis(labelFontSize=12, titleFontSize=13)
        .configure_legend(labelFontSize=12, titleFontSize=13)
    )
    st.altair_chart(chart, use_container_width=True)

# =====================================================
# GENERAL HELPERS
# =====================================================
def money(x):
    try:
        if pd.isna(x):
            return ""
        return f"${x:,.0f}"
    except Exception:
        return x

def sample_confidence(n):
    try:
        n = int(n)
    except Exception:
        return "Unknown"
    if n < 5:
        return "Limited (<5)"
    if n < 8:
        return "Moderate (5–7)"
    return "Strong (8+)"

def is_numeric_series(series):
    cleaned = series.dropna()
    if cleaned.empty:
        return False
    converted = pd.to_numeric(cleaned, errors="coerce")
    return converted.notna().all()

def display_table(df, currency_cols=None, percent_cols=None, number_cols=None, year_cols=None, smu_cols=None, decimal_cols=None):
    currency_cols = currency_cols or []
    percent_cols = percent_cols or []
    number_cols = number_cols or []
    year_cols = year_cols or []
    smu_cols = smu_cols or []
    decimal_cols = decimal_cols or []
    for col in df.columns:
        upper = str(col).upper()
        if "YEAR" in upper and col not in year_cols:
            year_cols.append(col)
        if "SMU" in upper and col not in smu_cols:
            smu_cols.append(col)
    fmt = {}
    def add_format(col, pattern):
        if col in df.columns and is_numeric_series(df[col]):
            fmt[col] = pattern
    for col in currency_cols:
        add_format(col, "${:,.0f}")
    for col in percent_cols:
        add_format(col, "{:.1f}%")
    for col in number_cols + year_cols + smu_cols:
        add_format(col, "{:,.0f}")
    for col in decimal_cols:
        add_format(col, "{:,.4f}")
    if fmt:
        st.dataframe(df.style.format(fmt), use_container_width=True)
    else:
        st.dataframe(df, use_container_width=True)
    return None


def sanitize_excel_value(value):
    """Prevent Excel formula injection in exported workbooks."""
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped.startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


def sanitize_excel_df(df):
    """Sanitize object/string columns before Excel export."""
    if df is None or not isinstance(df, pd.DataFrame):
        return df
    safe_df = df.copy()
    for col in safe_df.columns:
        if safe_df[col].dtype == "object" or pd.api.types.is_string_dtype(safe_df[col]):
            safe_df[col] = safe_df[col].map(sanitize_excel_value)
    return safe_df


if not hasattr(pd.DataFrame, "_cat_original_to_excel"):
    pd.DataFrame._cat_original_to_excel = pd.DataFrame.to_excel

    def _secure_to_excel(self, *args, **kwargs):
        return pd.DataFrame._cat_original_to_excel(sanitize_excel_df(self), *args, **kwargs)

    pd.DataFrame.to_excel = _secure_to_excel


def file_size_mb(uploaded_file):
    """Return uploaded file size in MB without permanently moving the file pointer."""
    try:
        current_pos = uploaded_file.tell()
        uploaded_file.seek(0, 2)
        size_bytes = uploaded_file.tell()
        uploaded_file.seek(current_pos)
        return size_bytes / (1024 * 1024)
    except Exception:
        return None


def validate_uploaded_file_security(uploaded_file, max_upload_mb=MAX_UPLOAD_MB):
    """Enforce .xlsx-only and max file-size checks."""
    if uploaded_file is None:
        return True
    if not uploaded_file.name.lower().endswith(".xlsx"):
        st.error("Only .xlsx files are allowed for security and compatibility reasons.")
        st.stop()
    uploaded_size_mb = file_size_mb(uploaded_file)
    if uploaded_size_mb is not None and uploaded_size_mb > max_upload_mb:
        st.error(f"Uploaded file is {uploaded_size_mb:.1f} MB. Maximum allowed size is {max_upload_mb} MB.")
        st.stop()
    return True


def reset_app_state():
    """Clear analysis state from Streamlit session."""
    st.session_state.run_clicked = False
    st.session_state.analysis = None


def render_export_acknowledgement(key="export_ack"):
    """Return True only when user acknowledges export authorization."""
    return st.checkbox(
        "I am authorized to download this export.",
        key=key,
    )


def safe_sheet_name(name):
    invalid = ["\\", "/", "*", "[", "]", ":", "?"]
    name = str(name)
    for ch in invalid:
        name = name.replace(ch, "-")
    return name[:31]

def standardize_text(value):
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip().upper())

def normalize_ccr_type(value, fallback_text=""):
    raw = standardize_text(value)
    fallback = standardize_text(fallback_text)
    combined = f"{raw} {fallback}".strip()
    if raw in CERTIFIED_REBUILD_TYPES:
        return raw
    if raw in REBUILD_TYPE_ALIASES:
        return REBUILD_TYPE_ALIASES[raw]
    for code in sorted(CERTIFIED_REBUILD_TYPES.keys(), key=len, reverse=True):
        code_pattern = re.escape(code).replace("\\ ", r"[\s_-]*")
        if re.search(rf"(^|[^A-Z0-9]){code_pattern}([^A-Z0-9]|$)", combined):
            return code
    for alias, code in REBUILD_TYPE_ALIASES.items():
        if alias in combined:
            return code
    return raw if raw else "UNKNOWN"

def normalize_region(value):
    raw = standardize_text(value)
    if not raw:
        return "UNKNOWN"
    aliases = {"WESTERN US AND CANADA": "WESTERN US & CANADA", "WESTERN US & CANADA": "WESTERN US & CANADA", "EASTERN US": "EASTERN US", "LATAM": "LATIN AMERICA", "MIDDLE EAST AND EURASIA": "MIDDLE EAST & EURASIA", "MIDDLE EAST & EURASIA": "MIDDLE EAST & EURASIA", "JAPAN ASIA": "JAPAN/ASIA", "JAPAN/ASIA": "JAPAN/ASIA", "ANZP INDONESIA": "ANZP/INDONESIA", "ANZP/INDONESIA": "ANZP/INDONESIA"}
    if raw in aliases:
        return aliases[raw]
    if raw in VALID_REGIONS:
        return raw
    for region in VALID_REGIONS:
        if region in raw:
            return region
    return raw

def add_vs_cmr(table, type_col="CCR TYPE", avg_col="Avg_Cost"):
    table = table.copy()
    if table.empty or type_col not in table.columns or avg_col not in table.columns:
        return table
    cmr_rows = table[table[type_col] == "CMR"]
    if not cmr_rows.empty:
        cmr_avg = cmr_rows[avg_col].iloc[0]
        if cmr_avg and not pd.isna(cmr_avg):
            table["Vs CMR %"] = ((table[avg_col] - cmr_avg) / cmr_avg) * 100
        else:
            table["Vs CMR %"] = np.nan
    else:
        table["Vs CMR %"] = np.nan
    return table

def add_vs_section_avg(table, avg_col="Avg_Cost"):
    table = table.copy()
    if table.empty or avg_col not in table.columns:
        return table
    section_avg = table[avg_col].mean()
    if section_avg and not pd.isna(section_avg):
        table["Vs Section Avg %"] = ((table[avg_col] - section_avg) / section_avg) * 100
    else:
        table["Vs Section Avg %"] = np.nan
    return table


# =====================================================
# V16 DECISION SUPPORT HELPERS
# =====================================================
def has_both_cmr_cpth(data, type_col="CCR TYPE"):
    if data is None or data.empty or type_col not in data.columns:
        return False
    types = set(data[type_col].dropna().astype(str).str.upper())
    return {"CMR", "CPT+H"}.issubset(types)


def build_dealer_rate_coverage_summary(processed_df):
    if processed_df is None or processed_df.empty or "Rate Source" not in processed_df.columns:
        return pd.DataFrame({"Metric": ["Dealer Rate Coverage Score"], "Value": ["No processed data available"]})
    total = max(len(processed_df), 1)
    direct = int((processed_df["Rate Source"] == "Dealer-Year Rate").sum())
    fallback = total - direct
    missing_dealer = int(processed_df["Dealer Code"].isna().sum()) if "Dealer Code" in processed_df.columns else 0
    exc = int((processed_df["Dealer Rate Exception Flag"] != "").sum()) if "Dealer Rate Exception Flag" in processed_df.columns else 0
    pct = round(direct / total * 100, 1)
    score = "Strong" if pct >= 95 else ("Moderate" if pct >= 80 else "Needs Review")
    return pd.DataFrame({"Metric": ["Dealer Rate Coverage Score", "Dealer-Year Match Rate %", "Dealer-Year Direct Match Rows", "Fallback Rate Rows", "Missing Dealer Code Rows", "Dealer Rate Exception Rows"], "Value": [score, pct, direct, fallback, missing_dealer, exc]})


def build_data_quality_score_summary(processed_df, outlier_count=0, insufficient_count=0):
    if processed_df is None or processed_df.empty:
        return pd.DataFrame({"Metric": ["Data Quality Score"], "Value": ["No processed data available"]})
    total = max(len(processed_df), 1)
    missing_dealer = int(processed_df["Dealer Code"].isna().sum()) if "Dealer Code" in processed_df.columns else 0
    rate_exc = int((processed_df["Dealer Rate Exception Flag"] != "").sum()) if "Dealer Rate Exception Flag" in processed_df.columns else 0
    fx_fallback = int((processed_df["FX Source"] == "Embedded fallback FX table").sum()) if "FX Source" in processed_df.columns else 0
    smu_flags = int(processed_df["Data Quality Exception Flag"].eq("SMU 0 or 1").sum()) if "Data Quality Exception Flag" in processed_df.columns else 0
    other_regions = int((processed_df["Region Display"] == "OTHER").sum()) if "Region Display" in processed_df.columns else 0
    score = 100
    score -= min(25, missing_dealer / total * 100)
    score -= min(25, rate_exc / total * 75)
    score -= min(15, fx_fallback / total * 50)
    score -= min(15, smu_flags / total * 50)
    score -= min(10, other_regions / total * 50)
    score -= min(10, outlier_count / total * 25)
    score -= min(10, insufficient_count / total * 10)
    score = round(max(0, min(100, score)), 1)
    label = "Strong" if score >= 90 else ("Moderate" if score >= 75 else "Needs Review")
    return pd.DataFrame({"Metric": ["Data Quality Score", "Data Quality Label", "Missing Dealer Code Rows", "Dealer Rate Exception Rows", "Fallback FX Rows", "SMU 0 or 1 Rows", "Unknown/Other Region Rows", "Cost Outlier Rows", "Insufficient Sample Rows"], "Value": [score, label, missing_dealer, rate_exc, fx_fallback, smu_flags, other_regions, int(outlier_count), int(insufficient_count)]})


def build_run_readiness_summary(processed_df, rate_coverage_summary, data_quality_score_summary):
    rows = []
    row_count = 0 if processed_df is None else len(processed_df)
    rows.append({"Readiness Check": "Rows After Filters", "Status": "Ready" if row_count > 0 else "Cannot Run", "Details": f"{row_count:,} processed row(s) available."})
    try:
        pct = float(rate_coverage_summary.loc[rate_coverage_summary["Metric"] == "Dealer-Year Match Rate %", "Value"].iloc[0])
        rows.append({"Readiness Check": "Dealer Rate Coverage", "Status": "Ready" if pct >= float(globals().get("dealer_rate_coverage_warning_threshold", 95.0)) else "Needs Review", "Details": f"Dealer-year match rate is {pct}%."})
    except Exception:
        rows.append({"Readiness Check": "Dealer Rate Coverage", "Status": "Needs Review", "Details": "Coverage could not be calculated."})
    try:
        dq = data_quality_score_summary.loc[data_quality_score_summary["Metric"] == "Data Quality Label", "Value"].iloc[0]
        rows.append({"Readiness Check": "Data Quality", "Status": "Ready" if dq == "Strong" else "Needs Review", "Details": f"Data quality label is {dq}."})
    except Exception:
        rows.append({"Readiness Check": "Data Quality", "Status": "Needs Review", "Details": "Score could not be calculated."})
    overall = "Cannot Run" if any(r["Status"] == "Cannot Run" for r in rows) else ("Needs Review" if any(r["Status"] == "Needs Review" for r in rows) else "Ready")
    rows.insert(0, {"Readiness Check": "Overall Run Readiness", "Status": overall, "Details": "Review detailed checks before relying on outputs."})
    return pd.DataFrame(rows)


def build_machine_benchmark_ranking(valid_df, processed_df, cost_col):
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    base = valid_df.groupby("SALES MODEL").agg(Avg_Cost=(cost_col, "mean"), Avg_SMU=("SMU AT REBUILD", "mean"), Valid_Rows=(cost_col, "count"), Cross_Type_Flags=("Cross-Type Exception Flag", lambda x: (x != "").sum())).reset_index()
    full = processed_df.groupby("SALES MODEL").agg(Total_Rows=(cost_col, "count"), Outlier_Rows=("Outlier", "sum"), Dealer_Rate_Exception_Rows=("Dealer Rate Exception Flag", lambda x: (x != "").sum())).reset_index()
    out = base.merge(full, on="SALES MODEL", how="left")
    out["Outlier Rate %"] = (out["Outlier_Rows"] / out["Total_Rows"] * 100).fillna(0)
    out["Dealer Rate Exception Rate %"] = (out["Dealer_Rate_Exception_Rows"] / out["Total_Rows"] * 100).fillna(0)
    overall = out["Avg_Cost"].mean()
    out["Vs Overall Avg %"] = ((out["Avg_Cost"] - overall) / overall * 100) if overall else np.nan
    out["Priority Score"] = (out["Vs Overall Avg %"].clip(lower=0).fillna(0) * .5 + out["Outlier Rate %"] * .25 + out["Dealer Rate Exception Rate %"] * .15 + out["Cross_Type_Flags"] * .10).round(1)
    out["Priority Label"] = np.where(out["Priority Score"] >= 25, "High", np.where(out["Priority Score"] >= 10, "Monitor", "Lower"))
    return out.sort_values(["Priority Score", "Avg_Cost"], ascending=False)


def build_executive_narrative(valid_df, summary_df, cost_col, outlier_count, cross_count, rate_cov, dq_score, show_adjusted):
    lines = []
    if valid_df is None or valid_df.empty:
        return ["No valid rows are available for executive narrative generation."]
    lines.append(f"The analysis includes {len(valid_df):,} valid rows after filters and cost-outlier removal.")
    lines.append(f"The average USD analysis cost is {money(valid_df[cost_col].mean())}.")
    if summary_df is not None and not summary_df.empty and "Avg_Cost" in summary_df.columns:
        top = summary_df.sort_values("Avg_Cost", ascending=False).iloc[0]
        lines.append(f"The highest average rebuild type cost is {top['CCR TYPE']} at {money(top['Avg_Cost'])}.")
    try:
        cov = rate_cov.loc[rate_cov["Metric"] == "Dealer-Year Match Rate %", "Value"].iloc[0]
        lines.append(f"Dealer labor-rate coverage is {cov}% direct dealer-year match.")
    except Exception:
        pass
    try:
        score = dq_score.loc[dq_score["Metric"] == "Data Quality Score", "Value"].iloc[0]
        label = dq_score.loc[dq_score["Metric"] == "Data Quality Label", "Value"].iloc[0]
        lines.append(f"The data quality score is {score}, rated {label}.")
    except Exception:
        pass
    lines.append(f"Cost outliers excluded from core averages: {outlier_count:,}.")
    lines.append(f"CPT+H cross-type review flags: {cross_count:,}.")
    lines.append("Cross-type CPT+H outliers are treated as outliers and excluded from reported CPT+H averages." if show_adjusted else "Cross-type CPT+H outliers are treated as outliers and exported for audit.")
    return lines


def build_cover_sheet(metadata, readiness, dq_score, rate_cov, export_mode="Full Analysis Workbook", export_reason="Not provided", scenario_name_value=None, user_role_view_value="Analyst", strict_mode_value=False):
    """Create a cover sheet without relying on UI globals, so template/export utilities can reuse it safely."""
    def look(table, key, default=""):
        try:
            return table.loc[table.iloc[:, 0] == key, table.columns[1]].iloc[0]
        except Exception:
            return default
    safe_scenario = scenario_name_value if scenario_name_value else "Not provided"
    return pd.DataFrame({
        "Field": ["App Version", "Methodology Version", "Outlier Rule Version", "Dealer Rate Version", "Security Control Version", "Export Format Version", "Export Timestamp", "Export Mode", "Export Reason", "Scenario Name", "User Role View", "Strict Mode", "Overall Run Readiness", "Data Quality Score", "Data Quality Label", "Dealer Rate Coverage Score", "Dealer-Year Match Rate %", "Highlight Legend"],
        "Value": [APP_VERSION, METHODOLOGY_VERSION, OUTLIER_RULE_VERSION, DEALER_RATE_VERSION, SECURITY_CONTROL_VERSION, EXPORT_FORMAT_VERSION, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), export_mode, export_reason, safe_scenario, user_role_view_value, "Yes" if strict_mode_value else "No", look(readiness, "Overall Run Readiness", "Unknown"), look(dq_score, "Data Quality Score", "Unknown"), look(dq_score, "Data Quality Label", "Unknown"), look(rate_cov, "Dealer Rate Coverage Score", "Unknown"), look(rate_cov, "Dealer-Year Match Rate %", "Unknown"), "Red = cost outlier; orange = cross-type exception; yellow = insufficient sample group."]
    })



# =====================================================
# V16.1 GOVERNANCE / FUTURE-PROOFING HELPERS
# =====================================================
def build_known_limitations_table():
    return pd.DataFrame({
        "Known Limitation": [
            "Cross-Type Outlier Dependency",
            "Dealer code dependency",
            "Labor-hour dependency",
            "SMU handling",
            "FX fallback",
            "Built-in dealer rate workbook dependency",
            "Outlier scope",
            "Cross-type scope",
            "Security scope",
        ],
        "Details": [
            "Cross-type outlier detection requires enough valid CMR records in the relevant machine scope.",
            "Dealer labor-rate matching depends on detecting a valid four-character dealer code from the DEALER field.",
            "Rows with missing REBUILD WORK HRS cannot be used for adjusted labor-cost analysis.",
            "SMU values of 0 or 1 are data-quality flags only; SMU is not used for statistical outlier removal.",
            "If live FX retrieval fails, embedded fallback FX rates may be used and flagged.",
            "The expanded built-in dealer-rate workbook must be deployed next to app.py; otherwise emergency generic rates are used.",
            "Cost outliers are detected by Machine + CCR TYPE using log(cost) IQR logic only.",
            "CPT+H cross-type review requires at least 3 valid CMR rows for the same machine.",
            " labeling and acknowledgements support governance but do not replace enterprise authentication/authorization.",
        ],
    })


def build_data_dictionary_table():
    return pd.DataFrame({
        "Field": [
            "Adjusted Total Cost USD",
            "Inflation-Adjusted Adjusted Total Cost USD",
            "PARTS DN USD",
            "Labor Cost USD",
            "Rate Source",
            "Dealer Rate Exception Flag",
            "Outlier",
            "Outlier Reason",
            "Insufficient Sample Group",
            "Cross-Type Exception Flag",
            "CMR Benchmark Cost",
            "Data Quality Exception Flag",
            "Dealer-Year Match Rate %",
            "Data Quality Score",
            "Machine Priority Score",
        ],
        "Definition": [
            "PARTS DN converted to USD plus labor cost converted/calculated in USD.",
            "Adjusted Total Cost USD after component-level CPI inflation adjustment to the selected base year.",
            "PARTS DN multiplied by the source-currency FX-to-USD rate.",
            "REBUILD WORK HRS multiplied by the dealer service-year labor rate in USD.",
            "Indicates whether a direct dealer-year rate or fallback rate was used.",
            "Flags rows where the dealer-year labor-rate match was missing or fallback logic was required.",
            "Boolean flag identifying rows excluded from average calculations by cost outlier logic.",
            "Text explanation for statistical cost outlier or insufficient-sample handling.",
            "TRUE when Machine + CCR TYPE has fewer than 5 rows; no statistical cost outlier removal is applied.",
            "Flags CPT+H rows above machine-level valid CMR median × configured threshold when enough CMR rows exist.",
            "Machine-level valid non-outlier CMR median cost used for CPT+H review.",
            "Flags data-quality issues such as SMU 0 or 1.",
            "Percent of processed rows that used a direct dealer-year labor-rate match.",
            "Composite score summarizing major data-quality risks for the run.",
            "Decision-support score ranking machines by cost position, outliers, dealer-rate exceptions, and cross-type flags.",
        ],
    })



# =====================================================
# V16.3 POWER BI READINESS / MACHINE GROUPING HELPERS
# =====================================================
def infer_machine_grouping_value(model):
    """Infer a simple built-in machine group from SALES MODEL."""
    text = str(model).strip().upper() if pd.notna(model) else "UNKNOWN"
    if not text or text == "NAN":
        return {"Machine Group": "UNKNOWN", "Machine Family": "UNKNOWN", "Machine Category": "Auto", "Machine Group Source": "Built-in", "Machine Group Notes": "Missing model"}
    # Common examples: 777F -> 777 Series, D11T -> D11 Series, 988K -> 988 Series.
    d_match = re.match(r"^(D\d+)", text)
    if d_match:
        family = d_match.group(1)
    else:
        n_match = re.search(r"(\d{2,4})", text)
        family = n_match.group(1) if n_match else text
    return {"Machine Group": f"{family} Series", "Machine Family": family, "Machine Category": "Auto", "Machine Group Source": "Built-in", "Machine Group Notes": "Auto-derived from SALES MODEL"}


def normalize_machine_grouping_columns(df):
    temp = df.copy()
    temp.columns = [str(c).strip() for c in temp.columns]
    rename = {}
    for col in temp.columns:
        cu = str(col).strip().upper().replace("_", " ")
        if cu in ["SALES MODEL", "MODEL", "MACHINE", "MACHINE MODEL"]:
            rename[col] = "SALES MODEL"
        elif cu in ["MACHINE GROUP", "GROUP"]:
            rename[col] = "Machine Group"
        elif cu in ["MACHINE FAMILY", "FAMILY"]:
            rename[col] = "Machine Family"
        elif cu in ["MACHINE CATEGORY", "CATEGORY"]:
            rename[col] = "Machine Category"
        elif cu in ["NOTES", "NOTE", "COMMENTS", "COMMENT"]:
            rename[col] = "Notes"
    return temp.rename(columns=rename)


def parse_machine_grouping_file(group_file):
    """Parse custom machine grouping workbook.

    Preferred V17.1 template format:
    - Instructions: user guidance only
    - Example: sample rows only
    - User Input: actual data to upload

    If a User Input sheet exists, only that sheet is parsed. If not, the parser
    falls back to scanning other sheets for backward compatibility.
    """
    if group_file is None:
        return pd.DataFrame(columns=["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "Notes"])
    group_file.seek(0)
    sheets = pd.read_excel(group_file, sheet_name=None)
    preferred_sheet = None
    for sname in sheets.keys():
        if str(sname).strip().lower() == "user input":
            preferred_sheet = sname
            break
    candidate_items = [(preferred_sheet, sheets[preferred_sheet])] if preferred_sheet else [(sname, sheet) for sname, sheet in sheets.items() if str(sname).strip().lower() not in ["instructions", "example"]]
    frames = []
    for sname, sheet in candidate_items:
        temp = normalize_machine_grouping_columns(sheet)
        if {"SALES MODEL", "Machine Group"}.issubset(set(temp.columns)):
            if "Machine Family" not in temp.columns:
                temp["Machine Family"] = temp["Machine Group"]
            if "Machine Category" not in temp.columns:
                temp["Machine Category"] = "Custom"
            if "Notes" not in temp.columns:
                temp["Notes"] = ""
            temp = temp[["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "Notes"]].copy()
            temp["SALES MODEL"] = temp["SALES MODEL"].astype(str).str.strip().str.upper()
            temp["Machine Group"] = temp["Machine Group"].astype(str).str.strip()
            temp = temp[(temp["SALES MODEL"] != "") & (temp["SALES MODEL"].str.upper() != "NAN") & (temp["Machine Group"] != "")]
            if not temp.empty:
                temp["Machine Group Source"] = f"Custom Upload: {sname}"
                frames.append(temp)
    if not frames:
        return pd.DataFrame(columns=["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "Notes", "Machine Group Source"])
    out = pd.concat(frames, ignore_index=True)
    out = out[out["SALES MODEL"] != ""].drop_duplicates("SALES MODEL", keep="last")
    return out


def create_machine_grouping_template():
    """Return a three-sheet custom machine grouping template workbook."""
    instructions = pd.DataFrame({
        "Section": [
            "Purpose", "How to Use", "How to Use", "How to Use", "How to Use", "How to Use",
            "Required Field", "Required Field", "Optional Field", "Optional Field", "Optional Field",
            "Validation Rule", "Validation Rule", "Validation Rule", "Upload Behavior", "Upload Behavior"
        ],
        "Instruction": [
            "Use this template to upload custom machine grouping logic into the Rebuild Analytics Platform.",
            "Read this Instructions sheet first.",
            "Review the Example sheet to see the correct format.",
            "Enter actual machine grouping data only on the User Input sheet.",
            "Do not rename, delete, or reorder required columns.",
            "Save the workbook as .xlsx and upload it in the Machine Grouping section of the app.",
            "SALES MODEL must match the SALES MODEL value in the rebuild workbook.",
            "Machine Group is the main grouping used for reporting, charts, and Power BI slicers.",
            "Machine Family is optional but recommended for hierarchy reporting.",
            "Machine Category is optional and can describe broad equipment type.",
            "Notes is optional and can document why the grouping was chosen.",
            "SALES MODEL cannot be blank.",
            "Machine Group cannot be blank.",
            "Each SALES MODEL should appear only once. If duplicates exist, the last valid row is used.",
            "Uploaded custom grouping overrides or enriches the app's built-in automatic grouping.",
            "Machines not included in User Input still use the built-in automatic grouping logic.",
        ]
    })
    example = pd.DataFrame({
        "SALES MODEL": ["777F", "793F", "D11T", "988K"],
        "Machine Group": ["777 Series", "793 Series", "D11 Series", "988 Series"],
        "Machine Family": ["777", "793", "D11", "988"],
        "Machine Category": ["Truck", "Truck", "Track-Type Tractor", "Wheel Loader"],
        "Notes": ["Example only", "Example only", "Example only", "Example only"],
    })
    user_input = pd.DataFrame(columns=["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "Notes"])
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="Instructions", index=False)
        example.to_excel(writer, sheet_name="Example", index=False)
        user_input.to_excel(writer, sheet_name="User Input", index=False)
        apply_excel_brand_formatting(writer.book)
    output.seek(0)
    return output.getvalue()


def apply_machine_grouping(df, group_file=None):
    """Apply built-in machine grouping, overridden by an optional custom grouping file."""
    out = df.copy()
    models = out["SALES MODEL"].dropna().astype(str).str.strip().str.upper().unique() if "SALES MODEL" in out.columns else []
    builtin = []
    for model in models:
        row = {"SALES MODEL": model}
        row.update(infer_machine_grouping_value(model))
        builtin.append(row)
    grouping = pd.DataFrame(builtin)
    custom = parse_machine_grouping_file(group_file)
    if not custom.empty:
        custom = custom.rename(columns={"Notes": "Machine Group Notes"})
        for col in ["Machine Family", "Machine Category", "Machine Group Source", "Machine Group Notes"]:
            if col not in custom.columns:
                custom[col] = "Custom" if col != "Machine Group Notes" else ""
        custom["Machine Group Source"] = custom.get("Machine Group Source", "Custom Upload")
        grouping = grouping.set_index("SALES MODEL")
        custom = custom.set_index("SALES MODEL")
        grouping.update(custom[[c for c in custom.columns if c in grouping.columns]])
        missing = custom.index.difference(grouping.index)
        if len(missing) > 0:
            grouping = pd.concat([grouping, custom.loc[missing]], axis=0)
        grouping = grouping.reset_index()
    out["SALES MODEL"] = out["SALES MODEL"].astype(str).str.strip().str.upper()
    out = out.merge(grouping, on="SALES MODEL", how="left")
    for col, default in [("Machine Group", "UNKNOWN"), ("Machine Family", "UNKNOWN"), ("Machine Category", "Auto"), ("Machine Group Source", "Built-in"), ("Machine Group Notes", "")]:
        if col not in out.columns:
            out[col] = default
        out[col] = out[col].fillna(default)
    return out, grouping


def build_filter_summary_table():
    return pd.DataFrame({
        "Filter": ["Scenario Name", "Machines", "Years", "Base Year", "Rebuild Types", "Regions", "Inflation Applied", "Dealer Rate Source", "Fallback Behavior", "Strict Mode", "Export Mode"],
        "Value": [
            scenario_name if scenario_name else "Not provided",
            machine_input if machine_input else "All",
            f"{start_year}–{end_year}",
            base_year,
            ", ".join(rebuild_filter),
            ", ".join(region_filter),
            "Yes" if apply_inflation else "No",
            "Built-in Expanded Dealer Rates" if use_default else "Uploaded Custom Dealer Rates",
            "Built-in default behavior" if use_default else rate_fallback_behavior,
            "Yes" if strict_mode else "No",
            export_mode,
        ]
    })


def build_powerbi_export_preview(tables, selected_tables=None):
    selected = selected_tables or list(tables.keys())
    rows = []
    for name in selected:
        table = tables.get(name, pd.DataFrame())
        if table is None or not isinstance(table, pd.DataFrame):
            table = pd.DataFrame()
        cols = [str(c) for c in table.columns]
        blank_cols = sum(1 for c in cols if c.strip() == "" or c.startswith("Unnamed"))
        dup_cols = len(cols) - len(set(cols))
        run_id = "Run ID" in cols or name in ["Data_Dictionary", "Known_Limitations", "Relationship_Guide", "PowerBI_Relationship_Checks", "DAX_Starter", "PowerBI_Instructions", "PowerBI_Report_Layout", "PowerBI_Table_Dictionary", "PowerBI_Sheet_Name_Map", "PowerBI_Pipeline_Guide", "PowerBI_Build_Checklist", "Export_Mode_Dictionary", "Required_Files", "Testing_Checklist", "Update_Process"]
        marker_rows = 0
        if not table.empty:
            first_col = table.columns[0]
            marker_rows = int(table[first_col].astype(str).str.contains("Confidential|Yellow|Removed Label", case=False, na=False).sum())
        status = "Ready"
        notes = []
        if table.empty:
            status = "Needs Review"
            notes.append("Empty table")
        if blank_cols:
            status = "Not Ready"
            notes.append(f"{blank_cols} blank/unnamed column(s)")
        if dup_cols:
            status = "Not Ready"
            notes.append(f"{dup_cols} duplicate column name(s)")
        if marker_rows:
            status = "Not Ready"
            notes.append(f"{marker_rows} marker row(s)")
        if not run_id:
            status = "Needs Review"
            notes.append("Run ID not present")
        rows.append({"Table": name, "Included": name in selected, "Rows": len(table), "Columns": len(cols), "Blank Headers": blank_cols, "Duplicate Headers": dup_cols, "Run ID Present": run_id, "Marker Rows": marker_rows, "Status": status, "Notes": "; ".join(notes) if notes else "OK"})
    return pd.DataFrame(rows)


def build_powerbi_readiness_score(preview_df):
    if preview_df is None or preview_df.empty:
        return pd.DataFrame({"Metric": ["Power BI Readiness"], "Value": ["Not Ready"]})
    not_ready = int((preview_df["Status"] == "Not Ready").sum())
    needs_review = int((preview_df["Status"] == "Needs Review").sum())
    total = max(len(preview_df), 1)
    score = max(0, round(100 - not_ready * 20 - needs_review * 7, 1))
    label = "Ready" if not_ready == 0 and needs_review == 0 else ("Needs Review" if not_ready == 0 else "Not Ready")
    return pd.DataFrame({"Metric": ["Power BI Readiness", "Power BI Readiness Score", "Tables Checked", "Not Ready Tables", "Needs Review Tables"], "Value": [label, score, total, not_ready, needs_review]})


def build_scenario_comparison_table(analysis, run_id=None, scenario_name_value=None):
    df = analysis["df"]
    valid = analysis["valid"]
    cost_col = analysis["cost_col"]
    rate_cov = analysis.get("dealer_rate_coverage_summary", pd.DataFrame())
    dq_score = analysis.get("data_quality_score_summary", pd.DataFrame())
    def lookup(table, metric, default=np.nan):
        try:
            return table.loc[table.iloc[:, 0] == metric, table.columns[1]].iloc[0]
        except Exception:
            return default
    total = len(df)
    outliers = int(df["Outlier"].sum()) if "Outlier" in df.columns else 0
    return pd.DataFrame({
        "Run ID": [run_id if run_id else datetime.now().strftime("RUN_%Y%m%d_%H%M%S")],
        "Scenario Name": [scenario_name_value if scenario_name_value else (scenario_name if scenario_name else "Not provided")],
        "Run Timestamp": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        "App Version": [APP_VERSION],
        "Avg Cost": [valid[cost_col].mean() if not valid.empty else np.nan],
        "Valid Rows": [len(valid)],
        "Total Rows": [total],
        "Outlier Rows": [outliers],
        "Outlier Rate %": [_percent_from_counts(outliers, total)],
        "Cross-Type Flags": [int((valid["Cross-Type Exception Flag"].astype(str).str.strip() != "").sum()) if "Cross-Type Exception Flag" in valid.columns else 0],
        "Dealer Rate Coverage %": [lookup(rate_cov, "Dealer-Year Match Rate %")],
        "Data Quality Score": [lookup(dq_score, "Data Quality Score")],
        "Start Year": [start_year],
        "End Year": [end_year],
        "Base Year": [base_year],
        "Dealer Rate Source": ["Built-in Expanded Dealer Rates" if use_default else "Uploaded Custom Dealer Rates"],
        "Inflation Applied": ["Yes" if apply_inflation else "No"],
    })


def build_powerbi_instructions_table():
    return pd.DataFrame({
        "Step": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
        "Instruction": [
            "Export the Power BI Dataset Export from the app.",
            "Save the export as Rebuild_Analytics_PowerBI_Dataset.xlsx in the approved SharePoint or OneDrive source folder.",
            "In Power BI Desktop, select Get Data > Excel Workbook and choose the fixed source workbook.",
            "Load all full detailed export tables. Do not use formatted review workbooks as Power BI sources.",
            "Use PowerBI_Sheet_Name_Map to reconcile shortened Excel sheet names with logical table names.",
            "Use Relationship_Guide and PowerBI_Relationship_Checks to create and validate model relationships.",
            "Use Dim_Machine_Group for clean machine-group slicers and group-level views.",
            "Use DAX_Starter to create initial measures and apply the recommended formats.",
            "Use PowerBI_Report_Layout and PowerBI_Build_Checklist to build and test report pages.",
            "After replacing the source workbook, refresh Power BI Desktop and validate slicers, totals, and exception pages."
        ],
        "Notes": [
            "The app is now Power BI-first; this export is the primary downstream output.",
            "Keep the file name and folder path stable so the Power BI connection does not break.",
            "If the file is synced locally, Power BI can connect to the local synced path; otherwise use the SharePoint/OneDrive location approved by your organization.",
            "Each exported sheet has headers on row 1 and no title/watermark rows.",
            "Some logical table names exceed Excel's 31-character sheet limit and are shortened only at the worksheet level.",
            "Recommended relationships are active, many-to-one, and single-direction unless the guide marks a relationship optional.",
            "Dim_Machine_Group has one row per Machine Group and avoids many-to-many behavior when slicing by group.",
            "Review cost column references if inflation is disabled or methodology changes.",
            "Build pages in order: Executive, Machine Detail, Region, Dealer, Exceptions, Handoff.",
            "Archive scenario packages separately from the fixed Power BI source workbook."
        ]
    })


def build_powerbi_report_layout_table():
    return pd.DataFrame([
        {"Page": "Executive Overview", "Section": "KPI Strip", "Visual Type": "Card", "Primary Table": "Fact_Rebuild_Rows", "Fields": "Average Cost, Valid Rows, Total Rows, Outlier Rows, Cross-Type Outliers", "Filters": "Run ID, Service Year, Region, Machine Group", "Sort": "Not applicable", "Purpose": "Show core run-level cost and quality metrics."},
        {"Page": "Executive Overview", "Section": "Rebuild Type Cost", "Visual Type": "Clustered column chart", "Primary Table": "Fact_Global_RebuildType_AvgCost", "Fields": "CCR Display, CCR Display Order, Avg_Cost, Count", "Filters": "Run ID, Service Year", "Sort": "CCR Display Order ascending", "Purpose": "Compare reported CMR, CPT+H, and CPT-O costs after outlier exclusions."},
        {"Page": "Executive Overview", "Section": "Machine Group Cost", "Visual Type": "Clustered column chart", "Primary Table": "Fact_MachineGroup_RebuildType_AvgCost", "Fields": "Machine Group, CCR Display, Avg_Cost, Count", "Filters": "Machine Group, Service Year, Region", "Sort": "Avg_Cost descending or CCR Display Order ascending", "Purpose": "Compare rebuild-type costs across machine groups."},
        {"Page": "Machine Detail", "Section": "Machine Selectors", "Visual Type": "Slicer", "Primary Table": "Dim_Machine", "Fields": "Machine Group, SALES MODEL", "Filters": "Not applicable", "Sort": "Machine Group then SALES MODEL", "Purpose": "Select machine scope for detail analysis."},
        {"Page": "Machine Detail", "Section": "Machine Rebuild Type Cost", "Visual Type": "Clustered column chart", "Primary Table": "Fact_Machine_RebuildType_AvgCost", "Fields": "SALES MODEL, CCR Display, CCR Display Order, Avg_Cost", "Filters": "SALES MODEL", "Sort": "CCR Display Order ascending", "Purpose": "Show one reported value for each rebuild type for the selected machine."},
        {"Page": "Machine Detail", "Section": "Machine Region Breakdown", "Visual Type": "Small multiples column chart or matrix", "Primary Table": "Fact_MachineRegion_RebuildType_AvgCost", "Fields": "SALES MODEL, Region, CCR Display, Avg_Cost, Count", "Filters": "SALES MODEL, Region", "Sort": "Region then CCR Display Order", "Purpose": "Show region rebuild-type cost differences that respond to machine slicers."},
        {"Page": "Region Performance", "Section": "Region Rebuild Type View", "Visual Type": "Small multiples column chart", "Primary Table": "Fact_MachineRegion_RebuildType_AvgCost", "Fields": "Region, SALES MODEL, CCR Display, Avg_Cost", "Filters": "Machine Group, SALES MODEL, Region", "Sort": "CCR Display Order ascending", "Purpose": "Analyze region costs by rebuild type while allowing machine selection."},
        {"Page": "Dealer Performance", "Section": "Dealer Cost Ranking", "Visual Type": "Bar chart", "Primary Table": "Fact_Dealer_Performance", "Fields": "DEALER, Dealer Code, SALES MODEL, Avg_Cost, Performance Score", "Filters": "Machine Group, SALES MODEL, Region, Dealer", "Sort": "Avg_Cost descending or Performance Score ascending", "Purpose": "Identify dealers with high cost or review-needed performance."},
        {"Page": "Exceptions & Data Quality", "Section": "Exception Summary", "Visual Type": "Cards and bar chart", "Primary Table": "Fact_Exception_Rows", "Fields": "Exception Type, Exception Detail, Cost", "Filters": "Machine Group, SALES MODEL, Region, CCR TYPE", "Sort": "Exception count descending", "Purpose": "Summarize rows requiring review."},
        {"Page": "Exceptions & Data Quality", "Section": "Cross-Type Outlier Audit", "Visual Type": "Table", "Primary Table": "Fact_CrossType_Outliers", "Fields": "SALES MODEL, Region, Dealer Code, CCR TYPE, CMR Benchmark Cost, Cross-Type Threshold Cost, Outlier Reason", "Filters": "SALES MODEL, Region, Dealer", "Sort": "Cost descending", "Purpose": "Audit CPT+H rows excluded by the cross-type outlier rule."},
        {"Page": "Handoff / Documentation", "Section": "Data Model", "Visual Type": "Table", "Primary Table": "PowerBI_Table_Dictionary", "Fields": "Table Name, Grain, Purpose, Key Fields", "Filters": "Not applicable", "Sort": "Table Name ascending", "Purpose": "Explain table purpose and grain for future maintainers."},
        {"Page": "Handoff / Documentation", "Section": "Pipeline", "Visual Type": "Table", "Primary Table": "PowerBI_Pipeline_Guide", "Fields": "Step, Action, Expected Result", "Filters": "Not applicable", "Sort": "Step ascending", "Purpose": "Document the SharePoint/OneDrive refresh workflow."},
    ])


def build_powerbi_sheet_name_map():
    rows = []
    for table_name in POWERBI_FULL_EXPORT_TABLES + ["Scenario_Comparison"]:
        excel_name = POWERBI_SHEET_NAME_MAP.get(table_name, safe_sheet_name(table_name))
        rows.append({
            "Logical Table Name": table_name,
            "Excel Sheet Name": excel_name,
            "Logical Name Length": len(table_name),
            "Excel Sheet Name Length": len(excel_name),
            "Name Shortened": excel_name != table_name,
            "Reason": "Excel worksheet names are limited to 31 characters" if excel_name != table_name else "No shortening required",
            "Use In Power BI": "Load this worksheet and rename the query/table to the logical table name if desired",
        })
    return pd.DataFrame(rows).drop_duplicates("Logical Table Name")


def build_powerbi_pipeline_guide():
    return pd.DataFrame([
        {"Step": 1, "Action": "Run analysis in the app", "Expected Result": "Analysis completes and Power BI Dataset Export is available", "Owner": "App user / analyst", "Notes": "Review data quality and exceptions before export."},
        {"Step": 2, "Action": "Download Power BI Dataset Export", "Expected Result": "Clean Excel workbook with one table per worksheet", "Owner": "App user / analyst", "Notes": "Every sheet uses row 1 as the header row."},
        {"Step": 3, "Action": "Rename export to Rebuild_Analytics_PowerBI_Dataset.xlsx", "Expected Result": "Stable file name for Power BI source", "Owner": "App user / analyst", "Notes": "Do not add dates or version suffixes to the source file name."},
        {"Step": 4, "Action": "Replace the existing file in the approved SharePoint/OneDrive source folder", "Expected Result": "Power BI source file path remains unchanged", "Owner": "Data owner", "Notes": "Archive prior versions separately."},
        {"Step": 5, "Action": "Open Power BI Desktop report", "Expected Result": "Report opens with existing data model", "Owner": "Power BI owner", "Notes": "Do not rebuild queries unless table structure changes."},
        {"Step": 6, "Action": "Refresh the Power BI model", "Expected Result": "Tables refresh from the replaced workbook", "Owner": "Power BI owner", "Notes": "Use PowerBI_Relationship_Checks if slicers or relationships behave unexpectedly."},
        {"Step": 7, "Action": "Validate key pages", "Expected Result": "Executive, Machine, Region, Dealer, and Exceptions pages update correctly", "Owner": "Power BI owner", "Notes": "Check machine slicers against Fact_MachineRegion_RebuildType_AvgCost."},
        {"Step": 8, "Action": "Publish or refresh in Power BI Service if licensing/workspace access allows", "Expected Result": "Shared report reflects latest approved dataset", "Owner": "Power BI owner / workspace owner", "Notes": "Sharing may require Pro/PPU or Premium/Fabric capacity depending on workspace setup."},
    ])


def build_powerbi_build_checklist():
    return pd.DataFrame([
        {"Order": 1, "Task": "Load all Power BI export worksheets", "Status Target": "Complete", "Validation": "All listed sheets load without promoted-header fixes", "Notes": "Headers should already be row 1."},
        {"Order": 2, "Task": "Rename shortened queries if desired", "Status Target": "Optional", "Validation": "Power BI query/table names match logical names in PowerBI_Sheet_Name_Map", "Notes": "Short worksheet names are only for Excel compatibility."},
        {"Order": 3, "Task": "Create/confirm core relationships", "Status Target": "Complete", "Validation": "Use Relationship_Guide and PowerBI_Relationship_Checks", "Notes": "Prefer many-to-one, single direction, active."},
        {"Order": 4, "Task": "Confirm dimension key uniqueness", "Status Target": "Complete", "Validation": "PowerBI_Relationship_Checks shows no duplicate dimension keys", "Notes": "Fix duplicates before creating many-to-many relationships."},
        {"Order": 5, "Task": "Create starter DAX measures", "Status Target": "Complete", "Validation": "Measures from DAX_Starter exist with recommended formatting", "Notes": "Review cost column name in Average Cost measure."},
        {"Order": 6, "Task": "Sort CCR Display by CCR Display Order", "Status Target": "Complete", "Validation": "CMR, CPT+H, CPT-O appear in business order", "Notes": "Apply sort-by-column where supported."},
        {"Order": 7, "Task": "Build Executive Overview page", "Status Target": "Complete", "Validation": "Cards and rebuild-type charts align with app summary", "Notes": "Use reported averages after outlier exclusions."},
        {"Order": 8, "Task": "Build Machine Detail page", "Status Target": "Complete", "Validation": "SALES MODEL slicer filters machine and region breakdown visuals", "Notes": "Use Fact_MachineRegion_RebuildType_AvgCost for region charts."},
        {"Order": 9, "Task": "Build Region Performance page", "Status Target": "Complete", "Validation": "Machine and region slicers affect charts", "Notes": "Use Dim_Machine and Dim_Region slicers."},
        {"Order": 10, "Task": "Build Dealer Performance page", "Status Target": "Complete", "Validation": "Dealer visuals respond to machine/region filters", "Notes": "Use Fact_Dealer_Performance."},
        {"Order": 11, "Task": "Build Exceptions & Data Quality page", "Status Target": "Complete", "Validation": "Outlier and cross-type audit tables reconcile to cards", "Notes": "Use Fact_CrossType_Outliers for CPT+H rule audit."},
        {"Order": 12, "Task": "Test source-file replacement refresh", "Status Target": "Complete", "Validation": "Replacing the fixed workbook and refreshing updates visuals", "Notes": "This validates the Level 2 pipeline."},
    ])


def build_powerbi_dax_starter(cost_col):
    return pd.DataFrame([
        {"Page": "All", "Measure Name": "Average Cost", "DAX Expression": f"Average Cost = AVERAGE(Fact_Rebuild_Rows[{cost_col}])", "Format": "Currency, 0 decimals", "Recommended Visual": "Card / bar chart value", "Business Meaning": "Average analysis cost in the current filter context."},
        {"Page": "All", "Measure Name": "Valid Rows", "DAX Expression": "Valid Rows = CALCULATE(COUNTROWS(Fact_Rebuild_Rows), Fact_Rebuild_Rows[Outlier] = FALSE())", "Format": "Whole number", "Recommended Visual": "Card", "Business Meaning": "Rows included in reported averages."},
        {"Page": "All", "Measure Name": "Total Rows", "DAX Expression": "Total Rows = COUNTROWS(Fact_Rebuild_Rows)", "Format": "Whole number", "Recommended Visual": "Card", "Business Meaning": "Processed rows after app filters."},
        {"Page": "Exceptions & Data Quality", "Measure Name": "Outlier Rows", "DAX Expression": "Outlier Rows = CALCULATE(COUNTROWS(Fact_Rebuild_Rows), Fact_Rebuild_Rows[Outlier] = TRUE())", "Format": "Whole number", "Recommended Visual": "Card", "Business Meaning": "Rows excluded by statistical or cross-type outlier rules."},
        {"Page": "Exceptions & Data Quality", "Measure Name": "Outlier Rate %", "DAX Expression": "Outlier Rate % = DIVIDE([Outlier Rows], [Total Rows])", "Format": "Percentage, 1 decimal", "Recommended Visual": "Card / KPI", "Business Meaning": "Share of processed rows excluded as outliers."},
        {"Page": "Exceptions & Data Quality", "Measure Name": "Statistical Cost Outliers", "DAX Expression": "Statistical Cost Outliers = CALCULATE(COUNTROWS(Fact_Rebuild_Rows), Fact_Rebuild_Rows[Statistical Cost Outlier Flag] = TRUE())", "Format": "Whole number", "Recommended Visual": "Card", "Business Meaning": "Rows excluded by log-cost IQR methodology."},
        {"Page": "Exceptions & Data Quality", "Measure Name": "Cross-Type Outliers", "DAX Expression": "Cross-Type Outliers = CALCULATE(COUNTROWS(Fact_Rebuild_Rows), Fact_Rebuild_Rows[Cross-Type Outlier Flag] = TRUE())", "Format": "Whole number", "Recommended Visual": "Card", "Business Meaning": "CPT+H rows excluded by machine-level CMR benchmark rule."},
        {"Page": "Dealer Performance", "Measure Name": "Dealer Rate Exceptions", "DAX Expression": "Dealer Rate Exceptions = CALCULATE(COUNTROWS(Fact_Rebuild_Rows), Fact_Rebuild_Rows[Dealer Rate Exception Flag] <> \"\")", "Format": "Whole number", "Recommended Visual": "Card / table", "Business Meaning": "Rows using fallback or missing dealer-year rate logic."},
        {"Page": "Machine Detail", "Measure Name": "Average SMU", "DAX Expression": "Average SMU = AVERAGE(Fact_Rebuild_Rows[SMU AT REBUILD])", "Format": "Whole number", "Recommended Visual": "Card / scatter plot", "Business Meaning": "Average service meter units in selected context."},
        {"Page": "Region Performance", "Measure Name": "Selected Machine Region Avg Cost", "DAX Expression": "Selected Machine Region Avg Cost = AVERAGE(Fact_MachRegion_RT_AvgCost[Avg_Cost])", "Format": "Currency, 0 decimals", "Recommended Visual": "Small multiples / matrix", "Business Meaning": "Average rebuild-type cost by selected machine and region using the machine-region summary table."},
    ])

def build_powerbi_table_dictionary():
    """Business-friendly dictionary for every Power BI export table, including grain."""
    return pd.DataFrame([
        {"Table Name": "Fact_Rebuild_Rows", "Excel Sheet Name": "Fact_Rebuild_Rows", "Grain": "One row per processed rebuild source record", "Purpose": "Primary processed row table with costs, rates, FX, CPI, data-quality flags, and outlier flags.", "Typical Power BI Use": "Main row-level fact table for measures, slicers, drill-through, and exception context.", "Key Fields": "Run ID, SALES MODEL, Machine Group, Dealer Code, Region, CCR TYPE, cost columns, Outlier, Outlier Rule Type"},
        {"Table Name": "Fact_Valid_Rebuild_Rows", "Excel Sheet Name": "Fact_Valid_Rebuild_Rows", "Grain": "One row per non-outlier rebuild record", "Purpose": "Rows that remain after statistical and cross-type outlier exclusions.", "Typical Power BI Use": "Use when visuals should match app-reported averages exactly.", "Key Fields": "Run ID, SALES MODEL, Region, CCR TYPE, cost columns, SMU AT REBUILD"},
        {"Table Name": "Fact_Global_RebuildType_AvgCost", "Excel Sheet Name": POWERBI_SHEET_NAME_MAP.get("Fact_Global_RebuildType_AvgCost", "Fact_Global_RebuildType_AvgCost"), "Grain": "One row per Run ID + CCR TYPE", "Purpose": "Global average cost by rebuild type after exclusions.", "Typical Power BI Use": "Executive rebuild-type average cost charts.", "Key Fields": "CCR Display, CCR Display Order, Avg_Cost, Count, Total Outliers Excluded"},
        {"Table Name": "Fact_Region_RebuildType_AvgCost", "Excel Sheet Name": POWERBI_SHEET_NAME_MAP.get("Fact_Region_RebuildType_AvgCost", "Fact_Region_RebuildType_AvgCost"), "Grain": "One row per Run ID + Region + CCR TYPE", "Purpose": "Regional average cost by rebuild type after exclusions.", "Typical Power BI Use": "Region rebuild-type comparison; use machine-region table when machine slicers are required.", "Key Fields": "Region, CCR Display, Avg_Cost, Count"},
        {"Table Name": "Fact_Machine_RebuildType_AvgCost", "Excel Sheet Name": POWERBI_SHEET_NAME_MAP.get("Fact_Machine_RebuildType_AvgCost", "Fact_Machine_RebuildType_AvgCost"), "Grain": "One row per Run ID + SALES MODEL + CCR TYPE", "Purpose": "Machine average cost by rebuild type after exclusions.", "Typical Power BI Use": "Machine detail rebuild-type charts.", "Key Fields": "SALES MODEL, CCR Display, Avg_Cost, Vs Machine CMR %, Count"},
        {"Table Name": "Fact_MachineGroup_RebuildType_AvgCost", "Excel Sheet Name": POWERBI_SHEET_NAME_MAP.get("Fact_MachineGroup_RebuildType_AvgCost", "Fact_MachineGroup_RebuildType_AvgCost"), "Grain": "One row per Run ID + Machine Group + CCR TYPE", "Purpose": "Machine group/family average cost by rebuild type.", "Typical Power BI Use": "Machine group comparison charts.", "Key Fields": "Machine Group, CCR Display, Avg_Cost, Count"},
        {"Table Name": "Fact_MachineRegion_RebuildType_AvgCost", "Excel Sheet Name": POWERBI_SHEET_NAME_MAP.get("Fact_MachineRegion_RebuildType_AvgCost", "Fact_MachineRegion_RebuildType_AvgCost"), "Grain": "One row per Run ID + SALES MODEL + Region + CCR TYPE", "Purpose": "Machine-region average cost by rebuild type so region charts respond to machine slicers.", "Typical Power BI Use": "Machine detail region charts and region drill-through pages.", "Key Fields": "SALES MODEL, Region, CCR Display, Avg_Cost, Count"},
        {"Table Name": "Fact_Machine_Insights", "Excel Sheet Name": "Fact_Machine_Insights", "Grain": "One row per generated machine insight", "Purpose": "Machine-level narrative insights in table form.", "Typical Power BI Use": "Insight cards/tables and drill-through notes.", "Key Fields": "SALES MODEL, Insight Category, Insight Text, Metric Name, Priority"},
        {"Table Name": "Fact_Machine_Ranking", "Excel Sheet Name": "Fact_Machine_Ranking", "Grain": "One row per Run ID + SALES MODEL", "Purpose": "Machine ranking and priority scoring output.", "Typical Power BI Use": "Executive ranking tables and priority-score visuals.", "Key Fields": "SALES MODEL, Avg_Cost, Priority Score, Priority Label"},
        {"Table Name": "Fact_Dealer_Performance", "Excel Sheet Name": "Fact_Dealer_Performance", "Grain": "One row per Run ID + SALES MODEL + Dealer Code + Region", "Purpose": "Dealer-level performance scoring and cost position metrics.", "Typical Power BI Use": "Dealer scorecards and review-needed dealer lists.", "Key Fields": "Dealer Code, DEALER, SALES MODEL, Avg_Cost, Performance Score"},
        {"Table Name": "Fact_Region_Performance", "Excel Sheet Name": "Fact_Region_Performance", "Grain": "One row per Run ID + SALES MODEL + Region", "Purpose": "Region-level performance summary by machine.", "Typical Power BI Use": "Region scorecards and machine-filtered region pages.", "Key Fields": "SALES MODEL, Region, Avg_Cost, Count, Outlier Rate %"},
        {"Table Name": "Fact_Exception_Rows", "Excel Sheet Name": "Fact_Exception_Rows", "Grain": "One row per exception event; one source record may appear multiple times", "Purpose": "Unified exception table for outliers, data quality, dealer rate exceptions, and FX fallback.", "Typical Power BI Use": "Exception review page and drill-through table.", "Key Fields": "Exception Type, Exception Detail, SALES MODEL, Region, Cost"},
        {"Table Name": "Fact_Outlier_Rows", "Excel Sheet Name": "Fact_Outlier_Rows", "Grain": "One row per row excluded as an outlier", "Purpose": "Audit table for all excluded statistical and cross-type outlier rows.", "Typical Power BI Use": "Outlier audit page.", "Key Fields": "SALES MODEL, CCR TYPE, Outlier Rule Type, Outlier Reason"},
        {"Table Name": "Fact_CrossType_Outliers", "Excel Sheet Name": "Fact_CrossType_Outliers", "Grain": "One row per CPT+H row excluded by cross-type rule", "Purpose": "Audit table for CPT+H rows above the machine-level CMR benchmark.", "Typical Power BI Use": "Cross-type outlier audit page.", "Key Fields": "SALES MODEL, CCR TYPE, CMR Benchmark Cost, Cross-Type Threshold Cost"},
        {"Table Name": "Dim_Machine", "Excel Sheet Name": "Dim_Machine", "Grain": "One row per machine model", "Purpose": "Machine model and grouping lookup.", "Typical Power BI Use": "Machine and model slicers.", "Key Fields": "SALES MODEL, Machine Group, Machine Family, Machine Category"},
        {"Table Name": "Dim_Machine_Group", "Excel Sheet Name": "Dim_Machine_Group", "Grain": "One row per Machine Group", "Purpose": "Clean machine-group lookup to avoid non-unique Machine Group relationships.", "Typical Power BI Use": "Machine group slicers and group-level summary visuals.", "Key Fields": "Machine Group, Machine Family, Machine Category, Machine Count"},
        {"Table Name": "Dim_Dealer", "Excel Sheet Name": "Dim_Dealer", "Grain": "One row per dealer/region row", "Purpose": "Dealer and region lookup table.", "Typical Power BI Use": "Dealer slicers and dealer pages.", "Key Fields": "Dealer Code, DEALER, Region"},
        {"Table Name": "Dim_Rebuild_Type", "Excel Sheet Name": "Dim_Rebuild_Type", "Grain": "One row per rebuild type", "Purpose": "Certified rebuild type reference table.", "Typical Power BI Use": "Rebuild type slicers and descriptions.", "Key Fields": "CCR TYPE, Rebuild Description"},
        {"Table Name": "Dim_Region", "Excel Sheet Name": "Dim_Region", "Grain": "One row per configured region", "Purpose": "Configured region reference table.", "Typical Power BI Use": "Region slicers and region charts.", "Key Fields": "Region, Configured Region Flag"},
        {"Table Name": "Dim_Service_Year", "Excel Sheet Name": "Dim_Service_Year", "Grain": "One row per service year", "Purpose": "Service year dimension.", "Typical Power BI Use": "Year slicers and trend visuals.", "Key Fields": "Service Year, Year, Year Label"},
        {"Table Name": "PowerBI_Relationship_Checks", "Excel Sheet Name": "PowerBI_Relationship_Checks", "Grain": "One row per suggested relationship", "Purpose": "Pre-check relationship readiness, unmatched keys, and duplicate dimension keys.", "Typical Power BI Use": "Troubleshooting model relationships.", "Key Fields": "Relationship, Status, Recommendation"},
        {"Table Name": "PowerBI_Sheet_Name_Map", "Excel Sheet Name": "PowerBI_Sheet_Name_Map", "Grain": "One row per exported logical table", "Purpose": "Map logical table names to Excel worksheet names.", "Typical Power BI Use": "Rename imported queries/tables consistently.", "Key Fields": "Logical Table Name, Excel Sheet Name, Name Shortened"},
        {"Table Name": "PowerBI_Pipeline_Guide", "Excel Sheet Name": "PowerBI_Pipeline_Guide", "Grain": "One row per pipeline step", "Purpose": "Document the SharePoint/OneDrive Level 2 Power BI source-file workflow.", "Typical Power BI Use": "Handoff and refresh process documentation.", "Key Fields": "Step, Action, Expected Result, Owner"},
        {"Table Name": "PowerBI_Build_Checklist", "Excel Sheet Name": "PowerBI_Build_Checklist", "Grain": "One row per report build/test task", "Purpose": "Checklist for building and validating the Power BI report.", "Typical Power BI Use": "Report build QA and handoff.", "Key Fields": "Order, Task, Validation, Notes"},
        {"Table Name": "DAX_Starter", "Excel Sheet Name": "DAX_Starter", "Grain": "One row per starter measure", "Purpose": "Starter DAX measures with page, format, visual, and business meaning.", "Typical Power BI Use": "Create consistent initial measures.", "Key Fields": "Page, Measure Name, DAX Expression, Format"},
        {"Table Name": "PowerBI_Report_Layout", "Excel Sheet Name": "PowerBI_Report_Layout", "Grain": "One row per recommended visual/section", "Purpose": "Detailed page-layout template for Power BI report building.", "Typical Power BI Use": "Report design blueprint.", "Key Fields": "Page, Section, Visual Type, Primary Table, Fields"},
    ])

def build_export_mode_dictionary():
    return pd.DataFrame([
        {"Export Mode": "Full Analysis Workbook", "Best For": "Human review and detailed Excel audit", "Notes": "Includes formatted sheets, tables, exceptions, references, and per-machine tabs."},
        {"Export Mode": "Summary Only", "Best For": "Fast manager review", "Notes": "Contains summary, machine ranking, key metadata, and support sheets without large raw data tabs."},
        {"Export Mode": "Exceptions Only", "Best For": "Outlier, data-quality, and cross-type review", "Notes": "Use when the user only needs rows requiring attention."},
        {"Export Mode": "Dealer Rate Audit", "Best For": "Validating dealer labor rate coverage", "Notes": "Includes rate validation, rate coverage, rate exceptions, and dealer rates used."},
        {"Export Mode": "Power BI Dataset Export", "Best For": "Power BI Desktop / Power BI semantic model creation", "Notes": "Clean table export with headers on row 1. Do not use formatted review workbook for Power BI imports."},
        {"Export Mode": "Scenario Archive Package", "Best For": "Handoff and scenario preservation", "Notes": "ZIP package with clean Power BI dataset, archive workbook, and README."},
    ])


def build_required_files_checklist():
    return pd.DataFrame([
        {"Item": "app.py", "Required": "Yes", "Purpose": "Main application code."},
        {"Item": "requirements.txt", "Required": "Yes", "Purpose": "Python package list for deployment."},
        {"Item": "expanded_dealer_base_rate_by_year_2016_2026.xlsx", "Required": "Yes for built-in dealer rates", "Purpose": "Expanded built-in dealer labor-rate workbook."},
        {"Item": "cat_logo.png", "Required": "Optional", "Purpose": "Approved logo file if allowed by internal branding rules."},
        {"Item": "Dealer_Rate_Template.xlsx", "Required": "Generated by app", "Purpose": "Template for custom dealer-rate uploads."},
        {"Item": "Machine_Grouping_Template.xlsx", "Required": "Generated by app", "Purpose": "Template for custom machine grouping overrides."},
    ])


def build_testing_checklist():
    return pd.DataFrame([
        {"Test": "App launches", "Expected Result": "Home page loads with no errors."},
        {"Test": "Dealer-rate template download", "Expected Result": "Template downloads and opens with Instructions, Example, and User Input sheets."},
        {"Test": "Machine-grouping template download", "Expected Result": "Template downloads and opens with Instructions, Example, and User Input sheets."},
        {"Test": "Rebuild workbook upload", "Expected Result": "Pre-run validation profile appears."},
        {"Test": "Run Analysis", "Expected Result": "Dashboard, Machine Detail, Dealer, Region, and Insights tabs populate."},
        {"Test": "Machine Detail chart", "Expected Result": "Chart shows one reported CMR, CPT+H, and CPT-O value when present; cross-type CPT+H outliers are excluded and auditable."},
        {"Test": "Machine Detail region breakdown", "Expected Result": "Region breakdown appears once, directly below the machine rebuild-type chart."},
        {"Test": "Power BI export", "Expected Result": "Power BI workbook sheets have headers on row 1 and no marker/watermark rows."},
        {"Test": "Power BI readiness", "Expected Result": "Selected tables show Ready or explain Needs Review / Not Ready."},
        {"Test": "Excel export", "Expected Result": "Workbook exports, outlier rows highlight red, cross-type rows orange, insufficient samples yellow."},
    ])


def build_update_process_table():
    return pd.DataFrame([
        {"Step": 1, "Action": "Back up current app.py", "Reason": "Provides rollback if a new version fails."},
        {"Step": 2, "Action": "Update code in a test branch/file", "Reason": "Avoids breaking the deployed app during edits."},
        {"Step": 3, "Action": "Run syntax check", "Reason": "Catches Python syntax errors before deployment."},
        {"Step": 4, "Action": "Test with a known sample workbook", "Reason": "Confirms methodology outputs are stable."},
        {"Step": 5, "Action": "Test Power BI export headers", "Reason": "Ensures row 1 is the true header row for every selected table."},
        {"Step": 6, "Action": "Test selected-machine export", "Reason": "Confirms machine handoff package still works."},
        {"Step": 7, "Action": "Commit to GitHub and reboot deployment", "Reason": "Makes the update live."},
        {"Step": 8, "Action": "Update changelog / handoff notes", "Reason": "Future users can understand what changed and why."},
    ])

def build_combined_global_ccr_type_summary(valid_df, adjusted_valid_df, cost_col):
    """Global CCR type summary with Standard and Adjusted CPT+H in the same table."""
    rows = []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    standard = valid_df.groupby("CCR TYPE", dropna=False).agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
    ).reset_index()
    standard["Cost View"] = "Standard"
    standard["Cross-Type Outliers Excluded"] = 0
    rows.append(standard)
    if adjusted_valid_df is not None and not adjusted_valid_df.empty and "CPT+H" in set(valid_df["CCR TYPE"].astype(str)):
        adj_cpth = adjusted_valid_df[adjusted_valid_df["CCR TYPE"] == "CPT+H"].copy()
        if not adj_cpth.empty:
            adj = adj_cpth.groupby("CCR TYPE", dropna=False).agg(
                Avg_Cost=(cost_col, "mean"),
                Avg_SMU=("SMU AT REBUILD", "mean"),
                Count=(cost_col, "count"),
            ).reset_index()
            adj["Cost View"] = "Adjusted"
            removed = int(((valid_df["CCR TYPE"] == "CPT+H") & (valid_df["Cross-Type Exception Flag"].astype(str).str.strip() != "")).sum()) if "Cross-Type Exception Flag" in valid_df.columns else 0
            adj["Cross-Type Outliers Excluded"] = removed
            rows.append(adj)
    out = pd.concat(rows, ignore_index=True, sort=False)
    out["Sample Confidence"] = out["Count"].apply(sample_confidence)
    out["Scope"] = "Global"
    cmr = out[(out["CCR TYPE"] == "CMR") & (out["Cost View"] == "Standard")]
    cmr_cost = cmr["Avg_Cost"].iloc[0] if not cmr.empty else np.nan
    out["Vs CMR %"] = ((out["Avg_Cost"] - cmr_cost) / cmr_cost * 100) if pd.notna(cmr_cost) and cmr_cost else np.nan
    out["Global CCR Avg Cost"] = out["Avg_Cost"]
    out["Vs Global CCR Avg %"] = 0.0
    sort_key = {"CMR": 1, "CPT+H": 2, "CPT-O": 3}
    out["_sort"] = out["CCR TYPE"].map(sort_key).fillna(99)
    out["_view_sort"] = out["Cost View"].map({"Standard": 1, "Adjusted": 2}).fillna(9)
    out = out.sort_values(["_sort", "_view_sort", "CCR TYPE"]).drop(columns=["_sort", "_view_sort"])
    out = add_ccr_display_columns(out)
    return out[["Scope", "CCR TYPE", "Cost View", "CCR Display", "CCR Display Order", "Avg_Cost", "Avg_SMU", "Count", "Sample Confidence", "Vs CMR %", "Global CCR Avg Cost", "Vs Global CCR Avg %", "Cross-Type Outliers Excluded"]]


def build_combined_region_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, global_ccr_summary=None):
    """Region + CCR type summary with Standard and Adjusted CPT+H together."""
    rows = []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    standard = valid_df.groupby(["Region", "CCR TYPE"], dropna=False).agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
    ).reset_index()
    standard["Cost View"] = "Standard"
    standard["Cross-Type Outliers Excluded"] = 0
    rows.append(standard)
    if adjusted_valid_df is not None and not adjusted_valid_df.empty:
        adj_cpth = adjusted_valid_df[adjusted_valid_df["CCR TYPE"] == "CPT+H"].copy()
        if not adj_cpth.empty:
            adj = adj_cpth.groupby(["Region", "CCR TYPE"], dropna=False).agg(
                Avg_Cost=(cost_col, "mean"),
                Avg_SMU=("SMU AT REBUILD", "mean"),
                Count=(cost_col, "count"),
            ).reset_index()
            adj["Cost View"] = "Adjusted"
            if "Cross-Type Exception Flag" in valid_df.columns:
                removed = valid_df[(valid_df["CCR TYPE"] == "CPT+H") & (valid_df["Cross-Type Exception Flag"].astype(str).str.strip() != "")].groupby("Region").size().to_dict()
            else:
                removed = {}
            adj["Cross-Type Outliers Excluded"] = adj["Region"].map(removed).fillna(0).astype(int)
            rows.append(adj)
    out = pd.concat(rows, ignore_index=True, sort=False)
    out["Sample Confidence"] = out["Count"].apply(sample_confidence)
    cmr_map = out[(out["CCR TYPE"] == "CMR") & (out["Cost View"] == "Standard")].set_index("Region")["Avg_Cost"].to_dict()
    out["Regional CMR Avg Cost"] = out["Region"].map(cmr_map)
    out["Vs Regional CMR %"] = np.where(out["Regional CMR Avg Cost"].notna() & (out["Regional CMR Avg Cost"] != 0), (out["Avg_Cost"] - out["Regional CMR Avg Cost"]) / out["Regional CMR Avg Cost"] * 100, np.nan)
    if global_ccr_summary is None or global_ccr_summary.empty:
        global_ccr_summary = build_combined_global_ccr_type_summary(valid_df, adjusted_valid_df, cost_col)
    lookup = global_ccr_summary.set_index(["CCR TYPE", "Cost View"])["Avg_Cost"].to_dict() if not global_ccr_summary.empty else {}
    out["Global CCR Avg Cost"] = out.apply(lambda r: lookup.get((r["CCR TYPE"], r["Cost View"]), np.nan), axis=1)
    out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    out = add_ccr_display_columns(out)
    return out[["Region", "CCR TYPE", "Cost View", "CCR Display", "CCR Display Order", "Avg_Cost", "Avg_SMU", "Count", "Sample Confidence", "Regional CMR Avg Cost", "Vs Regional CMR %", "Global CCR Avg Cost", "Vs Global CCR Avg %", "Cross-Type Outliers Excluded"]].sort_values(["Region", "CCR Display Order", "Cost View"])


def build_combined_machine_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, global_ccr_summary=None):
    """Machine + CCR type summary with Standard and Adjusted CPT+H together."""
    rows = []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    group_cols = [c for c in ["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "CCR TYPE"] if c in valid_df.columns]
    standard = valid_df.groupby(group_cols, dropna=False).agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
    ).reset_index()
    standard["Cost View"] = "Standard"
    standard["Cross-Type Outliers Excluded"] = 0
    rows.append(standard)
    if adjusted_valid_df is not None and not adjusted_valid_df.empty:
        adj_cpth = adjusted_valid_df[adjusted_valid_df["CCR TYPE"] == "CPT+H"].copy()
        if not adj_cpth.empty:
            adj_group_cols = [c for c in group_cols if c in adj_cpth.columns]
            adj = adj_cpth.groupby(adj_group_cols, dropna=False).agg(
                Avg_Cost=(cost_col, "mean"),
                Avg_SMU=("SMU AT REBUILD", "mean"),
                Count=(cost_col, "count"),
            ).reset_index()
            adj["Cost View"] = "Adjusted"
            if "Cross-Type Exception Flag" in valid_df.columns:
                removed = valid_df[(valid_df["CCR TYPE"] == "CPT+H") & (valid_df["Cross-Type Exception Flag"].astype(str).str.strip() != "")].groupby("SALES MODEL").size().to_dict()
            else:
                removed = {}
            adj["Cross-Type Outliers Excluded"] = adj["SALES MODEL"].map(removed).fillna(0).astype(int)
            rows.append(adj)
    out = pd.concat(rows, ignore_index=True, sort=False)
    out["Sample Confidence"] = out["Count"].apply(sample_confidence)
    cmr_map = out[(out["CCR TYPE"] == "CMR") & (out["Cost View"] == "Standard")].set_index("SALES MODEL")["Avg_Cost"].to_dict()
    out["Machine CMR Avg Cost"] = out["SALES MODEL"].map(cmr_map)
    out["Vs Machine CMR %"] = np.where(out["Machine CMR Avg Cost"].notna() & (out["Machine CMR Avg Cost"] != 0), (out["Avg_Cost"] - out["Machine CMR Avg Cost"]) / out["Machine CMR Avg Cost"] * 100, np.nan)
    if global_ccr_summary is None or global_ccr_summary.empty:
        global_ccr_summary = build_combined_global_ccr_type_summary(valid_df, adjusted_valid_df, cost_col)
    lookup = global_ccr_summary.set_index(["CCR TYPE", "Cost View"])["Avg_Cost"].to_dict() if not global_ccr_summary.empty else {}
    out["Global CCR Avg Cost"] = out.apply(lambda r: lookup.get((r["CCR TYPE"], r["Cost View"]), np.nan), axis=1)
    out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    out = add_ccr_display_columns(out)
    desired = ["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "CCR TYPE", "Cost View", "CCR Display", "CCR Display Order", "Avg_Cost", "Avg_SMU", "Count", "Sample Confidence", "Machine CMR Avg Cost", "Vs Machine CMR %", "Global CCR Avg Cost", "Vs Global CCR Avg %", "Cross-Type Outliers Excluded"]
    return out[[c for c in desired if c in out.columns]].sort_values(["SALES MODEL", "CCR Display Order", "Cost View"])



def build_combined_machine_group_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, global_ccr_summary=None):
    """Machine Group / Family + CCR type summary focused on rebuild-type average cost, not total average."""
    rows = []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    group_cols = [c for c in ["Machine Group", "Machine Family", "Machine Category", "CCR TYPE"] if c in valid_df.columns]
    if "Machine Group" not in group_cols:
        return pd.DataFrame()
    standard = valid_df.groupby(group_cols, dropna=False).agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
    ).reset_index()
    standard["Cost View"] = "Standard"
    standard["Cross-Type Outliers Excluded"] = 0
    rows.append(standard)
    if adjusted_valid_df is not None and not adjusted_valid_df.empty:
        adj_cpth = adjusted_valid_df[adjusted_valid_df["CCR TYPE"] == "CPT+H"].copy()
        if not adj_cpth.empty:
            adj_group_cols = [c for c in group_cols if c in adj_cpth.columns]
            adj = adj_cpth.groupby(adj_group_cols, dropna=False).agg(
                Avg_Cost=(cost_col, "mean"),
                Avg_SMU=("SMU AT REBUILD", "mean"),
                Count=(cost_col, "count"),
            ).reset_index()
            adj["Cost View"] = "Adjusted"
            if "Cross-Type Exception Flag" in valid_df.columns:
                removed = valid_df[(valid_df["CCR TYPE"] == "CPT+H") & (valid_df["Cross-Type Exception Flag"].astype(str).str.strip() != "")].groupby("Machine Group").size().to_dict()
            else:
                removed = {}
            adj["Cross-Type Outliers Excluded"] = adj["Machine Group"].map(removed).fillna(0).astype(int)
            rows.append(adj)
    out = pd.concat(rows, ignore_index=True, sort=False)
    out["Sample Confidence"] = out["Count"].apply(sample_confidence)
    cmr_map = out[(out["CCR TYPE"] == "CMR") & (out["Cost View"] == "Standard")].set_index("Machine Group")["Avg_Cost"].to_dict()
    out["Group CMR Avg Cost"] = out["Machine Group"].map(cmr_map)
    out["Vs Group CMR %"] = np.where(out["Group CMR Avg Cost"].notna() & (out["Group CMR Avg Cost"] != 0), (out["Avg_Cost"] - out["Group CMR Avg Cost"]) / out["Group CMR Avg Cost"] * 100, np.nan)
    if global_ccr_summary is None or global_ccr_summary.empty:
        global_ccr_summary = build_combined_global_ccr_type_summary(valid_df, adjusted_valid_df, cost_col)
    lookup = global_ccr_summary.set_index(["CCR TYPE", "Cost View"])["Avg_Cost"].to_dict() if not global_ccr_summary.empty else {}
    out["Global CCR Avg Cost"] = out.apply(lambda r: lookup.get((r["CCR TYPE"], r["Cost View"]), np.nan), axis=1)
    out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    out = add_ccr_display_columns(out)
    wanted = ["Machine Group", "Machine Family", "Machine Category", "CCR TYPE", "Cost View", "CCR Display", "CCR Display Order", "Avg_Cost", "Avg_SMU", "Count", "Sample Confidence", "Group CMR Avg Cost", "Vs Group CMR %", "Global CCR Avg Cost", "Vs Global CCR Avg %", "Cross-Type Outliers Excluded"]
    return out[[c for c in wanted if c in out.columns]].sort_values(["Machine Group", "CCR Display Order", "Cost View"])

def build_machine_insights_table(valid_df, processed_df, machine_ccr_summary, cost_col):
    """Generate row-based machine insights for app display and Power BI export."""
    rows = []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame(columns=["SALES MODEL", "Machine Group", "Insight Category", "Insight Text", "Metric Name", "Metric Value", "Priority"])
    for machine, mvalid in valid_df.groupby("SALES MODEL", dropna=False):
        mall = processed_df[processed_df["SALES MODEL"] == machine].copy() if processed_df is not None and not processed_df.empty else pd.DataFrame()
        group = mvalid["Machine Group"].dropna().iloc[0] if "Machine Group" in mvalid.columns and not mvalid["Machine Group"].dropna().empty else "UNKNOWN"
        avg_cost = mvalid[cost_col].mean()
        rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Average Cost", "Insight Text": f"{machine} average analysis cost is {money(avg_cost)} across {len(mvalid):,} valid row(s).", "Metric Name": "Average Cost", "Metric Value": avg_cost, "Priority": "Info"})
        ccr_rows = machine_ccr_summary[(machine_ccr_summary["SALES MODEL"] == machine) & (machine_ccr_summary["Cost View"] == "Standard")].copy() if machine_ccr_summary is not None and not machine_ccr_summary.empty else pd.DataFrame()
        if not ccr_rows.empty:
            top = ccr_rows.sort_values("Avg_Cost", ascending=False).iloc[0]
            rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Highest CCR Type", "Insight Text": f"Highest standard CCR type cost for {machine} is {top['CCR TYPE']} at {money(top['Avg_Cost'])}.", "Metric Name": "Highest CCR Avg Cost", "Metric Value": top["Avg_Cost"], "Priority": "Info"})
        outliers = int(mall["Outlier"].sum()) if not mall.empty and "Outlier" in mall.columns else 0
        if outliers:
            rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Outliers", "Insight Text": f"{machine} has {outliers:,} cost outlier row(s) excluded from core averages.", "Metric Name": "Cost Outliers", "Metric Value": outliers, "Priority": "Review"})
        cross_flags = int((mvalid["Cross-Type Exception Flag"].astype(str).str.strip() != "").sum()) if "Cross-Type Exception Flag" in mvalid.columns else 0
        if cross_flags:
            rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Cross-Type Flags", "Insight Text": f"{machine} has {cross_flags:,} CPT+H cross-type review flag(s).", "Metric Name": "Cross-Type Flags", "Metric Value": cross_flags, "Priority": "Review"})
        std_cpth = machine_ccr_summary[(machine_ccr_summary["SALES MODEL"] == machine) & (machine_ccr_summary["CCR TYPE"] == "CPT+H") & (machine_ccr_summary["Cost View"] == "Standard")] if machine_ccr_summary is not None and not machine_ccr_summary.empty else pd.DataFrame()
        adj_cpth = machine_ccr_summary[(machine_ccr_summary["SALES MODEL"] == machine) & (machine_ccr_summary["CCR TYPE"] == "CPT+H") & (machine_ccr_summary["Cost View"] == "Adjusted")] if machine_ccr_summary is not None and not machine_ccr_summary.empty else pd.DataFrame()
        if not std_cpth.empty and not adj_cpth.empty:
            std_val = std_cpth["Avg_Cost"].iloc[0]
            adj_val = adj_cpth["Avg_Cost"].iloc[0]
            diff_pct = (adj_val - std_val) / std_val * 100 if std_val else np.nan
            removed = int(adj_cpth["Cross-Type Outliers Excluded"].iloc[0]) if "Cross-Type Outliers Excluded" in adj_cpth.columns else 0
            rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Adjusted CPT+H", "Insight Text": f"{machine} standard CPT+H is {money(std_val)}; adjusted CPT+H is {money(adj_val)} after removing {removed:,} flagged row(s), a {diff_pct:.1f}% change.", "Metric Name": "Adjusted CPT+H Difference %", "Metric Value": diff_pct, "Priority": "Info" if removed == 0 else "Review"})
        if "Region" in mvalid.columns:
            reg = mvalid.groupby("Region")[cost_col].mean().sort_values(ascending=False)
            if len(reg) > 0:
                rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Highest Region", "Insight Text": f"Highest average region for {machine} is {reg.index[0]} at {money(reg.iloc[0])}.", "Metric Name": "Highest Region Avg Cost", "Metric Value": reg.iloc[0], "Priority": "Info"})
    return pd.DataFrame(rows)


# =====================================================
# V18 POWER BI-FIRST SUMMARY / OUTLIER HELPERS
# =====================================================
def _build_outlier_counts(processed_df, group_cols):
    cols = list(group_cols) if group_cols else []
    if processed_df is None or processed_df.empty:
        return pd.DataFrame(columns=cols + ["Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
    temp = processed_df.copy()
    for flag in ["Statistical Cost Outlier Flag", "Cross-Type Outlier Flag", "Outlier"]:
        if flag not in temp.columns:
            temp[flag] = False
    if cols:
        out = temp.groupby(cols, dropna=False).agg(
            Statistical_Outliers_Excluded=("Statistical Cost Outlier Flag", "sum"),
            Cross_Type_Outliers_Excluded=("Cross-Type Outlier Flag", "sum"),
            Total_Outliers_Excluded=("Outlier", "sum"),
        ).reset_index()
    else:
        out = pd.DataFrame({
            "Statistical_Outliers_Excluded": [int(temp["Statistical Cost Outlier Flag"].sum())],
            "Cross_Type_Outliers_Excluded": [int(temp["Cross-Type Outlier Flag"].sum())],
            "Total_Outliers_Excluded": [int(temp["Outlier"].sum())],
        })
    return out.rename(columns={"Statistical_Outliers_Excluded": "Statistical Outliers Excluded", "Cross_Type_Outliers_Excluded": "Cross-Type Outliers Excluded", "Total_Outliers_Excluded": "Total Outliers Excluded"})


def build_rebuild_type_avg_summary(valid_df, processed_df, cost_col, group_cols=None, cmr_label="CMR Avg Cost", vs_label="Vs CMR %"):
    group_cols = group_cols or []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame(columns=group_cols + ["CCR TYPE", "CCR Display", "CCR Display Order", "Avg_Cost", "Avg_SMU", "Count", "Sample Confidence", cmr_label, vs_label, "Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
    grouping = group_cols + ["CCR TYPE"]
    out = valid_df.groupby(grouping, dropna=False).agg(Avg_Cost=(cost_col, "mean"), Avg_SMU=("SMU AT REBUILD", "mean"), Count=(cost_col, "count")).reset_index()
    out["Sample Confidence"] = out["Count"].apply(sample_confidence)
    out = add_ccr_display_columns(out)
    if group_cols:
        cmr_map = out[out["CCR TYPE"] == "CMR"].set_index(group_cols)["Avg_Cost"].to_dict()
        out[cmr_label] = out.apply(lambda r: cmr_map.get(tuple(r[c] for c in group_cols), np.nan), axis=1)
    else:
        cmr_rows = out[out["CCR TYPE"] == "CMR"]
        cmr_val = cmr_rows["Avg_Cost"].iloc[0] if not cmr_rows.empty else np.nan
        out[cmr_label] = cmr_val
    out[vs_label] = np.where(out[cmr_label].notna() & (out[cmr_label] != 0), (out["Avg_Cost"] - out[cmr_label]) / out[cmr_label] * 100, np.nan)
    outlier_counts = _build_outlier_counts(processed_df, grouping)
    if not outlier_counts.empty:
        out = out.merge(outlier_counts, on=grouping, how="left")
    for col in ["Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"]:
        if col not in out.columns:
            out[col] = 0
        out[col] = out[col].fillna(0).astype(int)
    return out.sort_values([c for c in group_cols + ["CCR Display Order", "CCR TYPE"] if c in out.columns])


def build_combined_global_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, processed_df=None):
    processed_df = valid_df if processed_df is None else processed_df
    out = build_rebuild_type_avg_summary(valid_df, processed_df, cost_col, group_cols=[], cmr_label="Global CMR Avg Cost", vs_label="Vs CMR %")
    if not out.empty:
        out.insert(0, "Scope", "Global")
        out["Global CCR Avg Cost"] = out["Avg_Cost"]
        out["Vs Global CCR Avg %"] = 0.0
    return out


def build_combined_region_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, global_ccr_summary=None, processed_df=None):
    processed_df = valid_df if processed_df is None else processed_df
    out = build_rebuild_type_avg_summary(valid_df, processed_df, cost_col, group_cols=["Region"], cmr_label="Regional CMR Avg Cost", vs_label="Vs Regional CMR %")
    if global_ccr_summary is not None and not global_ccr_summary.empty:
        lookup = global_ccr_summary.set_index("CCR TYPE")["Avg_Cost"].to_dict()
        out["Global CCR Avg Cost"] = out["CCR TYPE"].map(lookup)
        out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    return out


def build_combined_machine_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, global_ccr_summary=None, processed_df=None):
    processed_df = valid_df if processed_df is None else processed_df
    group_cols = [c for c in ["SALES MODEL", "Machine Group", "Machine Family", "Machine Category"] if c in valid_df.columns]
    out = build_rebuild_type_avg_summary(valid_df, processed_df, cost_col, group_cols=group_cols, cmr_label="Machine CMR Avg Cost", vs_label="Vs Machine CMR %")
    if global_ccr_summary is not None and not global_ccr_summary.empty:
        lookup = global_ccr_summary.set_index("CCR TYPE")["Avg_Cost"].to_dict()
        out["Global CCR Avg Cost"] = out["CCR TYPE"].map(lookup)
        out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    return out


def build_combined_machine_group_ccr_type_summary(valid_df, adjusted_valid_df, cost_col, global_ccr_summary=None, processed_df=None):
    processed_df = valid_df if processed_df is None else processed_df
    group_cols = [c for c in ["Machine Group", "Machine Family", "Machine Category"] if c in valid_df.columns]
    out = build_rebuild_type_avg_summary(valid_df, processed_df, cost_col, group_cols=group_cols, cmr_label="Group CMR Avg Cost", vs_label="Vs Group CMR %")
    if global_ccr_summary is not None and not global_ccr_summary.empty:
        lookup = global_ccr_summary.set_index("CCR TYPE")["Avg_Cost"].to_dict()
        out["Global CCR Avg Cost"] = out["CCR TYPE"].map(lookup)
        out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    return out


def build_machine_region_ccr_type_summary(valid_df, processed_df, cost_col, global_ccr_summary=None):
    group_cols = [c for c in ["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "Region"] if c in valid_df.columns]
    out = build_rebuild_type_avg_summary(valid_df, processed_df, cost_col, group_cols=group_cols, cmr_label="Machine Region CMR Avg Cost", vs_label="Vs Machine Region CMR %")
    if global_ccr_summary is not None and not global_ccr_summary.empty:
        lookup = global_ccr_summary.set_index("CCR TYPE")["Avg_Cost"].to_dict()
        out["Global CCR Avg Cost"] = out["CCR TYPE"].map(lookup)
        out["Vs Global CCR Avg %"] = np.where(out["Global CCR Avg Cost"].notna() & (out["Global CCR Avg Cost"] != 0), (out["Avg_Cost"] - out["Global CCR Avg Cost"]) / out["Global CCR Avg Cost"] * 100, np.nan)
    return out


def build_machine_insights_table(valid_df, processed_df, machine_ccr_summary, cost_col):
    rows = []
    if valid_df is None or valid_df.empty:
        return pd.DataFrame(columns=["SALES MODEL", "Machine Group", "Insight Category", "Insight Text", "Metric Name", "Metric Value", "Priority"])
    for machine, mvalid in valid_df.groupby("SALES MODEL", dropna=False):
        mall = processed_df[processed_df["SALES MODEL"] == machine].copy() if processed_df is not None and not processed_df.empty else pd.DataFrame()
        group = mvalid["Machine Group"].dropna().iloc[0] if "Machine Group" in mvalid.columns and not mvalid["Machine Group"].dropna().empty else "UNKNOWN"
        avg_cost = mvalid[cost_col].mean()
        rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Average Cost", "Insight Text": f"{machine} average analysis cost is {money(avg_cost)} across {len(mvalid):,} valid row(s) after excluding statistical and cross-type outliers.", "Metric Name": "Average Cost", "Metric Value": avg_cost, "Priority": "Info"})
        ccr_rows = machine_ccr_summary[machine_ccr_summary["SALES MODEL"] == machine].copy() if machine_ccr_summary is not None and not machine_ccr_summary.empty else pd.DataFrame()
        if not ccr_rows.empty:
            top = ccr_rows.sort_values("Avg_Cost", ascending=False).iloc[0]
            rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Highest Rebuild Type", "Insight Text": f"Highest reported rebuild type cost for {machine} is {top['CCR TYPE']} at {money(top['Avg_Cost'])}.", "Metric Name": "Highest CCR Avg Cost", "Metric Value": top["Avg_Cost"], "Priority": "Info"})
        stat_outliers = int(mall["Statistical Cost Outlier Flag"].sum()) if not mall.empty and "Statistical Cost Outlier Flag" in mall.columns else 0
        cross_outliers = int(mall["Cross-Type Outlier Flag"].sum()) if not mall.empty and "Cross-Type Outlier Flag" in mall.columns else 0
        if stat_outliers or cross_outliers:
            rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Excluded Outliers", "Insight Text": f"{machine} excludes {stat_outliers:,} statistical cost outlier row(s) and {cross_outliers:,} cross-type CPT+H outlier row(s) from core averages.", "Metric Name": "Total Excluded Outliers", "Metric Value": stat_outliers + cross_outliers, "Priority": "Review" if cross_outliers else "Info"})
        if "Region" in mvalid.columns:
            reg = mvalid.groupby("Region")[cost_col].mean().sort_values(ascending=False)
            if len(reg) > 0:
                rows.append({"SALES MODEL": machine, "Machine Group": group, "Insight Category": "Highest Region", "Insight Text": f"Highest average region for {machine} is {reg.index[0]} at {money(reg.iloc[0])}.", "Metric Name": "Highest Region Avg Cost", "Metric Value": reg.iloc[0], "Priority": "Info"})
    return pd.DataFrame(rows)

def build_parameter_summary_table():
    return pd.DataFrame({
        "Parameter": [
            "Scenario Name",
            "User Role View",
            "Strict Mode Enabled",
            "Maximum Row Warning Threshold",
            "Dealer Rate Coverage Warning Threshold %",
            "Dealer Rate Coverage Strict Threshold %",
            "Data Quality Strict Minimum Score",
            "Cross-Type Threshold Multiplier",
            "Minimum CMR Rows for CPT+H Benchmark",
            "Machine Filter",
            "Start Year",
            "End Year",
            "Base Year",
            "Rebuild Type Filter",
            "Region Filter",
            "Inflation Applied",
            "Dealer Rate Source",
            "Export Mode",
            "Export Reason",
        ],
        "Value": [
            scenario_name if scenario_name else "Not provided",
            user_role_view,
            "Yes" if strict_mode else "No",
            max_rows_warning_threshold,
            dealer_rate_coverage_warning_threshold,
            dealer_rate_coverage_strict_threshold,
            data_quality_strict_min_score,
            cross_type_threshold_multiplier,
            min_cmr_rows_for_benchmark,
            machine_input if machine_input else "All",
            start_year,
            end_year,
            base_year,
            ", ".join(rebuild_filter),
            ", ".join(region_filter),
            "Yes" if apply_inflation else "No",
            "Built-in Expanded Dealer Rates" if use_default else "Uploaded Custom Dealer Rates",
            export_mode,
            export_reason_final,
        ],
    })


def build_role_policy_table():
    return pd.DataFrame({
        "Role View": ["Viewer", "Analyst", "Admin"],
        "Intended Capability": [
            "Dashboard and read-only review. Export controls should be limited in a secured deployment.",
            "Analysis review plus selected-machine and standard export workflows.",
            "Full export and advanced configuration review. Enterprise authentication should enforce this in production.",
        ],
        "Current App Behavior": [
            "Role is advisory only unless enterprise authentication is added.",
            "Role is advisory only unless enterprise authentication is added.",
            "Role is advisory only unless enterprise authentication is added.",
        ],
    })


def build_performance_safeguard_summary(source_rows, processed_rows):
    warnings = []
    if source_rows > max_rows_warning_threshold:
        warnings.append({"Safeguard": "Source Row Count", "Status": "Review", "Details": f"Source workbook contains {source_rows:,} rows, above threshold {max_rows_warning_threshold:,}. Consider Summary Only export if performance is slow."})
    else:
        warnings.append({"Safeguard": "Source Row Count", "Status": "OK", "Details": f"Source workbook contains {source_rows:,} rows."})
    if processed_rows > max_rows_warning_threshold:
        warnings.append({"Safeguard": "Processed Row Count", "Status": "Review", "Details": f"Processed data contains {processed_rows:,} rows. Large raw-data export may take longer."})
    else:
        warnings.append({"Safeguard": "Processed Row Count", "Status": "OK", "Details": f"Processed data contains {processed_rows:,} rows."})
    if export_mode == "Full Analysis Workbook" and processed_rows > max_rows_warning_threshold:
        warnings.append({"Safeguard": "Export Mode", "Status": "Review", "Details": "Full Analysis Workbook selected for a large run. Summary Only or Exceptions Only may be more efficient."})
    else:
        warnings.append({"Safeguard": "Export Mode", "Status": "OK", "Details": f"Selected export mode: {export_mode}."})
    return pd.DataFrame(warnings)


def apply_confidential_watermark(workbook, scenario_name_value=None):
    """No-op retained for compatibility. V16.2.3 removes confidentiality label rows from all exports."""
    return None


def evaluate_strict_mode(run_readiness_summary, dealer_rate_coverage_summary, data_quality_score_summary, processed_rows):
    """Raise a clear error when Strict Mode is enabled and configured thresholds are not met."""
    if not strict_mode:
        return
    issues = []
    try:
        coverage = float(dealer_rate_coverage_summary.loc[dealer_rate_coverage_summary["Metric"] == "Dealer-Year Match Rate %", "Value"].iloc[0])
        if coverage < float(dealer_rate_coverage_strict_threshold):
            issues.append(f"Dealer-year match rate is {coverage:.1f}%, below strict threshold {dealer_rate_coverage_strict_threshold:.1f}%.")
    except Exception:
        issues.append("Dealer-year match rate could not be evaluated.")
    try:
        dq_score = float(data_quality_score_summary.loc[data_quality_score_summary["Metric"] == "Data Quality Score", "Value"].iloc[0])
        if dq_score < float(data_quality_strict_min_score):
            issues.append(f"Data Quality Score is {dq_score:.1f}, below strict threshold {data_quality_strict_min_score:.1f}.")
    except Exception:
        issues.append("Data Quality Score could not be evaluated.")
    if processed_rows > max_rows_warning_threshold * 2:
        issues.append(f"Processed rows ({processed_rows:,}) exceed 2× the row warning threshold ({max_rows_warning_threshold:,}).")
    if issues:
        raise ValueError("Strict Mode stopped the analysis: " + " ".join(issues))


# =====================================================
# EXPORT FORMATTING HELPERS
# =====================================================
def _truthy_excel_value(value):
    """Treat Excel boolean/string/numeric values as true when appropriate."""
    if value is True:
        return True
    if value is False or value is None:
        return False
    text = str(value).strip().upper()
    return text in {"TRUE", "YES", "Y", "1"}


def apply_excel_brand_formatting(workbook):
    """
    Caterpillar-inspired workbook formatting plus row highlighting.

    Highlight rules:
    - Red fill: Outlier == TRUE
    - Orange fill: Cross-Type Exception Flag is populated
    - Yellow fill: Insufficient Sample Group == TRUE
    """
    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_font = Font(color="FFC500", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    outlier_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    cross_type_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    limited_sample_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    currency_keywords = ["COST", "RATE", "PARTS DN", "LABOR", "BENCHMARK", "AVG"]
    percent_keywords = ["%"]
    decimal_keywords = ["CPI", "FACTOR", "FX", "SCORE"]

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        header_map = {str(cell.value).upper().strip() if cell.value is not None else "": cell.column for cell in ws[1]}
        outlier_col = header_map.get("OUTLIER")
        cross_type_col = header_map.get("CROSS-TYPE EXCEPTION FLAG")
        insufficient_col = header_map.get("INSUFFICIENT SAMPLE GROUP")

        for col_idx in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col_idx).value).upper() if ws.cell(row=1, column=col_idx).value is not None else ""
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(len(header) + 4, 14), 42)
            for cell in ws[col_letter][1:]:
                cell.border = Border(bottom=thin)
                if "YEAR" in header or "SMU" in header or header in ["COUNT", "TOTAL_ROWS", "OUTLIER_ROWS", "VALID_ROWS", "VALUE"]:
                    cell.number_format = "#,##0"
                elif any(keyword in header for keyword in percent_keywords):
                    cell.number_format = "0.0%"
                elif any(keyword in header for keyword in decimal_keywords):
                    cell.number_format = "0.0000"
                elif any(keyword in header for keyword in currency_keywords):
                    cell.number_format = "$#,##0"

        for row_idx in range(2, ws.max_row + 1):
            row_fill = None
            if outlier_col and _truthy_excel_value(ws.cell(row=row_idx, column=outlier_col).value):
                row_fill = outlier_fill
            elif cross_type_col:
                cross_value = ws.cell(row=row_idx, column=cross_type_col).value
                if cross_value is not None and str(cross_value).strip() != "":
                    row_fill = cross_type_fill
            elif insufficient_col and _truthy_excel_value(ws.cell(row=row_idx, column=insufficient_col).value):
                row_fill = limited_sample_fill
            if row_fill is not None:
                for cell in ws[row_idx]:
                    cell.fill = row_fill

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")


# =====================================================
# FX / CPI / RATE HELPERS
# =====================================================
COMMON_CURRENCY_FALLBACK_TO_USD = {"USD": 1.00, "CAD": 0.74, "EUR": 1.09, "GBP": 1.27, "AUD": 0.66, "NZD": 0.61, "MXN": 0.058, "BRL": 0.19, "CLP": 0.0011, "COP": 0.00025, "PEN": 0.27, "JPY": 0.0065, "CNY": 0.14, "INR": 0.012, "ZAR": 0.055}

def normalize_currency_code(value, default_code="USD"):
    if pd.isna(value):
        return default_code.upper()
    code = str(value).strip().upper()
    if code in ["$", "US$", "USDOLLAR", "US DOLLAR", "U.S. DOLLAR", ""]:
        return "USD"
    if len(code) >= 3:
        return code[:3]
    return default_code.upper()

def detect_currency_column(df):
    candidates = ["CURRENCY", "Currency", "currency", "CURR", "Curr", "curr", "CURRENCY CODE", "Currency Code", "currency_code", "CCY"]
    for col in candidates:
        if col in df.columns:
            return col
    for col in df.columns:
        col_upper = str(col).upper()
        if "CURRENCY" in col_upper or col_upper in ["CURR", "CCY"]:
            return col
    return None

def frankfurter_year_average_to_usd(currency, year):
    currency = normalize_currency_code(currency)
    year = int(year)
    if currency == "USD":
        return 1.0, "USD"
    url = f"https://api.frankfurter.dev/v1/{year}-01-01..{year}-12-31?base={currency}&symbols=USD"
    with urllib.request.urlopen(url, timeout=20) as response:
        data = json.loads(response.read().decode("utf-8"))
    rates = []
    for values in data.get("rates", {}).values():
        if "USD" in values:
            rates.append(float(values["USD"]))
    if not rates:
        raise RuntimeError(f"No FX rates returned for {currency} {year}")
    return float(np.mean(rates)), "Frankfurter annual average"

def build_fx_lookup(currencies, years):
    rows = []
    cache = {}
    for currency in sorted(set([normalize_currency_code(c) for c in currencies if pd.notna(c)])):
        for year in sorted(set([int(y) for y in years if pd.notna(y)])):
            key = (currency, year)
            if key in cache:
                fx, source = cache[key]
            else:
                try:
                    fx, source = frankfurter_year_average_to_usd(currency, year)
                except Exception:
                    fx = COMMON_CURRENCY_FALLBACK_TO_USD.get(currency, 1.0)
                    source = "Embedded fallback FX table"
                cache[key] = (fx, source)
            rows.append({"Currency": currency, "Service Year": year, "FX to USD": fx, "FX Source": source})
    return pd.DataFrame(rows)

def fallback_bls_cpi_table(start_year, end_year):
    fallback = {2010: 218.056, 2011: 224.939, 2012: 229.594, 2013: 232.957, 2014: 236.736, 2015: 237.017, 2016: 240.007, 2017: 245.120, 2018: 251.107, 2019: 255.657, 2020: 258.811, 2021: 270.970, 2022: 292.655, 2023: 304.702, 2024: 313.689, 2025: 322.000, 2026: 330.080, 2027: 339.982, 2028: 350.181, 2029: 360.686, 2030: 371.506}
    return {y: fallback[y] for y in range(int(start_year), int(end_year) + 1) if y in fallback}

def fetch_bls_cpi_annual(start_year, end_year):
    url = "https://api.bls.gov/publicAPI/v2/timeseries/data/"
    payload = {"seriesid": ["CUUR0000SA0"], "startyear": str(int(start_year)), "endyear": str(int(end_year))}
    request = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), headers={"Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(request, timeout=20) as response:
        result = json.loads(response.read().decode("utf-8"))
    if result.get("status") != "REQUEST_SUCCEEDED":
        raise RuntimeError("BLS API request did not succeed")
    data = result["Results"]["series"][0]["data"]
    by_year = {}
    annual_m13 = {}
    for row in data:
        year = int(row["year"])
        period = row["period"]
        value = float(row["value"])
        if period == "M13":
            annual_m13[year] = value
        elif period.startswith("M") and period != "M13":
            by_year.setdefault(year, []).append(value)
    cpi = {}
    for year in range(int(start_year), int(end_year) + 1):
        if year in annual_m13:
            cpi[year] = annual_m13[year]
        elif year in by_year and len(by_year[year]) > 0:
            cpi[year] = float(np.mean(by_year[year]))
    return cpi

def get_cpi_table(start_year, end_year, base_year):
    min_year = min(int(start_year), int(base_year))
    max_year = max(int(end_year), int(base_year))
    source = "BLS API: CPI-U All Items, U.S. city average, not seasonally adjusted (CUUR0000SA0)"
    try:
        cpi = fetch_bls_cpi_annual(min_year, max_year)
        if int(base_year) not in cpi:
            raise RuntimeError("Base year CPI unavailable from BLS API")
        return cpi, source
    except Exception:
        return fallback_bls_cpi_table(min_year, max_year), "Fallback embedded CPI table because BLS API was unavailable"

def build_default_rate_table(start=2010, end=2030):
    """
    Return built-in expanded dealer-year labor rates from the bundled workbook.
    The workbook must sit in the same folder as app.py when deployed.
    Falls back to the old generic yearly rate table only if the workbook is missing/unreadable.
    """
    built_in_rate_path = Path("expanded_dealer_base_rate_by_year_2016_2026.xlsx")
    if built_in_rate_path.exists():
        try:
            rate_df = parse_multisheet_rate_workbook(built_in_rate_path)
            rate_df = clean_rate_table(rate_df)
            if not rate_df.empty:
                if start is not None and end is not None:
                    rate_df = rate_df[(rate_df["Service Year"] >= int(start)) & (rate_df["Service Year"] <= int(end))].copy()
                rate_df["Rate File Format"] = "Built-in Expanded Dealer Rates 2016-2026"
                rate_df["Notes"] = "Built-in expanded dealer base rate workbook: 2016-2026"
                return rate_df
        except Exception:
            pass

    # Emergency fallback only if the built-in workbook is unavailable.
    years = list(range(start, end + 1))
    return pd.DataFrame({
        "Dealer Code": ["DEFAULT"] * len(years),
        "Service Year": years,
        "Rate": [115 + (year - 2016) * 3 for year in years],
        "Rate Currency": ["USD"] * len(years),
        "Notes": ["Emergency generic fallback because built-in dealer workbook was not found/readable"] * len(years),
        "Rate File Sheet": ["Emergency Fallback"] * len(years),
        "Rate File Format": ["Emergency Generic Default"] * len(years),
    })


def normalize_dealer_code(value):
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    match = re.search(r"([A-Z][A-Z0-9]{3})", text)
    return match.group(1) if match else text


def normalize_rate_columns(df):
    """Normalize common dealer-rate column names to Dealer Code, Service Year, Rate, Rate Currency, Notes."""
    temp = df.copy()
    temp.columns = [str(col).strip() for col in temp.columns]
    rename_map = {}
    for col in temp.columns:
        cu = str(col).strip().upper().replace("_", " ")
        if cu in ["DEALER", "DEALER CODE", "DEALER ID", "CODE"] or ("DEALER" in cu and "CODE" in cu):
            rename_map[col] = "Dealer Code"
        elif cu in ["YEAR", "SERVICE YEAR", "RATE YEAR", "BASE RATE YEAR"] or ("YEAR" in cu and "RATE" in cu):
            rename_map[col] = "Service Year"
        elif cu in ["RATE", "LABOR RATE", "BASE RATE", "AVG BASE RATE", "AVERAGE BASE RATE", "DEALER AVERAGE BASE RATE"] or ("RATE" in cu and "SOURCE" not in cu and "CURRENCY" not in cu):
            rename_map[col] = "Rate"
        elif cu in ["CURRENCY", "RATE CURRENCY", "CURRENCY CODE", "CCY"]:
            rename_map[col] = "Rate Currency"
        elif cu in ["NOTE", "NOTES", "COMMENTS", "COMMENT"]:
            rename_map[col] = "Notes"
    temp = temp.rename(columns=rename_map)
    return temp


def parse_flat_rate_table(rate_file, sheet_name=None):
    """Parse a preferred flat dealer-rate table from one sheet.

    Preferred V17.1 template format:
    - Instructions: user guidance only
    - Example: sample rows only
    - User Input: actual data to upload

    If a User Input sheet exists, only that sheet is parsed. If not, the parser
    falls back to scanning other sheets for backward compatibility.
    """
    sheets = pd.read_excel(rate_file, sheet_name=None)
    candidate_items = []
    if sheet_name and sheet_name in sheets:
        candidate_items = [(sheet_name, sheets[sheet_name])]
    else:
        preferred_sheet = None
        for sname in sheets.keys():
            if str(sname).strip().lower() == "user input":
                preferred_sheet = sname
                break
        if preferred_sheet:
            candidate_items = [(preferred_sheet, sheets[preferred_sheet])]
        else:
            candidate_items = [(sname, sheet) for sname, sheet in sheets.items() if str(sname).strip().lower() not in ["instructions", "example"]]

    for sname, sheet in candidate_items:
        temp = normalize_rate_columns(sheet)
        if {"Dealer Code", "Service Year", "Rate"}.issubset(set(temp.columns)):
            out = temp.copy()
            if "Rate Currency" not in out.columns:
                out["Rate Currency"] = "USD"
            if "Notes" not in out.columns:
                out["Notes"] = ""
            out = out[["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes"]].copy()
            out["Dealer Code"] = out["Dealer Code"].astype(str).str.strip().str.upper()
            out = out[(out["Dealer Code"] != "") & (out["Dealer Code"].str.upper() != "NAN")]
            if not out.empty:
                out["Rate File Sheet"] = sname
                out["Rate File Format"] = "Flat Table"
                return out[["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes", "Rate File Sheet", "Rate File Format"]]
    return pd.DataFrame(columns=["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes", "Rate File Sheet", "Rate File Format"])


def parse_multisheet_rate_workbook(rate_file):
    """Parse legacy dealer-rate workbook where each sheet represents a dealer."""
    sheets = pd.read_excel(rate_file, sheet_name=None)
    rows = []
    for sheet_name, sheet in sheets.items():
        if str(sheet_name).strip().lower() in ["summary", "instructions", "example"]:
            continue
        temp = normalize_rate_columns(sheet)
        dealer_code = normalize_dealer_code(str(sheet_name).split("_")[0].strip())
        if "Dealer Code" not in temp.columns:
            temp["Dealer Code"] = dealer_code
        if "Rate Currency" not in temp.columns:
            temp["Rate Currency"] = "USD"
        if "Notes" not in temp.columns:
            temp["Notes"] = ""
        if "Service Year" in temp.columns and "Rate" in temp.columns:
            keep = temp[["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes"]].copy()
            keep["Rate File Sheet"] = sheet_name
            keep["Rate File Format"] = "Multi-Sheet Dealer Workbook"
            rows.append(keep)
    if not rows:
        return pd.DataFrame(columns=["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes", "Rate File Sheet", "Rate File Format"])
    return pd.concat(rows, ignore_index=True)


def clean_rate_table(rate_df):
    """Clean, type, and de-duplicate dealer rate table."""
    if rate_df is None or rate_df.empty:
        return pd.DataFrame(columns=["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes", "Rate File Sheet", "Rate File Format"])
    out = rate_df.copy()
    out["Dealer Code"] = out["Dealer Code"].apply(normalize_dealer_code)
    out["Service Year"] = pd.to_numeric(out["Service Year"], errors="coerce")
    out["Rate"] = pd.to_numeric(out["Rate"], errors="coerce")
    out["Rate Currency"] = out.get("Rate Currency", "USD")
    out["Rate Currency"] = out["Rate Currency"].apply(lambda x: normalize_currency_code(x, "USD"))
    if "Notes" not in out.columns:
        out["Notes"] = ""
    if "Rate File Sheet" not in out.columns:
        out["Rate File Sheet"] = "Unknown"
    if "Rate File Format" not in out.columns:
        out["Rate File Format"] = "Unknown"
    out = out.dropna(subset=["Service Year", "Rate"])
    out = out[out["Dealer Code"].astype(str).str.strip() != ""]
    out = out[out["Rate"] > 0]
    out["Service Year"] = out["Service Year"].astype(int)
    out = out.sort_values(["Dealer Code", "Service Year", "Rate File Sheet"]).drop_duplicates(["Dealer Code", "Service Year"], keep="last")
    return out[["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes", "Rate File Sheet", "Rate File Format"]]


def build_rate_table_from_workbook(rate_file, rate_format="Auto-detect"):
    """Build dealer rate table from either the preferred flat table or legacy multi-sheet workbook."""
    if rate_file is None:
        return build_default_rate_table()
    rate_file.seek(0)
    if rate_format == "Flat Table":
        rate_df = parse_flat_rate_table(rate_file)
    elif rate_format == "Multi-Sheet Dealer Workbook":
        rate_file.seek(0)
        rate_df = parse_multisheet_rate_workbook(rate_file)
    else:
        rate_df = parse_flat_rate_table(rate_file)
        if rate_df.empty:
            rate_file.seek(0)
            rate_df = parse_multisheet_rate_workbook(rate_file)
    return clean_rate_table(rate_df)


def validate_dealer_rate_table(rate_df, selected_start_year=None, selected_end_year=None):
    """Create validation summary for uploaded dealer labor rates."""
    if rate_df is None or rate_df.empty:
        return pd.DataFrame({"Check": ["Rate rows parsed"], "Value": [0], "Status": ["Review"]})
    checks = []
    checks.append({"Check": "Rate rows parsed", "Value": len(rate_df), "Status": "OK" if len(rate_df) > 0 else "Review"})
    checks.append({"Check": "Unique dealers", "Value": rate_df["Dealer Code"].nunique(), "Status": "OK"})
    checks.append({"Check": "Minimum service year", "Value": int(rate_df["Service Year"].min()) if not rate_df.empty else "", "Status": "OK"})
    checks.append({"Check": "Maximum service year", "Value": int(rate_df["Service Year"].max()) if not rate_df.empty else "", "Status": "OK"})
    checks.append({"Check": "Missing dealer codes", "Value": int((rate_df["Dealer Code"].astype(str).str.strip() == "").sum()), "Status": "OK" if int((rate_df["Dealer Code"].astype(str).str.strip() == "").sum()) == 0 else "Review"})
    checks.append({"Check": "Non-positive rates", "Value": int((rate_df["Rate"] <= 0).sum()), "Status": "OK" if int((rate_df["Rate"] <= 0).sum()) == 0 else "Review"})
    duplicate_count = int(rate_df.duplicated(["Dealer Code", "Service Year"]).sum())
    checks.append({"Check": "Duplicate dealer-year rows after cleaning", "Value": duplicate_count, "Status": "OK" if duplicate_count == 0 else "Review"})
    if selected_start_year is not None and selected_end_year is not None:
        outside = int(((rate_df["Service Year"] < selected_start_year) | (rate_df["Service Year"] > selected_end_year)).sum())
        checks.append({"Check": "Rows outside selected analysis year range", "Value": outside, "Status": "OK" if outside == 0 else "Info"})
    checks.append({"Check": "Detected rate format(s)", "Value": ", ".join(sorted(rate_df["Rate File Format"].dropna().unique())), "Status": "OK"})
    return pd.DataFrame(checks)


def create_dealer_rate_template():
    """Return a three-sheet preferred flat dealer-rate template workbook."""
    instructions = pd.DataFrame({
        "Section": [
            "Purpose", "How to Use", "How to Use", "How to Use", "How to Use", "How to Use", "How to Use",
            "Required Field", "Required Field", "Required Field", "Optional Field", "Optional Field",
            "Validation Rule", "Validation Rule", "Validation Rule", "Validation Rule", "Upload Behavior"
        ],
        "Instruction": [
            "Use this template to upload custom dealer labor rates into the Rebuild Analytics Platform.",
            "Read this Instructions sheet first.",
            "Review the Example sheet to see the correct format.",
            "Enter actual dealer-rate data only on the User Input sheet.",
            "Do not rename, delete, or reorder required columns.",
            "Do not type dollar signs or currency symbols into the Rate field; use a numeric value only.",
            "Save the workbook as .xlsx and upload it in the Custom Dealer Labor Rate Workbook section of the app.",
            "Dealer Code must match the dealer code used in the rebuild file, such as A000 or B123.",
            "Service Year must be a four-digit year, such as 2024 or 2025.",
            "Rate must be numeric and greater than zero.",
            "Rate Currency should be a three-letter currency code, such as USD, CAD, or EUR. If blank, the app expects USD.",
            "Notes is optional and can document source, approval, or rate context.",
            "Dealer Code cannot be blank.",
            "Service Year must be numeric.",
            "Rate must be numeric and greater than zero.",
            "Each Dealer Code + Service Year combination should be unique. If duplicates exist after cleaning, the last valid row is used.",
            "If a dealer-year rate is missing from the uploaded file, the app applies the selected fallback behavior and flags the affected rows.",
        ]
    })
    example = pd.DataFrame({
        "Dealer Code": ["A000", "A000", "B123"],
        "Service Year": [2024, 2025, 2025],
        "Rate": [150.00, 155.00, 162.50],
        "Rate Currency": ["USD", "USD", "USD"],
        "Notes": ["Example only", "Example only", "Example only"],
    })
    user_input = pd.DataFrame(columns=["Dealer Code", "Service Year", "Rate", "Rate Currency", "Notes"])
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="Instructions", index=False)
        example.to_excel(writer, sheet_name="Example", index=False)
        user_input.to_excel(writer, sheet_name="User Input", index=False)
        apply_excel_brand_formatting(writer.book)
        apply_confidential_watermark(writer.book)
    output.seek(0)
    return output.getvalue()


# =====================================================
# PRE-RUN VALIDATION / UI
# =====================================================
def read_rebuild_workbook(uploaded_file):
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=None)

def validate_uploaded_workbook(uploaded_file):
    try:
        rebuild = read_rebuild_workbook(uploaded_file)
    except Exception as exc:
        return None, pd.DataFrame({"Issue": [f"Unable to read workbook: {exc}"]})
    required = ["DELIVERED DATE", "ENROLL DATE", "IN SERVICE DATE", "SALES MODEL", "DEALER", "PARTS DN", "SMU AT REBUILD", "REBUILD WORK HRS"]
    issues = []
    rows = []
    all_frames = []
    for sheet_name, sheet in rebuild.items():
        sheet = sheet.copy()
        missing = [col for col in required if col not in sheet.columns]
        detected_type = normalize_ccr_type(sheet_name, sheet_name)
        rows.append({"Sheet": sheet_name, "Rows": len(sheet), "Detected CCR TYPE": detected_type, "Missing Required Columns": ", ".join(missing) if missing else "None"})
        if missing:
            issues.append({"Issue Type": "Missing Columns", "Sheet": sheet_name, "Details": ", ".join(missing)})
        sheet["_Detected CCR TYPE"] = detected_type
        all_frames.append(sheet)
    profile = pd.DataFrame(rows)
    if all_frames:
        combined = pd.concat(all_frames, ignore_index=True)
        if "REGION" in combined.columns:
            observed_regions = sorted(combined["REGION"].dropna().apply(normalize_region).unique())
            for region in [region for region in observed_regions if region not in VALID_REGIONS]:
                issues.append({"Issue Type": "Unconfigured Region", "Sheet": "All", "Details": region})
        observed_types = sorted(combined["_Detected CCR TYPE"].dropna().unique())
        for rebuild_type in [rtype for rtype in observed_types if rtype not in CERTIFIED_REBUILD_TYPES]:
            issues.append({"Issue Type": "Unknown Rebuild Type", "Sheet": "All", "Details": rebuild_type})
    issue_df = pd.DataFrame(issues) if issues else pd.DataFrame(columns=["Issue Type", "Sheet", "Details"])
    return profile, issue_df

# =====================================================
# SINGLE MACHINE EXPORT - V15.4
# =====================================================
def dealer_performance_for_df(machine_valid, machine_all, cost_col):
    if machine_valid.empty:
        return pd.DataFrame()
    dealer_summary = machine_valid.groupby("DEALER").agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
        Cross_Type_Flags=("Cross-Type Exception Flag", lambda x: (x != "").sum()),
    ).reset_index().sort_values("Avg_Cost", ascending=False)
    dealer_summary = add_vs_section_avg(dealer_summary)
    dealer_outliers = machine_all.groupby("DEALER").agg(
        Total_Rows=(cost_col, "count"),
        Outlier_Rows=("Outlier", "sum"),
        DQ_Rows=("Data Quality Exception Flag", lambda x: (x != "").sum()),
    ).reset_index()
    dealer_summary = dealer_summary.merge(dealer_outliers, on="DEALER", how="left")
    dealer_summary["Outlier Rate %"] = (dealer_summary["Outlier_Rows"] / dealer_summary["Total_Rows"]) * 100
    dealer_summary["DQ Rate %"] = (dealer_summary["DQ_Rows"] / dealer_summary["Total_Rows"]) * 100
    dealer_summary["Cross Flag Rate %"] = (dealer_summary["Cross_Type_Flags"] / dealer_summary["Count"]) * 100
    dealer_summary["Performance Score"] = (
        100
        - dealer_summary["Vs Section Avg %"].clip(lower=0).fillna(0) * 0.50
        - dealer_summary["Outlier Rate %"].fillna(0) * 0.30
        - dealer_summary["Cross Flag Rate %"].fillna(0) * 0.20
        - dealer_summary["DQ Rate %"].fillna(0) * 0.20
    ).clip(lower=0, upper=100)
    dealer_summary["Performance Label"] = np.where(dealer_summary["Performance Score"] >= 85, "Strong", np.where(dealer_summary["Performance Score"] >= 70, "Monitor", "Review Needed"))
    return dealer_summary

def region_performance_for_df(machine_valid, cost_col):
    if machine_valid.empty:
        return pd.DataFrame()
    region_summary = machine_valid.groupby("Region").agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
        Cross_Type_Flags=("Cross-Type Exception Flag", lambda x: (x != "").sum()),
    ).reset_index().sort_values("Avg_Cost", ascending=False)
    region_summary = add_vs_section_avg(region_summary)
    return region_summary

def build_machine_export(selected_machine, analysis):
    df = analysis["df"]
    valid = analysis["valid"]
    cost_col = analysis["cost_col"]
    metadata = analysis["metadata"].copy()
    fx_lookup = analysis["fx_lookup"].copy()
    cpi_table = analysis["cpi_table"]

    machine_all = df[df["SALES MODEL"] == selected_machine].copy()
    machine_valid = valid[valid["SALES MODEL"] == selected_machine].copy()
    machine_adjusted_valid = machine_valid[~((machine_valid["CCR TYPE"] == "CPT+H") & (machine_valid["Cross-Type Exception Flag"] == "CPT+H Cost Above Typical CMR"))].copy()

    standard_avg = machine_valid[cost_col].mean() if not machine_valid.empty else np.nan
    adjusted_avg = machine_adjusted_valid[cost_col].mean() if not machine_adjusted_valid.empty else np.nan
    adjusted_cpth_avg = machine_adjusted_valid[machine_adjusted_valid["CCR TYPE"] == "CPT+H"][cost_col].mean() if not machine_adjusted_valid.empty else np.nan

    machine_summary = pd.DataFrame({
        "Metric": [
            "Selected Machine", "Total Rows", "Valid Rows", "Standard Avg Cost", "Adjusted Avg Cost",
            "Adjusted CPT+H Avg", "Average SMU", "Cost Outliers", "Cross-Type Flags",
            "SMU Data Quality Flags", "Fallback Labor Rate Rows", "Fallback FX Rows"
        ],
        "Value": [
            selected_machine, len(machine_all), len(machine_valid), standard_avg, adjusted_avg, adjusted_cpth_avg,
            machine_valid["SMU AT REBUILD"].mean() if not machine_valid.empty else np.nan,
            int(machine_all["Outlier"].sum()) if "Outlier" in machine_all.columns else 0,
            int((machine_valid["Cross-Type Exception Flag"] != "").sum()) if "Cross-Type Exception Flag" in machine_valid.columns else 0,
            int((machine_all["Data Quality Exception Flag"] != "").sum()) if "Data Quality Exception Flag" in machine_all.columns else 0,
            int(machine_all["Rate Source"].isin(["Global Yearly Fallback Rate", "Overall Average Fallback Rate"]).sum()) if "Rate Source" in machine_all.columns else 0,
            int((machine_all["FX Source"] == "Embedded fallback FX table").sum()) if "FX Source" in machine_all.columns else 0,
        ]
    })

    
    if "Metric" in machine_summary.columns:
        machine_summary = machine_summary[~machine_summary["Metric"].astype(str).str.contains("Confidential|Removed Label", case=False, na=False)].copy()
    if "Value" in machine_summary.columns:
        machine_summary = machine_summary[~machine_summary["Value"].astype(str).str.contains("Confidential|Yellow", case=False, na=False)].copy()
    rebuild_summary = machine_valid.groupby("CCR TYPE").agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
    ).reset_index() if not machine_valid.empty else pd.DataFrame(columns=["CCR TYPE", "Avg_Cost", "Avg_SMU", "Count"])
    rebuild_summary = add_vs_cmr(rebuild_summary)
    if not rebuild_summary.empty:
        rebuild_summary["Sample Confidence"] = rebuild_summary["Count"].apply(sample_confidence)

    adjusted_summary = machine_adjusted_valid.groupby("CCR TYPE").agg(
        Adjusted_Avg_Cost=(cost_col, "mean"),
        Count=(cost_col, "count"),
    ).reset_index() if not machine_adjusted_valid.empty else pd.DataFrame(columns=["CCR TYPE", "Adjusted_Avg_Cost", "Count"])
    if not adjusted_summary.empty:
        adjusted_summary["CCR TYPE"] = adjusted_summary["CCR TYPE"].replace({"CPT+H": "CPT+H Adjusted"})
        adjusted_summary["Sample Confidence"] = adjusted_summary["Count"].apply(sample_confidence)

    standard_cpth = machine_valid[machine_valid["CCR TYPE"] == "CPT+H"][cost_col].mean() if not machine_valid.empty else np.nan
    adjusted_cpth = adjusted_cpth_avg
    cpth_diff = adjusted_cpth - standard_cpth if pd.notna(adjusted_cpth) and pd.notna(standard_cpth) else np.nan
    cpth_diff_pct = (cpth_diff / standard_cpth) if pd.notna(cpth_diff) and standard_cpth else np.nan
    adjusted_comparison = pd.DataFrame({
        "Metric": ["Standard CPT+H Avg", "Adjusted CPT+H Avg", "Difference $", "Difference %", "Cross-Type Outliers Excluded"],
        "Value": [standard_cpth, adjusted_cpth, cpth_diff, cpth_diff_pct, int(((machine_valid["CCR TYPE"] == "CPT+H") & (machine_valid["Cross-Type Exception Flag"] != "")).sum()) if not machine_valid.empty else 0]
    })

    dealer_summary = dealer_performance_for_df(machine_valid, machine_all, cost_col)
    region_summary = region_performance_for_df(machine_valid, cost_col)

    exceptions_summary = pd.DataFrame({
        "Metric": ["Cost Outliers", "CPT+H Cross-Type Outliers", "SMU 0 or 1", "Insufficient Sample Rows", "Rows Using Fallback FX", "Rows Using Global Yearly Fallback Rate", "Rows Using Overall Average Fallback Rate"],
        "Value": [
            int(machine_all["Outlier"].sum()) if "Outlier" in machine_all.columns else 0,
            int((machine_valid["Cross-Type Exception Flag"] != "").sum()) if "Cross-Type Exception Flag" in machine_valid.columns else 0,
            int(machine_all["Data Quality Exception Flag"].eq("SMU 0 or 1").sum()) if "Data Quality Exception Flag" in machine_all.columns else 0,
            int(machine_all["Insufficient Sample Group"].sum()) if "Insufficient Sample Group" in machine_all.columns else 0,
            int((machine_all["FX Source"] == "Embedded fallback FX table").sum()) if "FX Source" in machine_all.columns else 0,
            int((machine_all["Rate Source"] == "Global Yearly Fallback Rate").sum()) if "Rate Source" in machine_all.columns else 0,
            int((machine_all["Rate Source"] == "Overall Average Fallback Rate").sum()) if "Rate Source" in machine_all.columns else 0,
        ]
    })

    outlier_perf = machine_all.groupby("CCR TYPE").agg(
        Total_Rows=(cost_col, "count"),
        Outlier_Rows=("Outlier", "sum"),
        Avg_Cost_All=(cost_col, "mean"),
    ).reset_index() if not machine_all.empty else pd.DataFrame(columns=["CCR TYPE", "Total_Rows", "Outlier_Rows", "Avg_Cost_All"])
    valid_counts = machine_valid.groupby("CCR TYPE").size().reset_index(name="Valid_Rows") if not machine_valid.empty else pd.DataFrame(columns=["CCR TYPE", "Valid_Rows"])
    outlier_perf = outlier_perf.merge(valid_counts, on="CCR TYPE", how="left").fillna({"Valid_Rows": 0})
    if not outlier_perf.empty:
        outlier_perf["Outlier Rate %"] = (outlier_perf["Outlier_Rows"] / outlier_perf["Total_Rows"])

    cross_rows = machine_valid[machine_valid["Cross-Type Exception Flag"] != ""].copy() if "Cross-Type Exception Flag" in machine_valid.columns else pd.DataFrame()
    if cross_rows.empty:
        cross_rows = pd.DataFrame({"Message": ["No cross-type exceptions detected for this machine."]})

    run_metadata = pd.concat([
        pd.DataFrame({"Field": ["Selected Machine"], "Value": [selected_machine]}),
        metadata,
    ], ignore_index=True)

    methodology = pd.DataFrame({
        "Methodology Item": [
            "Cost Basis", "Labor Cost", "Currency", "Inflation", "Outliers", "Sample Size", "SMU", "Cross-Type Rule", "Adjusted CPT+H", "Dealer Score"
        ],
        "Description": [
            "Adjusted Total Cost USD = PARTS DN USD + Labor Cost USD. PLUS PARTS DN is ignored.",
            "Labor Cost USD = REBUILD WORK HRS × dealer service-year labor rate converted to USD when applicable.",
            "Source currency is converted to USD before inflation and analysis.",
            "BLS CPI-U All Items, U.S. city average, not seasonally adjusted. Inflation applied by component.",
            "Cost-only log(cost) + IQR by Machine + CCR TYPE.",
            "<5 rows: no removal; 5–7 rows: 2.0×IQR; 8+ rows: 1.5×IQR.",
            "SMU 0 or 1 is a data-quality flag only and is not a statistical outlier basis.",
            "CPT+H flagged when above machine-level valid CMR median × 1.10, if at least 3 valid CMR rows exist.",
            "V18 treats cross-type CPT+H exceptions as outliers. The app reports one CPT+H value after exclusions and exports excluded rows for audit.",
            "Score subtracts weighted impacts from above-average cost, outlier rate, cross-type rate, and data-quality rate.",
        ]
    })

    cpi_export = pd.DataFrame({"Year": list(cpi_table.keys()), "CPI": list(cpi_table.values())}).sort_values("Year")

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        machine_summary.to_excel(writer, sheet_name="Machine Summary", index=False)
        rebuild_summary.to_excel(writer, sheet_name="Rebuild Type Summary", index=False)
        adjusted_summary.to_excel(writer, sheet_name="Adjusted Cost Summary", index=False)
        adjusted_comparison.to_excel(writer, sheet_name="Adjusted CPT-H Impact", index=False)
        dealer_summary.to_excel(writer, sheet_name="Dealer Performance", index=False)
        region_summary.to_excel(writer, sheet_name="Region Performance", index=False)
        machine_all.to_excel(writer, sheet_name="Raw Data", index=False)
        machine_valid.to_excel(writer, sheet_name="Valid Raw Data", index=False)
        exceptions_summary.to_excel(writer, sheet_name="Exceptions", index=False)
        outlier_perf.to_excel(writer, sheet_name="Outlier Performance", index=False)
        cross_rows.to_excel(writer, sheet_name="Cross Type Flags", index=False)
        machine_rate_exceptions = machine_all[machine_all.get("Dealer Rate Exception Flag", "") != ""].copy() if "Dealer Rate Exception Flag" in machine_all.columns else pd.DataFrame()
        if machine_rate_exceptions.empty:
            machine_rate_exceptions = pd.DataFrame({"Message": ["No dealer rate exceptions detected for this machine."]})
        machine_rate_exceptions.to_excel(writer, sheet_name="Rate Exceptions", index=False)
        run_metadata.to_excel(writer, sheet_name="Run Metadata", index=False)
        methodology.to_excel(writer, sheet_name="Methodology", index=False)
        cpi_export.to_excel(writer, sheet_name="BLS CPI", index=False)
        fx_lookup.to_excel(writer, sheet_name="FX Rates", index=False)
        apply_excel_brand_formatting(writer.book)
        apply_confidential_watermark(writer.book, globals().get("scenario_name", None))
    output.seek(0)
    return output.getvalue()


# =====================================================
# POWER BI DATASET EXPORT HELPERS - V16.2
# =====================================================
def _clean_powerbi_columns(df):
    """Return a clean copy with Power BI-friendly column names and data types preserved."""
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    out = df.copy()
    # Power BI handles spaces, but removing line breaks and repeated whitespace makes the model cleaner.
    out.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in out.columns]
    return out


def _add_run_columns(df, run_id, scenario_name_value):
    out = _clean_powerbi_columns(df)
    if out.empty:
        return out
    if "Run ID" not in out.columns:
        out.insert(0, "Run ID", run_id)
    if "Scenario Name" not in out.columns:
        out.insert(1, "Scenario Name", scenario_name_value if scenario_name_value else "Not provided")
    return out


def _percent_from_counts(numerator, denominator):
    try:
        denominator = float(denominator)
        if denominator == 0:
            return np.nan
        return float(numerator) / denominator * 100
    except Exception:
        return np.nan


def build_powerbi_exception_rows(processed_df, cost_col, run_id, scenario_name_value):
    """Create one clean exception table for Power BI visuals."""
    rows = []
    if processed_df is None or processed_df.empty:
        return pd.DataFrame(columns=["Run ID", "Scenario Name", "SALES MODEL", "DEALER", "Dealer Code", "Region", "CCR TYPE", "Service Year", "Exception Type", "Exception Detail", "Cost", "SMU AT REBUILD"])

    base_cols = ["SALES MODEL", "DEALER", "Dealer Code", "Region", "CCR TYPE", "Service Year", "SMU AT REBUILD"]
    for _, r in processed_df.iterrows():
        base = {col: r.get(col, np.nan) for col in base_cols}
        base["Cost"] = r.get(cost_col, np.nan)
        if bool(r.get("Outlier", False)):
            rec = base.copy()
            rec["Exception Type"] = "Cost Outlier"
            rec["Exception Detail"] = r.get("Outlier Reason", "Cost outlier")
            rows.append(rec)
        if str(r.get("Cross-Type Exception Flag", "")).strip():
            rec = base.copy()
            rec["Exception Type"] = "Cross-Type Exception"
            rec["Exception Detail"] = r.get("Cross-Type Exception Flag", "")
            rows.append(rec)
        if str(r.get("Data Quality Exception Flag", "")).strip():
            rec = base.copy()
            rec["Exception Type"] = "Data Quality"
            rec["Exception Detail"] = r.get("Data Quality Exception Flag", "")
            rows.append(rec)
        if bool(r.get("Insufficient Sample Group", False)):
            rec = base.copy()
            rec["Exception Type"] = "Insufficient Sample"
            rec["Exception Detail"] = "Machine + CCR TYPE group has fewer than 5 records"
            rows.append(rec)
        if str(r.get("Dealer Rate Exception Flag", "")).strip():
            rec = base.copy()
            rec["Exception Type"] = "Dealer Rate Exception"
            rec["Exception Detail"] = r.get("Dealer Rate Exception Flag", "")
            rows.append(rec)
        if str(r.get("FX Source", "")).strip() == "Embedded fallback FX table":
            rec = base.copy()
            rec["Exception Type"] = "FX Fallback"
            rec["Exception Detail"] = "Embedded fallback FX table used"
            rows.append(rec)

    out = pd.DataFrame(rows)
    return _add_run_columns(out, run_id, scenario_name_value)


def build_powerbi_dealer_performance(valid_df, processed_df, cost_col, run_id, scenario_name_value):
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    group_cols = ["SALES MODEL", "Dealer Code", "DEALER", "Region"]
    valid_cols = [c for c in group_cols if c in valid_df.columns]
    proc_cols = [c for c in group_cols if c in processed_df.columns]
    perf = valid_df.groupby(valid_cols, dropna=False).agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
        Cross_Type_Flags=("Cross-Type Exception Flag", lambda x: (x.astype(str).str.strip() != "").sum()),
        Dealer_Rate_Exception_Rows=("Dealer Rate Exception Flag", lambda x: (x.astype(str).str.strip() != "").sum()),
        Data_Quality_Rows=("Data Quality Exception Flag", lambda x: (x.astype(str).str.strip() != "").sum()),
    ).reset_index()
    full = processed_df.groupby(proc_cols, dropna=False).agg(
        Total_Rows=(cost_col, "count"),
        Outlier_Rows=("Outlier", "sum"),
    ).reset_index()
    out = perf.merge(full, on=valid_cols, how="left")
    out["Outlier Rate %"] = out.apply(lambda r: _percent_from_counts(r.get("Outlier_Rows", 0), r.get("Total_Rows", 0)), axis=1)
    out["Cross Flag Rate %"] = out.apply(lambda r: _percent_from_counts(r.get("Cross_Type_Flags", 0), r.get("Count", 0)), axis=1)
    out["Dealer Rate Exception Rate %"] = out.apply(lambda r: _percent_from_counts(r.get("Dealer_Rate_Exception_Rows", 0), r.get("Count", 0)), axis=1)
    out["Data Quality Rate %"] = out.apply(lambda r: _percent_from_counts(r.get("Data_Quality_Rows", 0), r.get("Count", 0)), axis=1)
    section_avg = out["Avg_Cost"].mean()
    out["Vs Overall Avg %"] = ((out["Avg_Cost"] - section_avg) / section_avg * 100) if section_avg else np.nan
    out["Performance Score"] = (
        100
        - out["Vs Overall Avg %"].clip(lower=0).fillna(0) * 0.50
        - out["Outlier Rate %"].fillna(0) * 0.30
        - out["Cross Flag Rate %"].fillna(0) * 0.20
        - out["Dealer Rate Exception Rate %"].fillna(0) * 0.20
        - out["Data Quality Rate %"].fillna(0) * 0.20
    ).clip(lower=0, upper=100).round(1)
    out["Performance Label"] = np.where(out["Performance Score"] >= 85, "Strong", np.where(out["Performance Score"] >= 70, "Monitor", "Review Needed"))
    return _add_run_columns(out, run_id, scenario_name_value)


def build_powerbi_region_performance(valid_df, processed_df, cost_col, run_id, scenario_name_value):
    if valid_df is None or valid_df.empty:
        return pd.DataFrame()
    group_cols = ["SALES MODEL", "Region"]
    perf = valid_df.groupby(group_cols, dropna=False).agg(
        Avg_Cost=(cost_col, "mean"),
        Avg_SMU=("SMU AT REBUILD", "mean"),
        Count=(cost_col, "count"),
        Cross_Type_Flags=("Cross-Type Exception Flag", lambda x: (x.astype(str).str.strip() != "").sum()),
        Data_Quality_Rows=("Data Quality Exception Flag", lambda x: (x.astype(str).str.strip() != "").sum()),
    ).reset_index()
    full = processed_df.groupby(group_cols, dropna=False).agg(
        Total_Rows=(cost_col, "count"),
        Outlier_Rows=("Outlier", "sum"),
    ).reset_index()
    out = perf.merge(full, on=group_cols, how="left")
    out["Outlier Rate %"] = out.apply(lambda r: _percent_from_counts(r.get("Outlier_Rows", 0), r.get("Total_Rows", 0)), axis=1)
    overall = out["Avg_Cost"].mean()
    out["Vs Overall Avg %"] = ((out["Avg_Cost"] - overall) / overall * 100) if overall else np.nan
    return _add_run_columns(out, run_id, scenario_name_value)



def build_dim_machine_group(processed_df, run_id, scenario_label):
    cols = [c for c in ["Machine Group", "Machine Family", "Machine Category", "Machine Group Source", "Machine Group Notes"] if c in processed_df.columns]
    if not cols or "Machine Group" not in cols:
        return pd.DataFrame(columns=["Run ID", "Scenario Name", "Machine Group", "Machine Family", "Machine Category", "Machine Group Source", "Machine Group Notes"])
    temp = processed_df[cols].copy()
    temp = temp[temp["Machine Group"].astype(str).str.strip() != ""]
    if temp.empty:
        return _add_run_columns(temp, run_id, scenario_label)
    # One row per Machine Group to act as a clean slicer/reference dimension.
    agg_map = {col: (lambda x: next((str(v) for v in x.dropna().astype(str).unique() if str(v).strip()), "")) for col in cols if col != "Machine Group"}
    out = temp.groupby("Machine Group", dropna=False).agg(agg_map).reset_index()
    out["Machine Count"] = processed_df.groupby("Machine Group")["SALES MODEL"].nunique().reindex(out["Machine Group"]).fillna(0).astype(int).values if "SALES MODEL" in processed_df.columns else 0
    return _add_run_columns(out.sort_values("Machine Group"), run_id, scenario_label)


def build_powerbi_relationship_checks(tables, relationship_guide):
    rows = []
    for _, rel in relationship_guide.iterrows():
        fact_table = rel.get("From Table", "")
        fact_col = rel.get("From Column", "")
        dim_table = rel.get("To Table", "")
        dim_col = rel.get("To Column", "")
        fact = tables.get(fact_table, pd.DataFrame())
        dim = tables.get(dim_table, pd.DataFrame())
        status = "Ready"
        rec_notes = []
        fact_col_exists = isinstance(fact, pd.DataFrame) and fact_col in fact.columns
        dim_col_exists = isinstance(dim, pd.DataFrame) and dim_col in dim.columns
        fact_blank = dim_dup = unmatched = 0
        fact_rows = len(fact) if isinstance(fact, pd.DataFrame) else 0
        dim_rows = len(dim) if isinstance(dim, pd.DataFrame) else 0
        if not fact_col_exists:
            status = "Not Ready"
            rec_notes.append("Fact/source column missing")
        if not dim_col_exists:
            status = "Not Ready"
            rec_notes.append("Dimension/lookup column missing")
        if fact_col_exists and dim_col_exists:
            fact_vals = fact[fact_col]
            dim_vals = dim[dim_col]
            fact_blank = int(fact_vals.isna().sum() + (fact_vals.astype(str).str.strip() == "").sum())
            dim_dup = int(dim_vals.astype(str).str.strip().duplicated().sum())
            fact_clean = set(fact_vals.dropna().astype(str).str.strip()) - {""}
            dim_clean = set(dim_vals.dropna().astype(str).str.strip()) - {""}
            unmatched = len(fact_clean - dim_clean)
            if dim_dup > 0:
                status = "Not Ready"
                rec_notes.append("Dimension key is not unique")
            if unmatched > 0 and status == "Ready":
                status = "Needs Review"
                rec_notes.append("Fact has keys not found in dimension")
            if fact_blank > 0 and status == "Ready":
                status = "Needs Review"
                rec_notes.append("Fact has blank relationship keys")
        rows.append({
            "Relationship": f"{fact_table}[{fact_col}] -> {dim_table}[{dim_col}]",
            "Fact Table": fact_table,
            "Fact Column": fact_col,
            "Dimension Table": dim_table,
            "Dimension Column": dim_col,
            "Fact Rows": fact_rows,
            "Dimension Rows": dim_rows,
            "Fact Column Exists": fact_col_exists,
            "Dimension Column Exists": dim_col_exists,
            "Fact Blank Key Count": fact_blank,
            "Dimension Duplicate Key Count": dim_dup,
            "Fact Unmatched Key Count": unmatched,
            "Cardinality": rel.get("Cardinality", "Many-to-one (*:1)"),
            "Cross Filter Direction": rel.get("Cross Filter Direction", "Single"),
            "Active": rel.get("Active", "Yes"),
            "Status": status,
            "Recommendation": "; ".join(rec_notes) if rec_notes else "Ready to create relationship",
        })
    return pd.DataFrame(rows)

def build_powerbi_dataset_tables(analysis, scenario_name_value, export_reason_value, export_mode_value):
    """Build the single full detailed Power BI export. Row 1 is always headers in the writer."""
    processed_df = analysis["df"].copy()
    valid_df = analysis["valid"].copy()
    cost_col = analysis["cost_col"]
    run_id = datetime.now().strftime("RUN_%Y%m%d_%H%M%S")
    scenario_label = scenario_name_value if scenario_name_value else "Not provided"

    fact_rebuild = _add_run_columns(processed_df, run_id, scenario_label)
    fact_valid_rebuild = _add_run_columns(valid_df, run_id, scenario_label)

    machine_dim_cols = [c for c in ["SALES MODEL", "Machine Group", "Machine Family", "Machine Category", "Machine Group Source", "Machine Group Notes"] if c in processed_df.columns]
    dim_machine = processed_df[machine_dim_cols].drop_duplicates().sort_values("SALES MODEL").reset_index(drop=True) if machine_dim_cols else pd.DataFrame()
    dim_machine = _add_run_columns(dim_machine, run_id, scenario_label)
    dim_machine_group = build_dim_machine_group(processed_df, run_id, scenario_label)

    dealer_cols = [c for c in ["Dealer Code", "DEALER", "Region"] if c in processed_df.columns]
    dim_dealer = processed_df[dealer_cols].drop_duplicates().sort_values(dealer_cols).reset_index(drop=True) if dealer_cols else pd.DataFrame()
    dim_dealer = _add_run_columns(dim_dealer, run_id, scenario_label)

    dim_rebuild_type = analysis.get("rebuild_reference", pd.DataFrame()).rename(columns={"Description": "Rebuild Description"})
    dim_rebuild_type = _add_run_columns(dim_rebuild_type, run_id, scenario_label)

    dim_region = analysis.get("region_reference", pd.DataFrame()).rename(columns={"Configured Region": "Region"})
    if not dim_region.empty:
        dim_region["Configured Region Flag"] = True
    dim_region = _add_run_columns(dim_region, run_id, scenario_label)

    years = sorted(processed_df["Service Year"].dropna().astype(int).unique()) if "Service Year" in processed_df.columns else []
    dim_date = pd.DataFrame({"Service Year": years})
    if not dim_date.empty:
        dim_date["Year"] = dim_date["Service Year"]
        dim_date["Year Label"] = dim_date["Service Year"].astype(str)
    dim_date = _add_run_columns(dim_date, run_id, scenario_label)

    fact_exceptions = build_powerbi_exception_rows(processed_df, cost_col, run_id, scenario_label)
    fact_outliers = _add_run_columns(processed_df[processed_df["Outlier"] == True].copy(), run_id, scenario_label) if "Outlier" in processed_df.columns else pd.DataFrame()
    fact_cross_type_outliers = _add_run_columns(processed_df[processed_df.get("Cross-Type Outlier Flag", False) == True].copy(), run_id, scenario_label) if "Cross-Type Outlier Flag" in processed_df.columns else pd.DataFrame()

    fact_machine_summary = _add_run_columns(analysis.get("machine_benchmark_ranking", pd.DataFrame()).copy(), run_id, scenario_label)
    fact_dealer_performance = build_powerbi_dealer_performance(valid_df, processed_df, cost_col, run_id, scenario_label)
    fact_region_performance = build_powerbi_region_performance(valid_df, processed_df, cost_col, run_id, scenario_label)
    fact_rate_coverage = _add_run_columns(analysis.get("dealer_rate_coverage_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    fact_data_quality = _add_run_columns(analysis.get("data_quality_score_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    data_quality_summary = _add_run_columns(analysis.get("data_quality_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    global_ccr_type_avg = _add_run_columns(analysis.get("global_ccr_type_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    region_ccr_type_avg = _add_run_columns(analysis.get("region_ccr_type_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    machine_ccr_type_avg = _add_run_columns(analysis.get("machine_ccr_type_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    machine_group_ccr_type_avg = _add_run_columns(analysis.get("machine_group_ccr_type_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    machine_region_ccr_type_avg = _add_run_columns(analysis.get("machine_region_ccr_type_summary", pd.DataFrame()).copy(), run_id, scenario_label)
    machine_insights_export = _add_run_columns(analysis.get("machine_insights", pd.DataFrame()).copy(), run_id, scenario_label)

    run_metadata = analysis.get("metadata", pd.DataFrame()).copy()
    additional_metadata = pd.DataFrame({"Field": ["Run ID", "Scenario Name", "Export Mode", "Export Reason", "Power BI Export Format", "Power BI Notes", "Generated Timestamp"], "Value": [run_id, scenario_label, "Power BI Dataset Export", export_reason_value, "V18.1 Full Detailed Power BI Model Export", "Power BI is the primary output. All sheets are clean tables with headers on row 1. Long logical table names use shortened Excel sheet names where needed due to Excel sheet-name limits. Cross-type CPT+H exceptions are treated as outliers and audited in Fact_CrossType_Outliers.", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]})
    run_metadata = pd.concat([additional_metadata, run_metadata], ignore_index=True)

    relationship_guide = pd.DataFrame({
        "From Table": [
            "Fact_Rebuild_Rows", "Fact_Rebuild_Rows", "Fact_Rebuild_Rows", "Fact_Rebuild_Rows", "Fact_Rebuild_Rows",
            "Fact_Machine_RebuildType_AvgCost", "Fact_MachineRegion_RebuildType_AvgCost", "Fact_MachineRegion_RebuildType_AvgCost",
            "Fact_MachineGroup_RebuildType_AvgCost", "Fact_Dealer_Performance", "Fact_Region_Performance",
            "Fact_Exception_Rows", "Fact_CrossType_Outliers", "Dim_Machine", "Fact_Rebuild_Rows"
        ],
        "From Column": [
            "SALES MODEL", "Dealer Code", "CCR TYPE", "Region", "Service Year",
            "SALES MODEL", "SALES MODEL", "Region", "Machine Group", "Dealer Code", "Region",
            "SALES MODEL", "SALES MODEL", "Machine Group", "Machine Group"
        ],
        "To Table": [
            "Dim_Machine", "Dim_Dealer", "Dim_Rebuild_Type", "Dim_Region", "Dim_Service_Year",
            "Dim_Machine", "Dim_Machine", "Dim_Region", "Dim_Machine_Group", "Dim_Dealer", "Dim_Region",
            "Dim_Machine", "Dim_Machine", "Dim_Machine_Group", "Dim_Machine_Group"
        ],
        "To Column": [
            "SALES MODEL", "Dealer Code", "CCR TYPE", "Region", "Service Year",
            "SALES MODEL", "SALES MODEL", "Region", "Machine Group", "Dealer Code", "Region",
            "SALES MODEL", "SALES MODEL", "Machine Group", "Machine Group"
        ],
        "Cardinality": ["Many-to-one (*:1)"] * 15,
        "Cross Filter Direction": ["Single"] * 15,
        "Active": ["Yes"] * 15,
        "Relationship Notes": [
            "Core row-level machine relationship", "Core row-level dealer relationship", "Core row-level rebuild type relationship", "Core row-level region relationship", "Core service-year relationship",
            "Machine detail average table relationship", "Machine-region table responds to machine slicers", "Machine-region table responds to region slicers",
            "Group summary table relates to one-row-per-group dimension", "Dealer performance table relationship", "Region performance table relationship",
            "Exception table drill-through relationship", "Cross-type outlier audit relationship", "Optional dimension-to-dimension group relationship", "Optional direct machine-group slicer relationship"
        ]
    })
    dax_starter = build_powerbi_dax_starter(cost_col)


    tables = {
        "Fact_Rebuild_Rows": fact_rebuild,
        "Fact_Valid_Rebuild_Rows": fact_valid_rebuild,
        "Fact_Global_RebuildType_AvgCost": global_ccr_type_avg,
        "Fact_Region_RebuildType_AvgCost": region_ccr_type_avg,
        "Fact_Machine_RebuildType_AvgCost": machine_ccr_type_avg,
        "Fact_MachineGroup_RebuildType_AvgCost": machine_group_ccr_type_avg,
        "Fact_MachineRegion_RebuildType_AvgCost": machine_region_ccr_type_avg,
        "Fact_Machine_Insights": machine_insights_export,
        "Fact_Machine_Ranking": fact_machine_summary,
        "Fact_Dealer_Performance": fact_dealer_performance,
        "Fact_Region_Performance": fact_region_performance,
        "Fact_Exception_Rows": fact_exceptions,
        "Fact_Outlier_Rows": fact_outliers,
        "Fact_CrossType_Outliers": fact_cross_type_outliers,
        "Fact_Rate_Coverage": fact_rate_coverage,
        "Fact_Data_Quality": fact_data_quality,
        "DataQuality_Summary": data_quality_summary,
        "Dim_Machine": dim_machine,
        "Dim_Machine_Group": dim_machine_group,
        "Dim_Dealer": dim_dealer,
        "Dim_Rebuild_Type": dim_rebuild_type,
        "Dim_Region": dim_region,
        "Dim_Service_Year": dim_date,
        "Machine_Grouping": _add_run_columns(analysis.get("machine_grouping_lookup", pd.DataFrame()).copy(), run_id, scenario_label),
        "Run_Metadata": _clean_powerbi_columns(run_metadata),
        "Filter_Summary": _add_run_columns(analysis.get("filter_summary", pd.DataFrame()).copy(), run_id, scenario_label),
        "Relationship_Guide": relationship_guide,
        "PowerBI_Relationship_Checks": pd.DataFrame(),
        "DAX_Starter": dax_starter,
        "PowerBI_Instructions": build_powerbi_instructions_table(),
        "PowerBI_Report_Layout": build_powerbi_report_layout_table(),
        "PowerBI_Table_Dictionary": build_powerbi_table_dictionary(),
        "PowerBI_Sheet_Name_Map": build_powerbi_sheet_name_map(),
        "PowerBI_Pipeline_Guide": build_powerbi_pipeline_guide(),
        "PowerBI_Build_Checklist": build_powerbi_build_checklist(),
        "Export_Mode_Dictionary": build_export_mode_dictionary(),
        "Required_Files": build_required_files_checklist(),
        "Testing_Checklist": build_testing_checklist(),
        "Update_Process": build_update_process_table(),
        "Known_Limitations": _clean_powerbi_columns(analysis.get("known_limitations", pd.DataFrame()).copy()),
        "Data_Dictionary": _clean_powerbi_columns(analysis.get("data_dictionary", pd.DataFrame()).copy()),
        "Parameters": _add_run_columns(analysis.get("parameter_summary", pd.DataFrame()).copy(), run_id, scenario_label),
        "Scenario_Comparison": build_scenario_comparison_table(analysis, run_id, scenario_label),
    }
    tables["PowerBI_Relationship_Checks"] = build_powerbi_relationship_checks(tables, relationship_guide)
    preview = build_powerbi_export_preview(tables, list(tables.keys()))
    readiness = build_powerbi_readiness_score(preview)
    tables["PowerBI_Readiness"] = pd.concat([readiness.assign(Section="Score"), preview.assign(Section="Table Check")], ignore_index=True, sort=False)
    ordered = {name: tables[name] for name in POWERBI_FULL_EXPORT_TABLES if name in tables}
    # Include scenario comparison even if not in primary list for run-to-run analysis.
    ordered["Scenario_Comparison"] = tables["Scenario_Comparison"]
    return ordered


def _strip_powerbi_confidential_marker_rows(df):
    """Remove any accidental label marker rows before writing Power BI tables."""
    out = _clean_powerbi_columns(df)
    if out.empty:
        return out
    first_col = out.columns[0]
    marker_mask = out[first_col].astype(str).str.contains("Confidential|Yellow|Removed Label", case=False, na=False)
    return out.loc[~marker_mask].copy()


def write_powerbi_dataset_workbook(writer, analysis, scenario_name_value, export_reason_value, export_mode_value, selected_tables=None):
    tables = build_powerbi_dataset_tables(analysis, scenario_name_value, export_reason_value, export_mode_value)
    for sheet_name, table in tables.items():
        safe_name = POWERBI_SHEET_NAME_MAP.get(sheet_name, safe_sheet_name(sheet_name))
        clean_table = _strip_powerbi_confidential_marker_rows(table)
        clean_table.to_excel(writer, sheet_name=safe_name, index=False, startrow=0)
        ws = writer.book[safe_name]
        ws.freeze_panes = "A2"


def build_scenario_archive_package(analysis, scenario_name_value, export_reason_value, selected_tables=None):
    """Create a handoff ZIP containing Power BI data, archive summary, and instructions."""
    safe_scenario = re.sub(r"[^A-Za-z0-9_-]+", "_", scenario_name_value.strip()) if scenario_name_value else "Scenario"
    archive_buffer = BytesIO()

    pbi_buffer = BytesIO()
    with pd.ExcelWriter(pbi_buffer, engine="openpyxl") as writer:
        write_powerbi_dataset_workbook(writer, analysis, scenario_name_value, export_reason_value, "Power BI Dataset Export", selected_tables=selected_tables)
    pbi_buffer.seek(0)

    archive_xlsx = BytesIO()
    with pd.ExcelWriter(archive_xlsx, engine="openpyxl") as writer:
        analysis.get("metadata", pd.DataFrame()).to_excel(writer, sheet_name="Run Metadata", index=False)
        analysis.get("filter_summary", pd.DataFrame()).to_excel(writer, sheet_name="Filter Summary", index=False)
        analysis.get("parameter_summary", pd.DataFrame()).to_excel(writer, sheet_name="Parameters", index=False)
        analysis.get("summary", pd.DataFrame()).to_excel(writer, sheet_name="Summary", index=False)
        analysis.get("adjusted_summary", pd.DataFrame()).to_excel(writer, sheet_name="Adjusted Summary", index=False)
        analysis.get("machine_benchmark_ranking", pd.DataFrame()).to_excel(writer, sheet_name="Machine Ranking", index=False)
        analysis.get("global_ccr_type_summary", pd.DataFrame()).to_excel(writer, sheet_name="Global CCR Type Avg", index=False)
        analysis.get("region_ccr_type_summary", pd.DataFrame()).to_excel(writer, sheet_name="Region CCR Type Avg", index=False)
        analysis.get("machine_ccr_type_summary", pd.DataFrame()).to_excel(writer, sheet_name="Machine CCR Type Avg", index=False)
        analysis.get("machine_group_ccr_type_summary", pd.DataFrame()).to_excel(writer, sheet_name="Group CCR Type Avg", index=False)
        analysis.get("machine_insights", pd.DataFrame()).to_excel(writer, sheet_name="Machine Insights", index=False)
        analysis.get("dealer_rate_coverage_summary", pd.DataFrame()).to_excel(writer, sheet_name="Rate Coverage", index=False)
        analysis.get("data_quality_score_summary", pd.DataFrame()).to_excel(writer, sheet_name="Data Quality Score", index=False)
        analysis.get("machine_grouping_lookup", pd.DataFrame()).to_excel(writer, sheet_name="Machine Grouping", index=False)
        build_scenario_comparison_table(analysis, scenario_name_value=scenario_name_value).to_excel(writer, sheet_name="Scenario Comparison", index=False)
        build_powerbi_instructions_table().to_excel(writer, sheet_name="PowerBI Instructions", index=False)
        build_powerbi_report_layout_table().to_excel(writer, sheet_name="PowerBI Layout", index=False)
        build_powerbi_table_dictionary().to_excel(writer, sheet_name="PBI Table Dictionary", index=False)
        build_export_mode_dictionary().to_excel(writer, sheet_name="Export Modes", index=False)
        build_required_files_checklist().to_excel(writer, sheet_name="Required Files", index=False)
        build_testing_checklist().to_excel(writer, sheet_name="Testing Checklist", index=False)
        build_update_process_table().to_excel(writer, sheet_name="Update Process", index=False)
        analysis.get("data_dictionary", pd.DataFrame()).to_excel(writer, sheet_name="Data Dictionary", index=False)
        analysis.get("known_limitations", pd.DataFrame()).to_excel(writer, sheet_name="Known Limitations", index=False)
        apply_excel_brand_formatting(writer.book)
    archive_xlsx.seek(0)

    readme = f"""Rebuild Analytics Scenario Archive
Scenario: {scenario_name_value if scenario_name_value else 'Not provided'}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
App Version: {APP_VERSION}
Export Reason: {export_reason_value}

Files included:
1. Rebuild_Analysis_PowerBI_Dataset_{safe_scenario}.xlsx
   - Clean Power BI-ready tables with headers on row 1.
   - Use Relationship_Guide, DAX_Starter, PowerBI_Instructions, and PowerBI_Report_Layout.

2. Rebuild_Analysis_Scenario_Archive_{safe_scenario}.xlsx
   - Human-readable archive workbook with metadata, filters, summaries, grouping, methodology support tables, and handoff information.

Recommended handoff use:
- Store this ZIP with the source rebuild file and dealer-rate source used for the run.
- If importing to Power BI, use the Power BI dataset workbook rather than the archive workbook.
"""

    with zipfile.ZipFile(archive_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(f"Rebuild_Analysis_PowerBI_Dataset_{safe_scenario}.xlsx", pbi_buffer.getvalue())
        z.writestr(f"Rebuild_Analysis_Scenario_Archive_{safe_scenario}.xlsx", archive_xlsx.getvalue())
        z.writestr("README_Handoff.txt", readme)
    archive_buffer.seek(0)
    return archive_buffer.getvalue()

# =====================================================
# MAIN UI INPUTS
# =====================================================
st.markdown(
    """
    <div class="method-card">
        <div class="method-title">Active Analysis Logic</div>
        <div class="method-body">Cost = PARTS DN USD + Labor Cost USD. Inflation is applied by component. Outliers use log(cost) + IQR by Machine + CCR TYPE. CPT+H cross-type exceptions are treated as outliers, excluded from core averages, and audited in Power BI.</div>
    </div>
    """,
    unsafe_allow_html=True,
)

rebuild_file = st.file_uploader("Upload Rebuild File", type=["xlsx"])
if st.button("Clear uploaded data / reset analysis"):
    reset_app_state()
    st.success("Session analysis state cleared. Refresh the page if you also want to clear the file uploader selection.")
    st.stop()
validate_uploaded_file_security(rebuild_file)
with st.expander("Methodology Lock", expanded=False):
    st.markdown(METHOD_LOCK_TEXT)

st.subheader("Currency Conversion")
auto_currency = st.checkbox("Automatically convert source currency to USD", True)
default_source_currency = st.text_input("Default source currency if no currency column exists", "USD").upper().strip()
dealer_rate_currency_mode = st.selectbox("Dealer labor rate currency", ["USD", "Same as source currency"], index=0)

st.subheader("Dealer Labor Rates")
use_default = st.checkbox("Use Built-in Default Dealer Rates", True)
rate_file = None
rate_file_format = "Auto-detect"
rate_fallback_behavior = "Use yearly average fallback (recommended)"

st.download_button(
    "Download Dealer Rate Template",
    data=create_dealer_rate_template(),
    file_name="Dealer_Rate_Template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

if use_default:
    built_in_rate_file_available = Path("expanded_dealer_base_rate_by_year_2016_2026.xlsx").exists()
    if built_in_rate_file_available:
        st.warning("Built-in expanded dealer base rates 2016-2026 are being used from the bundled workbook. You can still upload a custom dealer labor rate workbook if needed.")
    else:
        st.error("Built-in dealer-rate workbook is missing from the app folder. The app can still run using emergency generic fallback rates, but production analysis should not rely on that fallback.")
else:
    rate_file_format = st.selectbox(
        "Dealer rate file format",
        ["Auto-detect", "Flat Table", "Multi-Sheet Dealer Workbook"],
        index=0,
        help="Recommended format is Flat Table with Dealer Code, Service Year, Rate, optional Rate Currency, and Notes.",
    )
    rate_fallback_behavior = st.selectbox(
        "Missing dealer-year rate fallback behavior",
        [
            "Use yearly average fallback (recommended)",
            "Use overall average fallback",
            "Use built-in default fallback",
            "Stop analysis if dealer-year rate is missing",
        ],
        index=0,
        help="Recommended default uses the uploaded rate file's yearly average when a dealer-year match is missing, and flags the row.",
    )
    rate_file = st.file_uploader("Upload Custom Dealer Labor Rate Workbook", type=["xlsx"])
    validate_uploaded_file_security(rate_file)
    if rate_file:
        try:
            preview_rates = build_rate_table_from_workbook(rate_file, rate_file_format)
            st.write("### Dealer Rate Upload Validation")
            display_table(validate_dealer_rate_table(preview_rates), number_cols=["Value"])
            st.write("### Dealer Rate Preview")
            display_table(preview_rates.head(50), currency_cols=["Rate"], number_cols=["Service Year"])
            st.caption("Preferred template sheet is User Input with columns: Dealer Code, Service Year, Rate, Rate Currency, Notes. Legacy flat and multi-sheet dealer workbooks are still supported.")
        except Exception as exc:
            st.error(f"Unable to parse dealer rate workbook: {exc}")
            st.stop()

st.subheader("Machine Grouping")
machine_group_file = None
st.download_button(
    "Download Machine Grouping Template",
    data=create_machine_grouping_template(),
    file_name="Machine_Grouping_Template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
machine_group_file = st.file_uploader(
    "Upload Custom Machine Grouping Workbook (optional)",
    type=["xlsx"],
    help="Optional override/enrichment file. Preferred template sheet is User Input with SALES MODEL, Machine Group, Machine Family, Machine Category, and Notes.",
)
validate_uploaded_file_security(machine_group_file)
if machine_group_file:
    try:
        machine_group_preview = parse_machine_grouping_file(machine_group_file)
        st.write("### Machine Grouping Preview")
        if machine_group_preview.empty:
            st.warning("No valid machine grouping rows detected. Required columns: SALES MODEL and Machine Group.")
        else:
            display_table(machine_group_preview.head(50))
    except Exception as exc:
        st.error(f"Unable to parse machine grouping workbook: {exc}")
        st.stop()

apply_inflation = st.checkbox("Apply BLS CPI-U Inflation", True)
base_year = st.number_input("Base Year for Inflation", 2010, 2030, 2026)
machine_input = st.text_input("Machines (optional, comma-separated; leave blank for all)")
start_year = st.number_input("Start Year", 2010, 2030, 2016)
end_year = st.number_input("End Year", 2010, 2030, 2026)
rebuild_filter = st.multiselect("Filter Rebuild Types", options=list(CERTIFIED_REBUILD_TYPES.keys()), default=["CMR", "CPT+H", "CPT-O"])
region_filter = st.multiselect("Filter Regions", options=VALID_REGIONS + ["UNKNOWN", "OTHER"], default=VALID_REGIONS + ["UNKNOWN", "OTHER"])

st.subheader("Scenario, Governance, and Future-Proof Controls")
scenario_name = st.text_input("Analysis Scenario Name", "")
user_role_view = st.selectbox("Role view / intended user type", ["Analyst", "Viewer", "Admin"], index=0, help="Advisory only unless enterprise authentication is added.")
strict_mode = st.checkbox("Strict Mode: stop analysis when configured quality thresholds are not met", False)
with st.expander("Advanced Threshold Configuration", expanded=False):
    st.caption("Defaults preserve the current methodology. Change these only when there is a business-approved reason.")
    cross_type_threshold_multiplier = st.number_input("CPT+H cross-type threshold multiplier", min_value=1.00, max_value=2.00, value=1.10, step=0.01)
    min_cmr_rows_for_benchmark = st.number_input("Minimum valid CMR rows for CPT+H benchmark", min_value=1, max_value=20, value=3, step=1)
    dealer_rate_coverage_warning_threshold = st.number_input("Dealer rate coverage warning threshold %", min_value=0.0, max_value=100.0, value=95.0, step=1.0)
    dealer_rate_coverage_strict_threshold = st.number_input("Dealer rate coverage strict threshold %", min_value=0.0, max_value=100.0, value=80.0, step=1.0)
    data_quality_strict_min_score = st.number_input("Strict Mode minimum data quality score", min_value=0.0, max_value=100.0, value=75.0, step=1.0)
    max_rows_warning_threshold = st.number_input("Large-run row warning threshold", min_value=1000, max_value=250000, value=DEFAULT_MAX_ROWS_WARNING, step=1000)

st.subheader("Export Controls")
export_mode = st.selectbox("Export mode", ["Power BI Dataset Export", "Full Analysis Workbook", "Summary Only", "Exceptions Only", "Dealer Rate Audit", "Scenario Archive Package"], index=0)
powerbi_selected_tables = POWERBI_FULL_EXPORT_TABLES.copy()
if export_mode in ["Power BI Dataset Export", "Scenario Archive Package"]:
    st.info("Power BI is the primary output. The app now produces one full detailed Power BI export with all required fact, dimension, audit, relationship-check, sheet-map, DAX, report-layout, pipeline, build-checklist, and handoff tables. Row 1 is always the header row on every exported sheet.")

export_reason = st.selectbox("Export reason", ["Manager review", "Dealer review", "Cost benchmarking", "Data validation", "Presentation support", "Other"], index=0)
export_reason_other = ""
if export_reason == "Other":
    export_reason_other = st.text_input("Describe export reason", "")
export_reason_final = export_reason_other.strip() if export_reason == "Other" and export_reason_other.strip() else export_reason
if user_role_view == "Viewer" and export_mode not in ["Summary Only", "Power BI Dataset Export"]:
    st.warning("Viewer role view is intended for limited distribution. Export mode has been changed to Summary Only for this session.")
    export_mode = "Summary Only"

if rebuild_file:
    with st.expander("Pre-Run Data Validation", expanded=True):
        profile, issue_df = validate_uploaded_workbook(rebuild_file)
        if profile is not None:
            st.write("### Workbook Profile")
            display_table(profile, number_cols=["Rows"])
        st.write("### Validation Issues")
        if issue_df.empty:
            st.success("No pre-run validation issues detected.")
        else:
            st.warning("Review validation issues before running analysis.")
            st.dataframe(issue_df, use_container_width=True)

if rebuild_file:
    st.write("### Preliminary Run Readiness Check")
    prelim_rows = [{"Readiness Check": "Workbook Uploaded", "Status": "Ready", "Details": "A rebuild workbook is loaded."}, {"Readiness Check": "Dealer Rate Source", "Status": "Ready", "Details": "Built-in expanded rates selected." if use_default else "Custom dealer rate workbook selected. Review upload validation."}]
    display_table(pd.DataFrame(prelim_rows))

if st.button("Run Analysis"):
    st.session_state.run_clicked = True
    st.session_state.analysis = None

# =====================================================
# ANALYSIS CORE
# =====================================================
def run_analysis(rebuild_file, rate_file):
    rebuild = read_rebuild_workbook(rebuild_file)
    frames = []
    for sheet_name, sheet in rebuild.items():
        sheet = sheet.copy()
        existing_ccr = None
        for col in sheet.columns:
            if str(col).upper().strip() in ["CCR TYPE", "CCRTYPE", "CCR_TYPE"]:
                existing_ccr = col
                break
        if existing_ccr:
            sheet["CCR TYPE"] = sheet[existing_ccr].apply(lambda value: normalize_ccr_type(value, sheet_name))
        else:
            sheet["CCR TYPE"] = normalize_ccr_type(sheet_name, sheet_name)
        frames.append(sheet)

    df = pd.concat(frames, ignore_index=True)
    required = ["DELIVERED DATE", "ENROLL DATE", "IN SERVICE DATE", "SALES MODEL", "DEALER", "PARTS DN", "SMU AT REBUILD", "REBUILD WORK HRS"]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    rows_uploaded = len(df)
    df = df[df["CCR TYPE"].isin(rebuild_filter)].copy()
    df["Service Date"] = df["DELIVERED DATE"].fillna(df["ENROLL DATE"]).fillna(df["IN SERVICE DATE"])
    df["Service Year"] = pd.to_datetime(df["Service Date"], errors="coerce").dt.year
    missing_service_date_count = int(df["Service Year"].isna().sum())
    df = df[(df["Service Year"] >= start_year) & (df["Service Year"] <= end_year)].copy()
    if machine_input:
        machine_list = [machine.strip() for machine in machine_input.split(",")]
        df = df[df["SALES MODEL"].isin(machine_list)].copy()
    if df.empty:
        raise ValueError("No rows remain after filters. Adjust machine, year, region, or rebuild type filters.")

    df["Dealer Code"] = df["DEALER"].astype(str).str.upper().str.extract(r"([A-Z][A-Z0-9]{3})")
    missing_dealer_code_count = int(df["Dealer Code"].isna().sum())
    df["Region"] = df["REGION"].apply(normalize_region) if "REGION" in df.columns else "UNKNOWN"
    df["Region Display"] = df["Region"].where(df["Region"].isin(VALID_REGIONS + ["UNKNOWN"]), "OTHER")
    unknown_region_count = int((df["Region Display"] == "OTHER").sum())
    df = df[df["Region Display"].isin(region_filter)].copy()
    if df.empty:
        raise ValueError("No rows remain after region filter.")

    # V16.3: built-in automatic machine grouping with optional custom grouping override.
    df, machine_grouping_lookup = apply_machine_grouping(df, machine_group_file)

    currency_col = detect_currency_column(df)
    if auto_currency and currency_col:
        df["Source Currency"] = df[currency_col].apply(lambda value: normalize_currency_code(value, default_source_currency))
    else:
        df["Source Currency"] = normalize_currency_code(default_source_currency)
    df["Service Year"] = pd.to_numeric(df["Service Year"], errors="coerce").astype("Int64")
    fx_lookup = build_fx_lookup(df["Source Currency"].dropna().unique(), df["Service Year"].dropna().unique())
    df = df.merge(fx_lookup, how="left", left_on=["Source Currency", "Service Year"], right_on=["Currency", "Service Year"])
    df["FX to USD"] = df["FX to USD"].fillna(1.0)
    df["FX Source"] = df["FX Source"].fillna("USD/no FX conversion")
    fallback_fx_count = int((df["FX Source"] == "Embedded fallback FX table").sum())

    if use_default or rate_file is None:
        rate_df = build_default_rate_table(2010, 2030)
        rate_file_mode = "Built-in Expanded Dealer Rates 2016-2026"
        effective_fallback_behavior = "Built-in expanded dealer rates; missing dealer-years use yearly average fallback, then overall average if needed"
    else:
        rate_df = build_rate_table_from_workbook(rate_file, rate_file_format)
        if rate_df.empty:
            raise ValueError("Custom dealer rate workbook did not contain any valid dealer-year rates. Use the downloadable template or select the correct rate format.")
        rate_file_mode = rate_file_format
        effective_fallback_behavior = rate_fallback_behavior

    rate_df["Service Year"] = pd.to_numeric(rate_df["Service Year"], errors="coerce").astype("Int64")
    rate_df["Rate"] = pd.to_numeric(rate_df["Rate"], errors="coerce")
    df = df.merge(rate_df, how="left", on=["Dealer Code", "Service Year"])
    df["Rate Source"] = "Dealer-Year Rate"
    df["Dealer Rate Exception Flag"] = ""
    missing_dealer_year = df["Rate"].isna()
    df.loc[missing_dealer_year, "Dealer Rate Exception Flag"] = "Missing dealer-year labor rate"

    if missing_dealer_year.any() and effective_fallback_behavior == "Stop analysis if dealer-year rate is missing":
        missing_count = int(missing_dealer_year.sum())
        raise ValueError(f"{missing_count} rows are missing dealer-year labor rates. Upload a complete rate file or choose a fallback behavior.")

    if missing_dealer_year.any():
        if effective_fallback_behavior == "Use built-in default fallback":
            default_rates = build_default_rate_table(2010, 2030).set_index("Service Year")["Rate"]
            df.loc[missing_dealer_year, "Rate"] = df.loc[missing_dealer_year, "Service Year"].map(default_rates)
            df.loc[missing_dealer_year, "Rate Source"] = "Built-in Default Fallback Rate"
        elif effective_fallback_behavior == "Use overall average fallback":
            overall_avg = rate_df["Rate"].mean()
            df.loc[missing_dealer_year, "Rate"] = overall_avg
            df.loc[missing_dealer_year, "Rate Source"] = "Overall Average Fallback Rate"
        else:
            yearly_avg = rate_df.groupby("Service Year")["Rate"].mean()
            df.loc[missing_dealer_year, "Rate"] = df.loc[missing_dealer_year, "Service Year"].map(yearly_avg)
            df.loc[missing_dealer_year, "Rate Source"] = "Global Yearly Fallback Rate"
            still_missing = df["Rate"].isna()
            if still_missing.any():
                overall_avg = rate_df["Rate"].mean()
                df.loc[still_missing, "Rate"] = overall_avg
                df.loc[still_missing, "Rate Source"] = "Overall Average Fallback Rate"
                df.loc[still_missing, "Dealer Rate Exception Flag"] = "Missing dealer-year and service-year average labor rate"

    missing_overall = df["Rate"].isna()
    if missing_overall.any():
        fallback_rate = build_default_rate_table(2010, 2030)["Rate"].mean()
        df.loc[missing_overall, "Rate"] = fallback_rate
        df.loc[missing_overall, "Rate Source"] = "Emergency Built-in Overall Fallback Rate"
        df.loc[missing_overall, "Dealer Rate Exception Flag"] = "Emergency fallback labor rate used"

    df["Base Rate Year Used"] = df["Service Year"]
    df["Avg Base Rate Used"] = df["Rate"]

    df["PARTS DN"] = pd.to_numeric(df["PARTS DN"], errors="coerce")
    df["REBUILD WORK HRS"] = pd.to_numeric(df["REBUILD WORK HRS"], errors="coerce")
    df["SMU AT REBUILD"] = pd.to_numeric(df["SMU AT REBUILD"], errors="coerce")
    missing_parts_count = int(df["PARTS DN"].isna().sum())
    missing_labor_hours_count = int(df["REBUILD WORK HRS"].isna().sum())
    df = df[(df["PARTS DN"] > 0) & (df["REBUILD WORK HRS"].notna())].copy()

    df["PARTS DN USD"] = df["PARTS DN"] * df["FX to USD"]
    df["Rate FX to USD"] = df["FX to USD"] if dealer_rate_currency_mode == "Same as source currency" else 1.0
    df["Rate USD"] = df["Rate"] * df["Rate FX to USD"]
    df["Labor Cost USD"] = df["REBUILD WORK HRS"] * df["Rate USD"]
    df["Adjusted Total Cost USD"] = df["PARTS DN USD"] + df["Labor Cost USD"]

    cpi_table, cpi_source = get_cpi_table(start_year, end_year, base_year)
    base_cpi = cpi_table.get(int(base_year), np.nan)
    if apply_inflation and pd.notna(base_cpi):
        df["Service Year CPI"] = df["Service Year"].apply(lambda year: cpi_table.get(int(year), np.nan) if pd.notnull(year) else np.nan)
        df["Inflation Factor"] = (base_cpi / df["Service Year CPI"]).fillna(1.0)
    else:
        df["Service Year CPI"] = np.nan
        df["Inflation Factor"] = 1.0
    df["Inflation-Adjusted Parts DN USD"] = df["PARTS DN USD"] * df["Inflation Factor"]
    df["Inflation-Adjusted Base Rate USD"] = df["Rate USD"] * df["Inflation Factor"]
    df["Inflation-Adjusted Labor Cost USD"] = df["REBUILD WORK HRS"] * df["Inflation-Adjusted Base Rate USD"]
    df["Inflation-Adjusted Adjusted Total Cost USD"] = df["Inflation-Adjusted Parts DN USD"] + df["Inflation-Adjusted Labor Cost USD"]
    cost_col = "Inflation-Adjusted Adjusted Total Cost USD" if apply_inflation else "Adjusted Total Cost USD"
    df = df[df[cost_col] > 0].copy()
    if df.empty:
        raise ValueError("No valid cost rows remain after processing.")

    df["Data Quality Exception Flag"] = ""
    df.loc[df["SMU AT REBUILD"].isin([0, 1]), "Data Quality Exception Flag"] = "SMU 0 or 1"
    df["Statistical Cost Outlier Flag"] = False
    df["Cross-Type Outlier Flag"] = False
    df["Outlier"] = False
    df["Outlier Rule Type"] = "Not Outlier"
    df["Outlier Reason"] = ""
    df["Excluded From Core Averages"] = False
    df["Excluded From Power BI Average Tables"] = False
    df["Insufficient Sample Group"] = False
    for (machine, ccr_type), group in df.groupby(["SALES MODEL", "CCR TYPE"], dropna=False):
        idx = group.index
        n = len(group)
        if n < 5:
            df.loc[idx, "Insufficient Sample Group"] = True
            df.loc[idx, "Outlier Rule Type"] = "Insufficient Sample - Not Removed"
            df.loc[idx, "Outlier Reason"] = "Insufficient sample; no statistical outlier removal"
            continue
        log_vals = np.log(group[cost_col])
        q1, q3 = np.percentile(log_vals, [25, 75])
        iqr = q3 - q1
        mult = 2.0 if n < 8 else 1.5
        low = q1 - mult * iqr
        high = q3 + mult * iqr
        mask = (log_vals < low) | (log_vals > high)
        outlier_idx = group.index[mask]
        df.loc[outlier_idx, "Statistical Cost Outlier Flag"] = True
        df.loc[outlier_idx, "Outlier Reason"] = "Cost outlier: log(cost) IQR rule"

    df["Cross-Type Exception Flag"] = ""
    df["CMR Benchmark Cost"] = np.nan
    df["Cross-Type Threshold Cost"] = np.nan
    for machine, group in df.groupby("SALES MODEL", dropna=False):
        cmr_valid = group[(group["CCR TYPE"] == "CMR") & (group["Statistical Cost Outlier Flag"] == False)]
        if len(cmr_valid) >= int(min_cmr_rows_for_benchmark):
            benchmark = cmr_valid[cost_col].median()
            threshold = benchmark * float(cross_type_threshold_multiplier)
            df.loc[group.index, "CMR Benchmark Cost"] = benchmark
            df.loc[group.index, "Cross-Type Threshold Cost"] = threshold
            mask = group["CCR TYPE"].eq("CPT+H") & group["Statistical Cost Outlier Flag"].eq(False) & (group[cost_col] > threshold)
            cross_idx = group.index[mask]
            df.loc[cross_idx, "Cross-Type Outlier Flag"] = True
            df.loc[cross_idx, "Cross-Type Exception Flag"] = "CPT+H Cost Above Typical CMR - Treated as Outlier"

    both = df["Statistical Cost Outlier Flag"] & df["Cross-Type Outlier Flag"]
    stat_only = df["Statistical Cost Outlier Flag"] & ~df["Cross-Type Outlier Flag"]
    cross_only = df["Cross-Type Outlier Flag"] & ~df["Statistical Cost Outlier Flag"]
    df.loc[stat_only, "Outlier Rule Type"] = "Statistical Cost Outlier"
    df.loc[cross_only, "Outlier Rule Type"] = "Cross-Type Cost Outlier"
    df.loc[both, "Outlier Rule Type"] = "Statistical and Cross-Type Cost Outlier"
    df.loc[cross_only, "Outlier Reason"] = "CPT+H cost above machine-level CMR benchmark threshold"
    df.loc[both, "Outlier Reason"] = "Cost outlier and CPT+H cross-type benchmark outlier"
    df["Outlier"] = df["Statistical Cost Outlier Flag"] | df["Cross-Type Outlier Flag"]
    df["Excluded From Core Averages"] = df["Outlier"]
    df["Excluded From Power BI Average Tables"] = df["Outlier"]

    valid = df[df["Outlier"] == False].copy()
    if valid.empty:
        raise ValueError("All rows were excluded as outliers. Review data or filters.")

    summary = valid.groupby("CCR TYPE").agg(Avg_Cost=(cost_col, "mean"), Avg_SMU=("SMU AT REBUILD", "mean"), Count=(cost_col, "count")).reset_index()
    summary = add_vs_cmr(summary)
    summary["Sample Confidence"] = summary["Count"].apply(sample_confidence)
    adjusted_valid = valid.copy()
    adjusted_summary = pd.DataFrame(columns=["CCR TYPE", "Adjusted_Avg_Cost", "Count", "Sample Confidence"])
    global_ccr_type_summary = build_combined_global_ccr_type_summary(valid, adjusted_valid, cost_col, processed_df=df)
    region_ccr_type_summary = build_combined_region_ccr_type_summary(valid, adjusted_valid, cost_col, global_ccr_type_summary, processed_df=df)
    machine_ccr_type_summary = build_combined_machine_ccr_type_summary(valid, adjusted_valid, cost_col, global_ccr_type_summary, processed_df=df)
    machine_group_ccr_type_summary = build_combined_machine_group_ccr_type_summary(valid, adjusted_valid, cost_col, global_ccr_type_summary, processed_df=df)
    machine_region_ccr_type_summary = build_machine_region_ccr_type_summary(valid, df, cost_col, global_ccr_type_summary)
    machine_insights = build_machine_insights_table(valid, df, machine_ccr_type_summary, cost_col)
    rebuild_reference = pd.DataFrame([{"CCR TYPE": key, "Description": value} for key, value in CERTIFIED_REBUILD_TYPES.items()])
    region_reference = pd.DataFrame({"Configured Region": VALID_REGIONS})
    metadata = pd.DataFrame({"Field": ["App Version", "Run Timestamp", "Rows Uploaded", "Rows After Filters", "Valid Rows", "Start Year", "End Year", "Base Year", "Machine Filter", "Rebuild Type Filter", "Region Filter", "Default Source Currency", "Dealer Rate Currency Mode", "Currency Column Detected", "BLS CPI Source", "FX Source", "Analysis Cost Used", "Dealer Rate Mode", "Dealer Rate Format", "Dealer Rate Fallback Behavior", "Dealer Rate Rows", "Dealer Rate Unique Dealers", "Scenario Name", "User Role View", "Strict Mode", "Methodology Version", "Outlier Rule Version", "Dealer Rate Version", "Security Control Version", "Export Format Version", "CPT+H Threshold Multiplier", "Minimum CMR Benchmark Rows", "Dealer Rate Coverage Warning Threshold %", "Dealer Rate Coverage Strict Threshold %", "Data Quality Strict Minimum Score", "Large-Run Row Warning Threshold"], "Value": [APP_VERSION, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), rows_uploaded, len(df), len(valid), start_year, end_year, base_year, machine_input if machine_input else "All", ", ".join(rebuild_filter), ", ".join(region_filter), default_source_currency, dealer_rate_currency_mode, currency_col if currency_col else "None", cpi_source, "Frankfurter annual average; embedded fallback if unavailable", cost_col, "Built-in Expanded 2016-2026" if use_default or rate_file is None else "Uploaded Custom", rate_file_mode, effective_fallback_behavior, len(rate_df), rate_df["Dealer Code"].nunique(), scenario_name if scenario_name else "Not provided", user_role_view, "Yes" if strict_mode else "No", METHODOLOGY_VERSION, OUTLIER_RULE_VERSION, DEALER_RATE_VERSION, SECURITY_CONTROL_VERSION, EXPORT_FORMAT_VERSION, cross_type_threshold_multiplier, min_cmr_rows_for_benchmark, dealer_rate_coverage_warning_threshold, dealer_rate_coverage_strict_threshold, data_quality_strict_min_score, max_rows_warning_threshold]})
    
    if "Field" in metadata.columns:
        metadata = metadata[~metadata["Field"].astype(str).str.contains("Confidential|Removed Label", case=False, na=False)].copy()
    if "Value" in metadata.columns:
        metadata = metadata[~metadata["Value"].astype(str).str.contains("Confidential|Yellow", case=False, na=False)].copy()
    data_quality_summary = pd.DataFrame({"Metric": ["Missing Service Date", "Missing Dealer Code", "Missing Parts DN", "Missing Labor Hours", "SMU 0 or 1", "Unknown/Other Regions", "Rows Using Fallback FX", "Rows Using Global Yearly Fallback Rate", "Rows Using Overall Average Fallback Rate", "Rows Using Built-in Default Fallback Rate", "Dealer Rate Exception Rows"], "Value": [missing_service_date_count, missing_dealer_code_count, missing_parts_count, missing_labor_hours_count, int(df["Data Quality Exception Flag"].eq("SMU 0 or 1").sum()), unknown_region_count, fallback_fx_count, int((df["Rate Source"] == "Global Yearly Fallback Rate").sum()), int((df["Rate Source"] == "Overall Average Fallback Rate").sum()), int((df["Rate Source"] == "Built-in Default Fallback Rate").sum()), int((df["Dealer Rate Exception Flag"] != "").sum())]})
    dealer_rate_coverage_summary = build_dealer_rate_coverage_summary(df)
    data_quality_score_summary = build_data_quality_score_summary(df, outlier_count=int(df["Outlier"].sum()), insufficient_count=int(df["Insufficient Sample Group"].sum()))
    run_readiness_summary = build_run_readiness_summary(df, dealer_rate_coverage_summary, data_quality_score_summary)
    parameter_summary = build_parameter_summary_table()
    known_limitations = build_known_limitations_table()
    data_dictionary = build_data_dictionary_table()
    role_policy = build_role_policy_table()
    performance_safeguards = build_performance_safeguard_summary(rows_uploaded, len(df))
    evaluate_strict_mode(run_readiness_summary, dealer_rate_coverage_summary, data_quality_score_summary, len(df))
    show_adjusted_cpth = False
    machine_benchmark_ranking = build_machine_benchmark_ranking(valid, df, cost_col)
    executive_narrative = build_executive_narrative(valid, summary, cost_col, int(df["Outlier"].sum()), int(df["Cross-Type Outlier Flag"].sum()), dealer_rate_coverage_summary, data_quality_score_summary, show_adjusted_cpth)

    return {"df": df, "valid": valid, "adjusted_valid": adjusted_valid, "summary": summary, "adjusted_summary": adjusted_summary, "cost_col": cost_col, "cpi_table": cpi_table, "cpi_source": cpi_source, "base_cpi": base_cpi, "fx_lookup": fx_lookup, "currency_col": currency_col, "rebuild_reference": rebuild_reference, "region_reference": region_reference, "metadata": metadata, "data_quality_summary": data_quality_summary, "outlier_count": int(df["Outlier"].sum()), "cross_count": int(df["Cross-Type Outlier Flag"].sum()), "dq_count": int(df["Data Quality Exception Flag"].eq("SMU 0 or 1").sum()), "insufficient_count": int(df["Insufficient Sample Group"].sum()), "global_year_fallback_count": int((df["Rate Source"] == "Global Yearly Fallback Rate").sum()), "overall_fallback_count": int((df["Rate Source"] == "Overall Average Fallback Rate").sum()), "rate_df": rate_df, "dealer_rate_validation": validate_dealer_rate_table(rate_df, start_year, end_year), "dealer_rate_exception_rows": df[df["Dealer Rate Exception Flag"] != ""].copy(), "rate_file_mode": rate_file_mode, "effective_fallback_behavior": effective_fallback_behavior, "dealer_rate_coverage_summary": dealer_rate_coverage_summary, "data_quality_score_summary": data_quality_score_summary, "run_readiness_summary": run_readiness_summary, "show_adjusted_cpth": show_adjusted_cpth, "machine_benchmark_ranking": machine_benchmark_ranking, "executive_narrative": executive_narrative, "parameter_summary": parameter_summary, "known_limitations": known_limitations, "data_dictionary": data_dictionary, "role_policy": role_policy, "performance_safeguards": performance_safeguards, "global_ccr_type_summary": global_ccr_type_summary, "region_ccr_type_summary": region_ccr_type_summary, "machine_ccr_type_summary": machine_ccr_type_summary, "machine_group_ccr_type_summary": machine_group_ccr_type_summary, "machine_region_ccr_type_summary": machine_region_ccr_type_summary, "machine_insights": machine_insights, "machine_grouping_lookup": machine_grouping_lookup, "filter_summary": build_filter_summary_table()}

# =====================================================
# RENDER RESULTS
# =====================================================
if st.session_state.run_clicked and rebuild_file:
    if st.session_state.analysis is None:
        try:
            st.session_state.analysis = run_analysis(rebuild_file, rate_file)
        except Exception as exc:
            st.error(str(exc))
            st.stop()

    analysis = st.session_state.analysis
    df = analysis["df"]
    valid = analysis["valid"]
    summary = analysis["summary"]
    adjusted_summary = analysis["adjusted_summary"]
    cost_col = analysis["cost_col"]
    cpi_table = analysis["cpi_table"]
    cpi_source = analysis["cpi_source"]
    base_cpi = analysis["base_cpi"]
    fx_lookup = analysis["fx_lookup"]
    currency_col = analysis["currency_col"]
    rebuild_reference = analysis["rebuild_reference"]
    region_reference = analysis["region_reference"]
    metadata = analysis["metadata"]
    data_quality_summary = analysis["data_quality_summary"]
    outlier_count = analysis["outlier_count"]
    cross_count = analysis["cross_count"]
    dq_count = analysis["dq_count"]
    insufficient_count = analysis["insufficient_count"]
    global_year_fallback_count = analysis["global_year_fallback_count"]
    overall_fallback_count = analysis["overall_fallback_count"]
    rate_df = analysis.get("rate_df", pd.DataFrame())
    dealer_rate_validation = analysis.get("dealer_rate_validation", pd.DataFrame())
    dealer_rate_exception_rows = analysis.get("dealer_rate_exception_rows", pd.DataFrame())
    rate_file_mode = analysis.get("rate_file_mode", "Unknown")
    effective_fallback_behavior = analysis.get("effective_fallback_behavior", "Unknown")
    dealer_rate_coverage_summary = analysis.get("dealer_rate_coverage_summary", pd.DataFrame())
    data_quality_score_summary = analysis.get("data_quality_score_summary", pd.DataFrame())
    run_readiness_summary = analysis.get("run_readiness_summary", pd.DataFrame())
    show_adjusted_cpth = analysis.get("show_adjusted_cpth", False)
    machine_benchmark_ranking = analysis.get("machine_benchmark_ranking", pd.DataFrame())
    executive_narrative = analysis.get("executive_narrative", [])
    parameter_summary = analysis.get("parameter_summary", pd.DataFrame())
    known_limitations = analysis.get("known_limitations", pd.DataFrame())
    data_dictionary = analysis.get("data_dictionary", pd.DataFrame())
    role_policy = analysis.get("role_policy", pd.DataFrame())
    performance_safeguards = analysis.get("performance_safeguards", pd.DataFrame())
    global_ccr_type_summary = analysis.get("global_ccr_type_summary", pd.DataFrame())
    region_ccr_type_summary = analysis.get("region_ccr_type_summary", pd.DataFrame())
    machine_ccr_type_summary = analysis.get("machine_ccr_type_summary", pd.DataFrame())
    machine_group_ccr_type_summary = analysis.get("machine_group_ccr_type_summary", pd.DataFrame())
    machine_region_ccr_type_summary = analysis.get("machine_region_ccr_type_summary", pd.DataFrame())
    machine_insights = analysis.get("machine_insights", pd.DataFrame())
    filter_summary = analysis.get("filter_summary", pd.DataFrame())
    machine_grouping_lookup = analysis.get("machine_grouping_lookup", pd.DataFrame())

    tabs = st.tabs(["Executive Dashboard", "Machine Detail", "Dealer Performance", "Region Performance", "Exceptions & Data Quality", "Executive Insights", "Power BI Readiness", "How to Use", "Methodology", "Governance & Dictionary", "Reference"])
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11 = tabs

    with tab1:
        st.subheader("Run Summary")
        display_table(metadata)
        st.write("### Active Filter Summary")
        if not filter_summary.empty:
            fcols = st.columns(4)
            filter_lookup = dict(zip(filter_summary["Filter"], filter_summary["Value"]))
            fcols[0].metric("Scenario", str(filter_lookup.get("Scenario Name", "Not provided"))[:40])
            fcols[1].metric("Years", str(filter_lookup.get("Years", "")))
            fcols[2].metric("Machines", str(filter_lookup.get("Machines", "All"))[:40])
            fcols[3].metric("Inflation", str(filter_lookup.get("Inflation Applied", "")))
            display_table(filter_summary)
        st.write("### Run Readiness Check")
        display_table(run_readiness_summary)
        st.write("### Saved Parameter Summary")
        display_table(parameter_summary)
        st.write("### Performance Safeguards")
        display_table(performance_safeguards)
        st.write("### Dealer Rate Coverage Score")
        display_table(dealer_rate_coverage_summary)
        st.write("### Data Quality Score")
        display_table(data_quality_score_summary)
        st.subheader("Executive Summary")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Valid Rows", f"{len(valid):,}")
        col2.metric("Cost Outliers", f"{outlier_count:,}")
        col3.metric("Avg USD Analysis Cost", money(valid[cost_col].mean()))
        col4.metric("Cross-Type Outliers", f"{cross_count:,}")
        st.write("### Combined Global CCR Type Average Cost")
        display_table(global_ccr_type_summary, currency_cols=["Avg_Cost", "Global CCR Avg Cost"], percent_cols=["Vs CMR %", "Vs Global CCR Avg %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
        st.info("V18 methodology reports one CPT+H value. Cross-type CPT+H exceptions are treated as outliers, excluded from this table, and audited separately in Power BI.")
        st.write("### Machine Group Average Cost by Rebuild Type")
        display_table(machine_group_ccr_type_summary, currency_cols=["Avg_Cost", "Group CMR Avg Cost", "Global CCR Avg Cost"], percent_cols=["Vs Group CMR %", "Vs Global CCR Avg %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
        if not machine_group_ccr_type_summary.empty:
            group_options = sorted(machine_group_ccr_type_summary["Machine Group"].dropna().unique().tolist())
            selected_group_for_chart = st.selectbox("Machine Group Rebuild-Type Chart", group_options, key="machine_group_ccr_chart_selector")
            group_chart_data = machine_group_ccr_type_summary[machine_group_ccr_type_summary["Machine Group"] == selected_group_for_chart].copy()
            cat_rebuild_type_bar_chart(group_chart_data, value_col="Avg_Cost")
        st.write("### Machine Benchmark Ranking")
        display_table(machine_benchmark_ranking, currency_cols=["Avg_Cost"], percent_cols=["Outlier Rate %", "Dealer Rate Exception Rate %", "Vs Overall Avg %"], smu_cols=["Avg_SMU"], number_cols=["Valid_Rows", "Total_Rows", "Outlier_Rows", "Cross_Type_Flags", "Dealer_Rate_Exception_Rows", "Priority Score"])
        st.write("### SMU vs USD Analysis Cost by Rebuild Type")
        scatter = valid[["SMU AT REBUILD", cost_col, "CCR TYPE"]].dropna().rename(columns={"SMU AT REBUILD": "SMU", cost_col: "Cost"})
        cat_scatter_chart(scatter, "SMU", "Cost", "CCR TYPE", tooltip=["CCR TYPE", "SMU", "Cost"])
        st.write("### Currency and Inflation Settings")
        settings = pd.DataFrame({"Setting": ["Currency Column Detected", "Default Source Currency", "Dealer Rate Currency Mode", "FX Source", "CPI Source", "BLS Series", "Inflation Applied", "Base Year", "Base Year CPI", "Cost Source", "Labor Cost", "PLUS PARTS DN", "Analysis Cost Used"], "Value": [str(currency_col) if currency_col else "None - default source currency used", default_source_currency, dealer_rate_currency_mode, "Frankfurter annual average; embedded fallback if unavailable", cpi_source, "CUUR0000SA0 - CPI-U All Items, U.S. city average, not seasonally adjusted", "Yes" if apply_inflation else "No", str(base_year), f"{base_cpi:,.3f}" if pd.notna(base_cpi) else "N/A", "PARTS DN USD + Labor Cost USD", "Labor Hours × Dealer Service-Year Rate converted to USD where applicable", "Ignored entirely", cost_col]})
        st.dataframe(settings, use_container_width=True)

    with tab2:
        st.subheader("Machine Detail")
        machine_list = sorted(valid["SALES MODEL"].dropna().unique())
        selected_machine = st.selectbox("Select Machine", machine_list, key="machine_detail_selector")
        dfm = valid[valid["SALES MODEL"] == selected_machine].copy()
        dfm_all = df[df["SALES MODEL"] == selected_machine].copy()
        st.write(f"### Machine: {selected_machine}")
        mcol1, mcol2, mcol3, mcol4 = st.columns(4)
        machine_show_adjusted_cpth = False
        mcol1.metric("Reported Avg Cost", money(dfm[cost_col].mean()))
        mcol2.metric("Statistical Outliers", f"{int(dfm_all.get('Statistical Cost Outlier Flag', pd.Series(dtype=bool)).sum()):,}")
        mcol3.metric("Cross-Type Outliers", f"{int(dfm_all.get('Cross-Type Outlier Flag', pd.Series(dtype=bool)).sum()):,}")
        mcol4.metric("Total Outliers", f"{int(dfm_all['Outlier'].sum()):,}")

        machine_export_bytes = build_machine_export(selected_machine, analysis)
        safe_machine = str(selected_machine).replace("/", "-").replace("\\", "-").replace(" ", "_")
        if render_export_acknowledgement("machine_export_ack"):
            st.download_button(
                "Download Selected Machine Workbook",
                data=machine_export_bytes,
                file_name=f"Rebuild_Analysis_{safe_machine}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("Confirm export authorization to enable the selected-machine workbook download.")

        machine_type_summary = dfm.groupby("CCR TYPE").agg(Avg_Cost=(cost_col, "mean"), Avg_SMU=("SMU AT REBUILD", "mean"), Count=(cost_col, "count")).reset_index()
        machine_type_summary = add_vs_cmr(machine_type_summary)
        machine_type_summary["Sample Confidence"] = machine_type_summary["Count"].apply(sample_confidence)
        st.write("### Combined Machine CCR Type Average Cost")
        selected_machine_ccr = machine_ccr_type_summary[machine_ccr_type_summary["SALES MODEL"] == selected_machine].copy() if not machine_ccr_type_summary.empty else pd.DataFrame()
        display_table(selected_machine_ccr, currency_cols=["Avg_Cost", "Machine CMR Avg Cost", "Global CCR Avg Cost"], percent_cols=["Vs Machine CMR %", "Vs Global CCR Avg %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
        if not selected_machine_ccr.empty:
            cat_rebuild_type_bar_chart(selected_machine_ccr, value_col="Avg_Cost")
        st.write("### Region Breakdown by Rebuild Type (Shown Once for Selected Machine)")
        selected_machine_region_ccr = machine_region_ccr_type_summary[machine_region_ccr_type_summary["SALES MODEL"] == selected_machine].copy() if not machine_region_ccr_type_summary.empty else pd.DataFrame()
        if selected_machine_region_ccr.empty:
            st.info("No regional rebuild-type breakdown available for this machine.")
        else:
            display_table(selected_machine_region_ccr, currency_cols=["Avg_Cost", "Regional CMR Avg Cost", "Global CCR Avg Cost"], percent_cols=["Vs Regional CMR %", "Vs Global CCR Avg %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
            for region_name in selected_machine_region_ccr["Region"].dropna().unique():
                st.write(f"#### {region_name}")
                region_chart_data = selected_machine_region_ccr[selected_machine_region_ccr["Region"] == region_name].copy()
                cat_rebuild_type_bar_chart(region_chart_data, value_col="Avg_Cost", height=300)
        selected_machine_insights = machine_insights[machine_insights["SALES MODEL"] == selected_machine].copy() if not machine_insights.empty else pd.DataFrame()
        st.write("### Machine-Level Insights")
        display_table(selected_machine_insights, currency_cols=["Metric Value"])
        st.info("One CPT+H value is reported for this machine. Cross-type CPT+H outliers are excluded and available for audit.")

        st.write("### Machine SMU vs USD Analysis Cost")
        machine_chart = dfm[["SMU AT REBUILD", cost_col, "CCR TYPE"]].dropna().rename(columns={"SMU AT REBUILD": "SMU", cost_col: "Cost"})
        cat_scatter_chart(machine_chart, "SMU", "Cost", "CCR TYPE", tooltip=["CCR TYPE", "SMU", "Cost"])

        st.write("### Cross-Type Outlier Audit")
        benchmark_value = dfm_all["CMR Benchmark Cost"].dropna().iloc[0] if not dfm_all["CMR Benchmark Cost"].dropna().empty else np.nan
        machine_cross = dfm_all[dfm_all.get("Cross-Type Outlier Flag", False) == True]
        st.dataframe(pd.DataFrame({"Metric": ["CMR Benchmark Cost", "CPT+H Cross-Type Outliers", "Machine Outliers"], "Value": [money(benchmark_value) if not pd.isna(benchmark_value) else "N/A", f"{len(machine_cross):,}", f"{int(dfm_all['Outlier'].sum()):,}"]}), use_container_width=True)
        if not machine_cross.empty:
            display_table(machine_cross[["DEALER", "Region", "CCR TYPE", "SMU AT REBUILD", cost_col, "CMR Benchmark Cost", "Cross-Type Exception Flag"]], currency_cols=[cost_col, "CMR Benchmark Cost"], smu_cols=["SMU AT REBUILD"])
        else:
            st.write("No CPT+H cross-type outliers for this machine.")

        st.write("### Outlier Performance")
        outlier_perf = dfm_all.groupby("CCR TYPE").agg(Total_Rows=(cost_col, "count"), Outlier_Rows=("Outlier", "sum"), Avg_Cost_All=(cost_col, "mean")).reset_index()
        valid_counts = dfm.groupby("CCR TYPE").size().reset_index(name="Valid_Rows")
        outlier_perf = outlier_perf.merge(valid_counts, on="CCR TYPE", how="left").fillna({"Valid_Rows": 0})
        outlier_perf["Outlier Rate %"] = (outlier_perf["Outlier_Rows"] / outlier_perf["Total_Rows"]) * 100
        display_table(outlier_perf, currency_cols=["Avg_Cost_All"], percent_cols=["Outlier Rate %"], number_cols=["Total_Rows", "Outlier_Rows", "Valid_Rows"])

    with tab3:
        st.subheader("Dealer Analysis")
        dealer_machine_options = ["All Machines"] + sorted(valid["SALES MODEL"].dropna().unique().tolist())
        dealer_machine = st.selectbox("Focus Dealer Analysis on Machine", dealer_machine_options, key="dealer_machine_selector")
        dealer_df = valid.copy() if dealer_machine == "All Machines" else valid[valid["SALES MODEL"] == dealer_machine]
        dealer_full_df = df.copy() if dealer_machine == "All Machines" else df[df["SALES MODEL"] == dealer_machine]
        dealer_summary = dealer_performance_for_df(dealer_df, dealer_full_df, cost_col)
        st.write(f"### Dealer Summary - {dealer_machine}")
        display_table(dealer_summary, currency_cols=["Avg_Cost"], percent_cols=["Vs Section Avg %", "Outlier Rate %", "DQ Rate %", "Cross Flag Rate %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Cross_Type_Flags", "Total_Rows", "Outlier_Rows", "DQ_Rows", "Performance Score"])
        dealer_type = dealer_df.groupby(["DEALER", "CCR TYPE"]).agg(Avg_Cost=(cost_col, "mean"), Avg_SMU=("SMU AT REBUILD", "mean"), Count=(cost_col, "count")).reset_index().sort_values(["DEALER", "CCR TYPE"])
        st.write("### Dealer by Rebuild Type")
        display_table(dealer_type, currency_cols=["Avg_Cost"], smu_cols=["Avg_SMU"], number_cols=["Count"])
        st.write("### Dealer Average Cost Chart")
        if not dealer_summary.empty:
            cat_bar_chart(dealer_summary.rename(columns={"Avg_Cost": "Average Cost"}), "DEALER", "Average Cost", None, tooltip=["DEALER", "Average Cost"])
        else:
            st.info("No dealer summary data available.")

    with tab4:
        st.subheader("Region Analysis")
        region_machine_options = ["All Machines"] + sorted(valid["SALES MODEL"].dropna().unique().tolist())
        region_machine = st.selectbox("Focus Region Analysis on Machine", region_machine_options, key="region_machine_selector")
        region_df = valid.copy() if region_machine == "All Machines" else valid[valid["SALES MODEL"] == region_machine]
        region_summary = region_performance_for_df(region_df, cost_col)
        st.write(f"### Region Summary - {region_machine}")
        display_table(region_summary, currency_cols=["Avg_Cost"], percent_cols=["Vs Section Avg %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Cross_Type_Flags"])
        region_type = region_ccr_type_summary.copy()
        if region_machine != "All Machines":
            region_adjusted_valid = region_df[~((region_df["CCR TYPE"] == "CPT+H") & (region_df["Cross-Type Exception Flag"] == "CPT+H Cost Above Typical CMR"))].copy()
            region_type = build_combined_region_ccr_type_summary(region_df, region_adjusted_valid, cost_col, global_ccr_type_summary)
        st.write("### Region by Rebuild Type")
        display_table(region_type, currency_cols=["Avg_Cost", "Regional CMR Avg Cost", "Global CCR Avg Cost"], percent_cols=["Vs Regional CMR %", "Vs Global CCR Avg %"], smu_cols=["Avg_SMU"], number_cols=["Count", "Statistical Outliers Excluded", "Cross-Type Outliers Excluded", "Total Outliers Excluded"])
        st.write("### Separate Region Charts by Rebuild Type")
        if region_type.empty:
            st.info("No region rebuild-type chart data available.")
        else:
            for region_name in region_type["Region"].dropna().unique():
                st.write(f"#### {region_name}")
                region_chart_data = region_type[region_type["Region"] == region_name].copy()
                cat_rebuild_type_bar_chart(region_chart_data, value_col="Avg_Cost", height=300)

    with tab5:
        st.subheader("Exception Summary")
        exception_summary = pd.DataFrame({"Metric": ["Cost Outliers", "CPT+H Cross-Type Outliers", "SMU 0 or 1", "Insufficient Sample Rows", "Rows Using Global Yearly Fallback Rate", "Rows Using Overall Average Fallback Rate"], "Value": [outlier_count, cross_count, dq_count, insufficient_count, global_year_fallback_count, overall_fallback_count]})
        st.dataframe(exception_summary, use_container_width=True)
        st.write("### Data Quality Summary")
        st.dataframe(data_quality_summary, use_container_width=True)
        st.write("### Data Quality Score")
        display_table(data_quality_score_summary)
        st.write("### Performance Safeguards")
        display_table(performance_safeguards)
        st.write("### Dealer Rate Coverage Score")
        display_table(dealer_rate_coverage_summary)
        st.write("### Dealer Rate Upload Validation")
        if not dealer_rate_validation.empty:
            display_table(dealer_rate_validation, number_cols=["Value"])
        else:
            st.write("No dealer rate validation table available.")
        st.write("### Dealer Rate Exception Rows")
        if not dealer_rate_exception_rows.empty:
            display_table(dealer_rate_exception_rows, currency_cols=["Rate", "Rate USD", "Labor Cost USD", "Adjusted Total Cost USD", "Inflation-Adjusted Adjusted Total Cost USD"], smu_cols=["SMU AT REBUILD"], year_cols=["Service Year"])
        else:
            st.write("No dealer rate exception rows detected.")
        st.write("### Uploaded Dealer Rates Used")
        if not rate_df.empty:
            display_table(rate_df.head(200), currency_cols=["Rate"], number_cols=["Service Year"])
        else:
            st.write("No uploaded/custom dealer rate table available.")
        st.write("### Outlier Performance by Machine + Rebuild Type")
        outlier_perf_all = df.groupby(["SALES MODEL", "CCR TYPE"]).agg(Total_Rows=(cost_col, "count"), Outlier_Rows=("Outlier", "sum"), Avg_Cost_All=(cost_col, "mean")).reset_index()
        outlier_perf_all["Outlier Rate %"] = (outlier_perf_all["Outlier_Rows"] / outlier_perf_all["Total_Rows"]) * 100
        display_table(outlier_perf_all, currency_cols=["Avg_Cost_All"], percent_cols=["Outlier Rate %"], number_cols=["Total_Rows", "Outlier_Rows"])
        st.write("### Cross-Type Exception Rows")
        cross_rows = valid[valid["Cross-Type Exception Flag"] != ""]
        if not cross_rows.empty:
            display_table(cross_rows[["SALES MODEL", "DEALER", "Region", "CCR TYPE", "SMU AT REBUILD", cost_col, "CMR Benchmark Cost", "Cross-Type Exception Flag"]], currency_cols=[cost_col, "CMR Benchmark Cost"], smu_cols=["SMU AT REBUILD"])
        else:
            st.write("No cross-type exceptions detected.")
        st.write("### Outlier Rows")
        outlier_rows = df[df["Outlier"] == True]
        if not outlier_rows.empty:
            display_table(outlier_rows, currency_cols=["PARTS DN", "PARTS DN USD", "Rate", "Rate USD", "Labor Cost USD", "Adjusted Total Cost USD", "Inflation-Adjusted Adjusted Total Cost USD"], smu_cols=["SMU AT REBUILD"], year_cols=["Service Year"])
        else:
            st.write("No cost outliers detected.")

    with tab6:
        st.subheader("Executive Insights")
        st.write("### Executive Summary Narrative")
        for line in executive_narrative:
            st.write("•", line)
        st.write("### Supporting Insights")
        insights = []
        if not summary.empty:
            top_type = summary.sort_values("Avg_Cost", ascending=False).iloc[0]
            insights.append(f"Highest average rebuild type cost is {top_type['CCR TYPE']} at {money(top_type['Avg_Cost'])}.")
        machine_avg = valid.groupby("SALES MODEL")[cost_col].mean().sort_values(ascending=False)
        if len(machine_avg) > 0:
            insights.append(f"Highest average machine cost is {machine_avg.index[0]} at {money(machine_avg.iloc[0])}.")
        try:
            review_dealers = dealer_summary[dealer_summary["Performance Label"] == "Review Needed"] if "Performance Label" in dealer_summary.columns else pd.DataFrame()
            if not review_dealers.empty:
                insights.append(f"{len(review_dealers)} dealer(s) are marked Review Needed based on score, cost position, outlier rate, cross-type flags, and data quality.")
        except Exception:
            pass
        insights.append(f"Cost outliers excluded from core averages: {outlier_count:,}.")
        insights.append(f"CPT+H cross-type review flags: {cross_count:,}.")
        insights.append(f"Rows using fallback labor rates: {global_year_fallback_count + overall_fallback_count:,}.")
        for insight in insights:
            st.write("•", insight)
        st.write("### Machine-Level Insights")
        display_table(machine_insights, currency_cols=["Metric Value"])

    with tab7:
        st.subheader("Power BI Export Preview & Readiness")
        pbi_tables = build_powerbi_dataset_tables(analysis, scenario_name, export_reason_final, export_mode)
        selected_for_preview = powerbi_selected_tables if export_mode in ["Power BI Dataset Export", "Scenario Archive Package"] else DEFAULT_POWERBI_TABLES_STANDARD
        pbi_preview = build_powerbi_export_preview(pbi_tables, selected_for_preview)
        pbi_readiness = build_powerbi_readiness_score(pbi_preview)
        st.write("### Power BI Readiness Score")
        display_table(pbi_readiness)
        try:
            readiness_label = pbi_readiness.loc[pbi_readiness["Metric"] == "Power BI Readiness", "Value"].iloc[0]
            if readiness_label == "Ready":
                st.success("Power BI export is ready. Headers should load correctly from row 1.")
            elif readiness_label == "Needs Review":
                st.warning("Power BI export is usable, but some selected tables should be reviewed.")
            else:
                st.error("Power BI export is not ready. Review table-level issues before importing.")
        except Exception:
            pass
        st.write("### Included Power BI Tables")
        display_table(pbi_preview, number_cols=["Rows", "Columns", "Blank Headers", "Duplicate Headers", "Marker Rows"])
        st.write("### Machine Grouping Lookup")
        display_table(machine_grouping_lookup)
        st.write("### Scenario Comparison Row")
        display_table(build_scenario_comparison_table(analysis, scenario_name_value=scenario_name), currency_cols=["Avg Cost"], percent_cols=["Outlier Rate %", "Dealer Rate Coverage %"], number_cols=["Valid Rows", "Total Rows", "Outlier Rows", "Cross-Type Flags"])

    with tab8:
        st.subheader("How to Use This App")
        st.markdown("""
        1. Upload the rebuild workbook.
        2. Select the dealer labor-rate source. Built-in expanded 2016–2026 rates are selected by default.
        3. Review validation and readiness checks.
        4. Run the analysis.
        5. Review the Executive Dashboard, dealer rate coverage score, and data quality score.
        6. Use Machine Detail for selected-machine review.
        7. Review Exceptions & Data Quality before sharing results.
        8. If needed, manually change advanced thresholds; defaults preserve the standard methodology.
        9. Use Strict Mode only for formal review where the configured thresholds should stop weak runs.
        10. Choose the least-detailed export mode and provide an export reason. Use Power BI Dataset Export when building Power BI reports, or Scenario Archive Package for handoff.
        11. Use the Power BI Readiness tab to validate selected Power BI tables before export.
            12. Use custom machine grouping when business-approved group names are needed.
            13. Interpret Excel highlights: red = cost outlier, orange = cross-type exception, yellow = insufficient sample group.
        """)

    with tab9:
        st.subheader("Methodology")
        st.markdown(METHOD_LOCK_TEXT)
        st.markdown("""**Visual style:** Charts, checkboxes, filter controls, tabs, and workbook headers use a Caterpillar-inspired black, yellow, and gray palette.  
**Important:** Official Caterpillar logo usage should follow internal brand/asset approval rules.  
**V16.5 update:** Refines rebuild-type charts so CMR, CPT+H Standard, CPT+H Adjusted, and CPT-O appear as separate sections; adds machine-group average cost by rebuild type; and places separate region-by-rebuild-type charts directly below the machine rebuild-type chart.""")

    with tab10:
        st.subheader("Governance & Dictionary")
        st.markdown("""
This section is intended for handoff, auditability, and future maintenance. It explains the meaning of key fields, export tables, roles, required files, testing expectations, and safe update process.
""")
        st.write("### Known Limitations")
        display_table(known_limitations)
        st.write("### Data Dictionary")
        display_table(data_dictionary)
        st.write("### Power BI Table Dictionary")
        display_table(build_powerbi_table_dictionary())
        st.write("### Export Mode Dictionary")
        display_table(build_export_mode_dictionary())
        st.write("### Role-Based Export Design")
        display_table(role_policy)
        st.write("### Required Files Checklist")
        display_table(build_required_files_checklist())
        st.write("### Testing Checklist")
        display_table(build_testing_checklist())
        st.write("### Safe Update Process")
        display_table(build_update_process_table())
        st.write("### Saved Parameter Summary")
        display_table(parameter_summary)
        st.write("### Performance Safeguards")
        display_table(performance_safeguards)

    with tab11:
        st.subheader("Reference")
        st.markdown("""
Use this section to audit supported rebuild types, configured regions, observed values in the current run, machine grouping, dealer-rate source, and Power BI export references.
""")
        st.write("### Certified Rebuild Types")
        st.dataframe(rebuild_reference, use_container_width=True)
        st.write("### Configured Regions")
        st.dataframe(region_reference, use_container_width=True)
        st.write("### Observed Rebuild Types in Current Run")
        st.dataframe(pd.DataFrame({"Observed CCR TYPE": sorted(valid["CCR TYPE"].dropna().unique())}), use_container_width=True)
        st.write("### Observed Regions in Current Run")
        st.dataframe(pd.DataFrame({"Observed Region": sorted(valid["Region"].dropna().unique())}), use_container_width=True)
        st.write("### Machine Grouping Reference")
        display_table(machine_grouping_lookup)
        st.write("### Dealer Labor Rate Source Reference")
        display_table(pd.DataFrame({"Reference Item": ["Dealer Rate Mode", "Dealer Rate Format", "Fallback Behavior", "Dealer Rate Rows", "Unique Dealers"], "Value": ["Built-in Expanded Dealer Rates" if use_default else "Uploaded Custom Dealer Rates", rate_file_mode, effective_fallback_behavior, len(rate_df), rate_df["Dealer Code"].nunique() if not rate_df.empty else 0]}))
        st.write("### CPI / Inflation Reference")
        display_table(pd.DataFrame({"Year": list(cpi_table.keys()), "CPI": list(cpi_table.values())}).sort_values("Year"), year_cols=["Year"], decimal_cols=["CPI"])
        st.write("### FX Reference")
        display_table(fx_lookup, year_cols=["Service Year"], decimal_cols=["FX to USD"])
        st.write("### Power BI Export Table Reference")
        display_table(build_powerbi_table_dictionary())
        st.write("### App Version Reference")
        display_table(pd.DataFrame({"Item": ["App Version", "Methodology Version", "Outlier Rule Version", "Dealer Rate Version", "Security Control Version", "Export Format Version"], "Value": [APP_VERSION, METHODOLOGY_VERSION, OUTLIER_RULE_VERSION, DEALER_RATE_VERSION, SECURITY_CONTROL_VERSION, EXPORT_FORMAT_VERSION]}))

    output = BytesIO()
    export_metadata = pd.concat([
        metadata.copy(),
        pd.DataFrame({"Field": ["Export Mode", "Export Reason", "Scenario Name", "User Role View", "Strict Mode", "Methodology Version", "Outlier Rule Version", "Dealer Rate Version", "Security Control Version", "Export Format Version"], "Value": [export_mode, export_reason_final, scenario_name if scenario_name else "Not provided", user_role_view, "Yes" if strict_mode else "No", METHODOLOGY_VERSION, OUTLIER_RULE_VERSION, DEALER_RATE_VERSION, SECURITY_CONTROL_VERSION, EXPORT_FORMAT_VERSION]}),
    ], ignore_index=True)
    cover_sheet = build_cover_sheet(
        export_metadata,
        run_readiness_summary,
        data_quality_score_summary,
        dealer_rate_coverage_summary,
        export_mode,
        export_reason_final,
        scenario_name_value=scenario_name,
        user_role_view_value=user_role_view,
        strict_mode_value=strict_mode,
    )
    if export_mode == "Power BI Dataset Export":
        # Dedicated clean Power BI writer. Do not write cover page, watermark rows, merged cells,
        # decorative formatting, or normal review-workbook tabs before the dataset tables.
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            write_powerbi_dataset_workbook(writer, analysis, scenario_name, export_reason_final, export_mode, selected_tables=powerbi_selected_tables)
    elif export_mode == "Scenario Archive Package":
        output = BytesIO(build_scenario_archive_package(analysis, scenario_name, export_reason_final, selected_tables=powerbi_selected_tables))
    else:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            cover_sheet.to_excel(writer, sheet_name="Cover Page", index=False)
            export_metadata.to_excel(writer, sheet_name="Run Metadata", index=False)
            run_readiness_summary.to_excel(writer, sheet_name="Run Readiness", index=False)
            dealer_rate_coverage_summary.to_excel(writer, sheet_name="Rate Coverage", index=False)
            data_quality_score_summary.to_excel(writer, sheet_name="Data Quality Score", index=False)
            parameter_summary.to_excel(writer, sheet_name="Parameters", index=False)
            performance_safeguards.to_excel(writer, sheet_name="Performance Checks", index=False)
            if export_mode == "Summary Only":
                summary.to_excel(writer, sheet_name="Summary", index=False)
                if show_adjusted_cpth:
                    adjusted_summary.to_excel(writer, sheet_name="Adjusted Summary", index=False)
                else:
                    pd.DataFrame({"Message": ["Adjusted CPT+H comparison hidden because both CMR and CPT+H are required."]}).to_excel(writer, sheet_name="Adjusted CPT-H Hidden", index=False)
                machine_benchmark_ranking.to_excel(writer, sheet_name="Machine Ranking", index=False)
                global_ccr_type_summary.to_excel(writer, sheet_name="Global CCR Type Avg", index=False)
                region_ccr_type_summary.to_excel(writer, sheet_name="Region CCR Type Avg", index=False)
                machine_ccr_type_summary.to_excel(writer, sheet_name="Machine CCR Type Avg", index=False)
                machine_group_ccr_type_summary.to_excel(writer, sheet_name="Group CCR Type Avg", index=False)
                machine_insights.to_excel(writer, sheet_name="Machine Insights", index=False)
                pd.DataFrame({"Executive Narrative": executive_narrative}).to_excel(writer, sheet_name="Executive Narrative", index=False)
            elif export_mode == "Exceptions Only":
                exception_summary.to_excel(writer, sheet_name="Exceptions", index=False)
                data_quality_summary.to_excel(writer, sheet_name="Data Quality", index=False)
                dealer_rate_exception_rows.to_excel(writer, sheet_name="Rate Exceptions", index=False)
                outlier_perf_all.to_excel(writer, sheet_name="Outlier Performance", index=False)
                cross_rows.to_excel(writer, sheet_name="Cross Type Flags", index=False)
                df[df["Outlier"] == True].to_excel(writer, sheet_name="Outlier Rows", index=False)
            elif export_mode == "Dealer Rate Audit":
                dealer_rate_validation.to_excel(writer, sheet_name="Rate Validation", index=False)
                dealer_rate_coverage_summary.to_excel(writer, sheet_name="Rate Coverage Detail", index=False)
                dealer_rate_exception_rows.to_excel(writer, sheet_name="Rate Exceptions", index=False)
                rate_df.to_excel(writer, sheet_name="Dealer Rates Used", index=False)
                data_quality_summary.to_excel(writer, sheet_name="Data Quality", index=False)
            else:
                summary.to_excel(writer, sheet_name="Summary", index=False)
                if show_adjusted_cpth:
                    adjusted_summary.to_excel(writer, sheet_name="Adjusted Summary", index=False)
                else:
                    pd.DataFrame({"Message": ["Adjusted CPT+H comparison hidden because both CMR and CPT+H are required."]}).to_excel(writer, sheet_name="Adjusted CPT-H Hidden", index=False)
                machine_benchmark_ranking.to_excel(writer, sheet_name="Machine Ranking", index=False)
                global_ccr_type_summary.to_excel(writer, sheet_name="Global CCR Type Avg", index=False)
                region_ccr_type_summary.to_excel(writer, sheet_name="Region CCR Type Avg", index=False)
                machine_ccr_type_summary.to_excel(writer, sheet_name="Machine CCR Type Avg", index=False)
                machine_group_ccr_type_summary.to_excel(writer, sheet_name="Group CCR Type Avg", index=False)
                machine_insights.to_excel(writer, sheet_name="Machine Insights", index=False)
                pd.DataFrame({"Executive Narrative": executive_narrative}).to_excel(writer, sheet_name="Executive Narrative", index=False)
                exception_summary.to_excel(writer, sheet_name="Exceptions", index=False)
                data_quality_summary.to_excel(writer, sheet_name="Data Quality", index=False)
                dealer_rate_validation.to_excel(writer, sheet_name="Rate Validation", index=False)
                dealer_rate_exception_rows.to_excel(writer, sheet_name="Rate Exceptions", index=False)
                rate_df.to_excel(writer, sheet_name="Dealer Rates Used", index=False)
                outlier_perf_all.to_excel(writer, sheet_name="Outlier Performance", index=False)
                cross_rows.to_excel(writer, sheet_name="Cross Type Flags", index=False)
                df[df["Outlier"] == True].to_excel(writer, sheet_name="Outlier Rows", index=False)
                pd.DataFrame({"Year": list(cpi_table.keys()), "CPI": list(cpi_table.values())}).sort_values("Year").to_excel(writer, sheet_name="BLS CPI", index=False)
                fx_lookup.to_excel(writer, sheet_name="FX Rates", index=False)
                rebuild_reference.to_excel(writer, sheet_name="Rebuild Type Reference", index=False)
                region_reference.to_excel(writer, sheet_name="Region Reference", index=False)
                known_limitations.to_excel(writer, sheet_name="Known Limitations", index=False)
                data_dictionary.to_excel(writer, sheet_name="Data Dictionary", index=False)
                build_powerbi_table_dictionary().to_excel(writer, sheet_name="PBI Table Dictionary", index=False)
                build_export_mode_dictionary().to_excel(writer, sheet_name="Export Modes", index=False)
                build_required_files_checklist().to_excel(writer, sheet_name="Required Files", index=False)
                build_testing_checklist().to_excel(writer, sheet_name="Testing Checklist", index=False)
                build_update_process_table().to_excel(writer, sheet_name="Update Process", index=False)
                role_policy.to_excel(writer, sheet_name="Role Design", index=False)
                for machine in valid["SALES MODEL"].dropna().unique():
                    valid[valid["SALES MODEL"] == machine].to_excel(writer, sheet_name=safe_sheet_name(machine), index=False)
            apply_excel_brand_formatting(writer.book)
            apply_confidential_watermark(writer.book, scenario_name)
    if render_export_acknowledgement("full_export_ack"):
        safe_scenario = re.sub(r"[^A-Za-z0-9_-]+", "_", scenario_name.strip()) if scenario_name else "Scenario"
        if export_mode == "Power BI Dataset Export":
            download_label = "Download Power BI Dataset Workbook"
            export_prefix = "Rebuild_Analysis_PowerBI_Dataset"
            file_ext = "xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif export_mode == "Scenario Archive Package":
            download_label = "Download Scenario Archive Package"
            export_prefix = "Rebuild_Analysis_Scenario_Archive"
            file_ext = "zip"
            mime_type = "application/zip"
        else:
            download_label = "Download Workbook"
            export_prefix = "Rebuild_Analysis"
            file_ext = "xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        st.download_button(download_label, data=output.getvalue(), file_name=f"{export_prefix}_{safe_scenario}_{export_mode.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d')}.{file_ext}", mime=mime_type)
    else:
        st.info("Confirm export authorization to enable the full workbook download.")
    st.markdown("""<div class="cat-footer"><strong>Rebuild Analytics Platform</strong> | Internal analytics tool | Cost, inflation, dealer, region, outlier, cross-type exception, governance, performance-safeguard, and Power BI dataset reporting</div>""", unsafe_allow_html=True)
else:
    st.info("Upload file and run analysis")

